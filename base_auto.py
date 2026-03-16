import pandas as pd
import numpy as np
import os
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side

def tratar_e_consolidar_bases(caminho_base_rio, caminho_share, caminho_base_principal, datas_filtro=None):
    print("Iniciando leitura e tratamento dos dados...")

    # ==========================================
    # 1. EXTRAÇÃO
    # ==========================================
    df1 = pd.read_excel(caminho_base_rio, sheet_name='Planilha1')
    df1.rename(columns={'PASSAGEIRO': 'PAX'}, inplace=True)
    
    # NOVA REGRA: Filtrar a base para manter apenas onde o destino é Rio de Janeiro.
    # Usamos o strip() para arrancar espaços em branco perdidos e upper() para evitar problemas com letras minúsculas.
    df1 = df1[df1['DESTINO'].str.strip().str.upper() == 'RIO DE JANEIRO'].copy()
    
    if 'ORIGEM' not in df1.columns:
        df1['ORIGEM'] = 'SÃO PAULO'
        
    df2 = pd.read_excel(caminho_share, sheet_name='Base Relatorio   RIO x SAO')
    df2['DATA'] = df2['DATA SP']
    cols_df2 = ['Nº Mês', 'MÊS', 'EMPRESA', 'ORIGEM', 'DESTINO', 'SERVIÇO', 'HORÁRIO', 'PAX', 'DATA']
    df2 = df2[[c for c in cols_df2 if c in df2.columns]]

    df_consolidado = pd.concat([df1, df2], ignore_index=True)

    # ==========================================
    # --- NOVO: FILTRO DE DATAS ---
    # ==========================================
    if datas_filtro:
        # 1. Garante que o Pandas entenda a coluna como uma Data verdadeira
        df_consolidado['DATA'] = pd.to_datetime(df_consolidado['DATA'], errors='coerce')
        
        # 2. Converte a nossa lista de strings (ex: '2026-03-11') para o mesmo formato
        datas_limite = pd.to_datetime(datas_filtro)
        
        # 3. O ".isin()" funciona como um crachá de segurança: só passa quem está na lista
        df_consolidado = df_consolidado[df_consolidado['DATA'].isin(datas_limite)].copy()
        
        # 4. Se a tabela ficar vazia após o filtro, a gente para o processo para não dar erro lá na frente
        if df_consolidado.empty:
            print("⚠️ Nenhum dado encontrado para as datas selecionadas. Encerrando processamento.")
            return
    # ==========================================

    # ==========================================
    # 2. LIMPEZA E REGRAS DE NEGÓCIO
    # ==========================================
    
    # Transformando TUDO em texto para não perder a empresa "1001" (que é um número)
    df_consolidado['EMPRESA'] = df_consolidado['EMPRESA'].astype(str)

    mask_tag_sp = df_consolidado['EMPRESA'].str.contains(r'\(BF\)|\(BAF\)|\(RIO\)', na=False, regex=True)
    df_consolidado.loc[mask_tag_sp & (df_consolidado['DESTINO'].str.strip().str.upper() == 'SÃO PAULO'), 'DESTINO'] = 'SÃO PAULO (BARRA FUNDA)'
    df_consolidado.loc[mask_tag_sp & (df_consolidado['ORIGEM'].str.strip().str.upper() == 'SÃO PAULO'), 'ORIGEM'] = 'SÃO PAULO (BARRA FUNDA)'
    df_consolidado.loc[(df_consolidado['EMPRESA'].str.strip().str.upper() == 'ITAPEMIRIM'), 'EMPRESA'] = 'KAISSARA'

    df_consolidado['EMPRESA'] = df_consolidado['EMPRESA'].str.replace(r'\s*\(.*?\)\s*', '', regex=True)
    df_consolidado['EMPRESA'] = df_consolidado['EMPRESA'].str.normalize('NFKD').str.encode('ascii', errors='ignore').str.decode('utf-8').str.upper().str.strip()

    mapa_regras = {
        '1001': {'GRUPO': 'JCA', 'MODALIDADE': 'RODOVIARIO'},
        'AGUIA BRANCA': {'GRUPO': 'AGUIA BRANCA', 'MODALIDADE': 'RODOVIARIO'},
        'EXPRESSO DO SUL': {'GRUPO': 'JCA', 'MODALIDADE': 'RODOVIARIO'},
        'CATARINENSE': {'GRUPO': 'JCA', 'MODALIDADE': 'RODOVIARIO'},
        'PENHA': {'GRUPO': 'COMPORTE', 'MODALIDADE': 'RODOVIARIO'},
        'AGUIA FLEX': {'GRUPO': 'AGUIA BRANCA', 'MODALIDADE': 'DIGITAL'},
        'KAISSARA': {'GRUPO': 'SUZANTUR', 'MODALIDADE': 'RODOVIARIO'},
        'WEMOBI': {'GRUPO': 'JCA', 'MODALIDADE': 'DIGITAL'},
        'ADAMANTINA': {'GRUPO': 'ADAMANTINA', 'MODALIDADE': 'RODOVIARIO'},
        'FLIXBUS': {'GRUPO': 'FLIXBUS', 'MODALIDADE': 'DIGITAL'},
        'RIO DOCE': {'GRUPO': 'RIO DOCE', 'MODALIDADE': 'RODOVIARIO'},
        'NOTAVEL': {'GRUPO': 'NOTÁVEL', 'MODALIDADE': 'RODOVIARIO'}
    }
    
    df_consolidado['GRUPO'] = df_consolidado['EMPRESA'].map(lambda x: mapa_regras.get(x, {}).get('GRUPO', 'OUTROS'))
    df_consolidado['MODALIDADE'] = df_consolidado['EMPRESA'].map(lambda x: mapa_regras.get(x, {}).get('MODALIDADE', 'OUTROS'))

    # Trava de Segurança: Se houver algum 'OUTROS', ele para o código e te avisa.
    empresas_nao_mapeadas = df_consolidado[df_consolidado['GRUPO'] == 'OUTROS']['EMPRESA'].unique()
    if len(empresas_nao_mapeadas) > 0:
        raise ValueError(f"🚨 ERRO CRÍTICO: As seguintes empresas não estão no mapa: {empresas_nao_mapeadas}. Adicione no código antes de rodar novamente.")

    df_consolidado = df_consolidado[df_consolidado['PAX'] <= 69].copy()
    df_consolidado.loc[df_consolidado['PAX'] == 69, 'PAX'] = 68

    df_consolidado['IPV'] = np.where(df_consolidado['PAX'] > 54, 
                                     df_consolidado['PAX'] / 68, 
                                     df_consolidado['PAX'] / 54)
    df_consolidado['IPV'] = df_consolidado['IPV'].clip(upper=1.0)

    # Convertendo datas e horas para formatos puros do Python para o Excel entender depois
    df_consolidado['DATA'] = pd.to_datetime(df_consolidado['DATA'])
    df_consolidado['HORÁRIO'] = pd.to_datetime(df_consolidado['HORÁRIO'].astype(str), errors='coerce').apply(lambda x: x.time() if pd.notnull(x) else None)
    
    df_consolidado['Nº Mês'] = df_consolidado['DATA'].dt.month
    df_consolidado['Ano'] = df_consolidado['DATA'].dt.year
    df_consolidado['SEMANA'] = df_consolidado['DATA'].dt.isocalendar().week 
    
    meses_pt = {1:'01-JANEIRO', 2:'02-FEVEREIRO', 3:'03-MARÇO', 4:'04-ABRIL', 
                5:'05-MAIO', 6:'06-JUNHO', 7:'07-JULHO', 8:'08-AGOSTO', 
                9:'09-SETEMBRO', 10:'10-OUTUBRO', 11:'11-NOVEMBRO', 12:'12-DEZEMBRO'}
    df_consolidado['MÊS'] = df_consolidado['Nº Mês'].map(meses_pt)
    
    # Extraindo apenas a data pura (sem horas zerdas) para a coluna
    df_consolidado['DATA'] = df_consolidado['DATA'].dt.date

    # ==========================================
    # 3. VERIFICAÇÃO DE ANOMALIAS (Todas as pequenas)
    # ==========================================
    soma_diaria = df_consolidado.groupby(['DATA', 'EMPRESA'])['PAX'].sum().reset_index()
    datas_unicas = soma_diaria['DATA'].unique()
    gigantes_nomes = ['AGUIA BRANCA', '1001', 'EXPRESSO DO SUL']
    
    for data in datas_unicas:
        df_dia = soma_diaria[soma_diaria['DATA'] == data]
        max_gigantes = df_dia[df_dia['EMPRESA'].isin(gigantes_nomes)]['PAX'].max()
        
        if pd.isna(max_gigantes): 
            max_gigantes = 0
            
        # Pega todas as empresas que NÃO são gigantes e compara o volume
        outras_empresas_dia = df_dia[~df_dia['EMPRESA'].isin(gigantes_nomes)]
        for _, row in outras_empresas_dia.iterrows():
            if row['PAX'] > max_gigantes + 300:
                print(f"⚠️ ANOMALIA - {row['DATA'].strftime('%d/%m/%Y')}: {row['EMPRESA']} teve {row['PAX']} pax. (Teto das grandes: {max_gigantes}). Revisar fórmula!")

    # ==========================================
    # 4. EXPORTAÇÃO E FORMATAÇÃO VISUAL
    # ==========================================
    print("Iniciando formatação e salvamento no Excel...")
    ordem_colunas = ['DATA', 'Nº Mês', 'MÊS', 'EMPRESA', 'ORIGEM', 'DESTINO', 'SERVIÇO', 'HORÁRIO', 'PAX', 'SEMANA', 'IPV', 'GRUPO', 'MODALIDADE', 'Ano']
    
    if 'SERVIÇO' not in df_consolidado.columns:
        df_consolidado['SERVIÇO'] = 1
    else:
        df_consolidado['SERVIÇO'] = df_consolidado['SERVIÇO'].fillna(1)

    df_final = df_consolidado[ordem_colunas]

    book = load_workbook(caminho_base_principal)
    sheet = book['Base Relatorio   RIO x SAO']
    start_row = sheet.max_row + 1

    # Definindo estilos do OpenPyXL
    estilo_borda = Side(border_style="thin", color="000000")
    borda_fina = Border(top=estilo_borda, left=estilo_borda, right=estilo_borda, bottom=estilo_borda)
    alinhamento_centro = Alignment(horizontal='center', vertical='center')

    for row_idx, row_data in enumerate(df_final.values, start=start_row):
        for col_idx, value in enumerate(row_data, start=1):
            cell = sheet.cell(row=row_idx, column=col_idx, value=value)
            
            # 1. Centralizando todas as células
            cell.alignment = alinhamento_centro
            
            nome_coluna_atual = ordem_colunas[col_idx-1]
            
            # 2. Bordas em todas (menos no 'Ano')
            if nome_coluna_atual != 'Ano':
                cell.border = borda_fina
                
            # 3. Formatações Específicas
            if nome_coluna_atual == 'DATA':
                # 'dd/mmm' faz o excel mostrar "05/mar" (abreviado sem ano)
                cell.number_format = 'dd/mmm' 
            elif nome_coluna_atual == 'HORÁRIO':
                # 'hh:mm' exibe as horas e minutos, ocultando os segundos
                cell.number_format = 'hh:mm'
            elif nome_coluna_atual == 'IPV':
                # '0%' arredonda a visualização para números inteiros acompanhados do símbolo %
                cell.number_format = '0%'

    caminho_salvamento = os.path.join(os.path.dirname(caminho_base_principal), "TESTE_Base_RIO_x_SAO.xlsx")
    book.save(caminho_salvamento)
    print(f"✅ Sucesso! O arquivo formatado está em: {caminho_salvamento}")

# Execução (apenas ajuste o caminho da sua pasta de Downloads)
if __name__ == '__main__':
    pasta_testes = Path.home() / "Downloads"
    
    arq_base_rio = os.path.join(pasta_testes, "BASE RIO MARÇO 26.xlsx")
    arq_share = os.path.join(pasta_testes, "Share - Mercado RIO - SPO -2026.11.xlsx")
    arq_principal = os.path.join(pasta_testes, "Base RIO x SAO.xlsx")

    # --- ESTRATÉGIA DE DATAS ---
    # Opção A: Passando as datas na mão (Descomente para usar)
    minhas_datas = ['2026-03-05', '2026-03-06'] 

    # Opção B: Automático (Hoje e Ontem) - Aconselhado para automação diária
    # hoje = datetime.now().strftime('%Y-%m-%d')
    # ontem = (datetime.now() - timedelta(days=1)).strftime('%Y-%m-%d')
    # minhas_datas = [ontem, hoje]
    
    print(f"Filtrando dados apenas para as datas: {minhas_datas}")
    
    # Chamando a função com o novo parâmetro ativado
    tratar_e_consolidar_bases(arq_base_rio, arq_share, arq_principal, datas_filtro=minhas_datas)
