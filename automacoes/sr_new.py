import sys
import os
import base64
import json
# Feedback em ASCII puro para o dashboard
print("PROGRESS:{\"p\": 1, \"m\": \"Carregando Modulos SR...\"}", flush=True)

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import shutil
import re
import unicodedata
import polars as pl
from datetime import datetime, timedelta, date, time
from pathlib import Path
from core.banco import BASE_DIR, ASSETS_DIR
from core.google_auth import obter_servico_gmail
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Side


def normalizar_data_br(data_valor):
    """Normaliza datas para dd/mm/YYYY aceitando ISO e variações comuns."""
    if data_valor is None:
        return None

    if isinstance(data_valor, datetime):
        return data_valor.strftime('%d/%m/%Y')

    texto = str(data_valor).strip()
    if not texto:
        return None

    if 'T' in texto:
        try:
            return datetime.fromisoformat(texto.replace('Z', '')).strftime('%d/%m/%Y')
        except ValueError:
            pass

    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%Y/%m/%d'):
        try:
            return datetime.strptime(texto, fmt).strftime('%d/%m/%Y')
        except ValueError:
            continue

    raise ValueError(
        f"Data inválida: {data_valor}. Use dd/mm/YYYY, YYYY-mm-dd, dd-mm-YYYY ou ISO."
    )


def baixar_anexos_especificos(service, pasta_alvo, nomes_procurados, data_inicio=None, data_fim=None, callback_progresso=None):
    """
    Busca e baixa anexos que contenham as palavras-chave no nome, selecionando os mais recentes.
    data_inicio: 'YYYY/MM/DD'
    data_fim: 'YYYY/MM/DD'
    """
    if not os.path.exists(pasta_alvo):
        os.makedirs(pasta_alvo)

    if not data_inicio:
        data_inicio = datetime.now().strftime('%Y/%m/%d')
    
    query = f'has:attachment after:{data_inicio}'
    if data_fim:
        # Adiciona 1 dia para o 'before' ser inclusive (Gmail before é exclusivo)
        fim_dt = datetime.strptime(data_fim, '%Y/%m/%d') + timedelta(days=1)
        query += f' before:{fim_dt.strftime("%Y/%m/%d")}'
    
    print(f"DEBUG: Query Gmail: {query}")
    resultado = service.users().messages().list(userId='me', q=query).execute()
    mensagens = resultado.get('messages', [])

    # Dicionário para guardar o melhor candidato para cada termo: {termo: {date: X, filename: Y, id: Z, msgId: W}}
    candidatos = {termo: {"date": 0, "filename": None, "attachmentId": None, "messageId": None} for termo in nomes_procurados}

    def processar_partes_recursivo(parts, msg_id, msg_date):
        for parte in parts:
            nome_arquivo = parte.get('filename')
            if nome_arquivo:
                # Log de auditoria para o usuário
                print(f"DEBUG: Arquivo ignorado ou analisado: {nome_arquivo}")
                for termo in nomes_procurados:
                    if termo.lower() in nome_arquivo.lower():
                        if msg_date > candidatos[termo]["date"]:
                            if 'body' in parte and 'attachmentId' in parte['body']:
                                candidatos[termo] = {
                                    "date": msg_date,
                                    "filename": nome_arquivo,
                                    "attachmentId": parte['body']['attachmentId'],
                                    "messageId": msg_id
                                }
            # Se houver sub-partes, busca nelas também
            if 'parts' in parte:
                processar_partes_recursivo(parte['parts'], msg_id, msg_date)

    for m in mensagens:
        msg_completa = service.users().messages().get(userId='me', id=m['id']).execute()
        msg_date = int(msg_completa.get('internalDate', 0))
        payload = msg_completa.get('payload', {})
        
        # Inicia a busca recursiva a partir das partes do payload
        if 'parts' in payload:
            processar_partes_recursivo(payload['parts'], m['id'], msg_date)
        elif 'filename' in payload and payload['filename']:
            # Caso raro onde o payload principal é o próprio arquivo
            processar_partes_recursivo([payload], m['id'], msg_date)
    
    arquivos_baixados = {}
    total_candidatos = sum(1 for info in candidatos.values() if info["filename"])
    total_candidatos = max(total_candidatos, 1)
    contador_baixados = 0
    for termo, info in candidatos.items():
        if info["filename"]:
            try:
                anexo = service.users().messages().attachments().get(
                    userId='me', messageId=info["messageId"], id=info["attachmentId"]).execute()
                
                dados_decodificados = base64.urlsafe_b64decode(anexo['data'])
                caminho_final = os.path.join(pasta_alvo, info["filename"])
                
                with open(caminho_final, 'wb') as f:
                    f.write(dados_decodificados)
                
                print(f"[OK] Download realizado (Versao mais recente): {info['filename']}")
                arquivos_baixados[termo] = caminho_final
                contador_baixados += 1
                if callback_progresso:
                    progresso = 0.30 + (0.15 * (contador_baixados / total_candidatos))
                    callback_progresso(progresso, f"Anexo baixado ({contador_baixados}/{total_candidatos}): {Path(caminho_final).name}")
            except Exception as e:
                print(f"[ERRO] Falha ao baixar candidato {info['filename']}: {e}")

    return arquivos_baixados

# ==========================================
# TRATAMENTO DE DADOS
# ==========================================

MESES_PT = {
    1: '01-JANEIRO', 2: '02-FEVEREIRO', 3: '03-MARÇO', 4: '04-ABRIL',
    5: '05-MAIO', 6: '06-JUNHO', 7: '07-JULHO', 8: '08-AGOSTO',
    9: '09-SETEMBRO', 10: '10-OUTUBRO', 11: '11-NOVEMBRO', 12: '12-DEZEMBRO'
}


def _normalizar_data_excel(valor):
    if valor is None:
        return None
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor

    texto = str(valor).strip()
    if not texto or texto.lower() in {'none', 'nan', 'nat'}:
        return None

    if 'T' in texto:
        try:
            return datetime.fromisoformat(texto.replace('Z', '')).date()
        except ValueError:
            pass

    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%Y/%m/%d', '%Y-%m-%d %H:%M:%S'):
        try:
            return datetime.strptime(texto, fmt).date()
        except ValueError:
            continue
    return None


def _normalizar_horario(valor):
    if valor is None:
        return ''
    if isinstance(valor, datetime):
        return valor.strftime('%H:%M')
    if isinstance(valor, time):
        return valor.strftime('%H:%M')

    if isinstance(valor, (int, float)) and 0 <= float(valor) < 1:
        segundos = int(round(float(valor) * 24 * 60 * 60))
        horas = (segundos // 3600) % 24
        minutos = (segundos % 3600) // 60
        return f"{horas:02d}:{minutos:02d}"

    texto = str(valor).strip()
    if not texto or texto.lower() in {'none', 'nan', 'nat'}:
        return ''

    for fmt in ('%H:%M:%S', '%H:%M', '%H:%M:%S.%f', '%Y-%m-%d %H:%M:%S'):
        try:
            return datetime.strptime(texto, fmt).strftime('%H:%M')
        except ValueError:
            continue

    if re.match(r'^\d{2}:\d{2}', texto):
        return texto[:5]
    return texto


def _normalizar_texto_sem_acento(valor):
    texto = str(valor or '').strip().upper()
    return unicodedata.normalize('NFKD', texto).encode('ascii', errors='ignore').decode('utf-8').strip()


def _normalizar_cidade(valor):
    texto = str(valor or '').strip().upper()
    chave = _normalizar_texto_sem_acento(valor)

    if chave in {'RIO', 'RJ', 'RIO JANEIRO', 'RIO DE JANEIRO', 'RIO DE JANERIO'}:
        return 'RIO DE JANEIRO'
    if chave in {'SP', 'SAO PAULO'}:
        return 'SÃO PAULO'
    if chave == 'SAO PAULO (BARRA FUNDA)':
        return 'SÃO PAULO (BARRA FUNDA)'
    return texto


def _normalizar_empresa(valor):
    texto = re.sub(r'\s*\(.*?\)\s*', ' ', str(valor or ''))
    texto = unicodedata.normalize('NFKD', texto).encode('ascii', errors='ignore').decode('utf-8')
    return ' '.join(texto.upper().split())


def _ler_planilha_polars(caminho_arquivo, nome_aba):
    try:
        retorno = pl.read_excel(caminho_arquivo, sheet_name=nome_aba)
    except Exception as e:
        raise RuntimeError(
            f"Falha ao ler planilha '{nome_aba}' em {caminho_arquivo}. "
            f"Confirme suporte do polars.read_excel para o arquivo de origem."
        ) from e

    if isinstance(retorno, dict):
        if nome_aba in retorno:
            return retorno[nome_aba]
        return next(iter(retorno.values()))
    return retorno


def _garantir_colunas(df, colunas):
    faltantes = [c for c in colunas if c not in df.columns]
    if faltantes:
        df = df.with_columns([pl.lit(None).alias(c) for c in faltantes])
    return df.select(colunas)


def _gerar_chave_unica(row_dict):
    data_val = row_dict.get('DATA')
    if isinstance(data_val, datetime):
        data_txt = data_val.date().isoformat()
    elif isinstance(data_val, date):
        data_txt = data_val.isoformat()
    else:
        data_norm = _normalizar_data_excel(data_val)
        data_txt = data_norm.isoformat() if data_norm else ''

    horario_txt = _normalizar_horario(row_dict.get('HORÁRIO'))
    empresa = _normalizar_empresa(row_dict.get('EMPRESA'))
    origem = _normalizar_cidade(row_dict.get('ORIGEM'))
    destino = _normalizar_cidade(row_dict.get('DESTINO'))

    try:
        pax_txt = str(int(float(row_dict.get('PAX'))))
    except Exception:
        pax_txt = str(row_dict.get('PAX') or '').strip()

    try:
        ipv_txt = str(round(float(row_dict.get('IPV')), 4))
    except Exception:
        ipv_txt = str(row_dict.get('IPV') or '').strip()

    try:
        servico_txt = str(int(float(row_dict.get('SERVIÇO', 1))))
    except Exception:
        servico_txt = str(row_dict.get('SERVIÇO') or '1').strip()

    ano_txt = str(row_dict.get('Ano') or '').strip()

    return f"{data_txt}|{horario_txt}|{empresa}|{origem}|{destino}|{pax_txt}|{ipv_txt}|{servico_txt}|{ano_txt}"


def _normalizar_df_chaves(df):
    return df.with_columns([
        pl.col('DATA').map_elements(_normalizar_data_excel, return_dtype=pl.Date).alias('DATA'),
        pl.col('HORÁRIO').map_elements(_normalizar_horario, return_dtype=pl.Utf8).fill_null('').alias('HORÁRIO'),
        pl.col('EMPRESA').map_elements(_normalizar_empresa, return_dtype=pl.Utf8).fill_null('').alias('EMPRESA'),
        pl.col('ORIGEM').map_elements(_normalizar_cidade, return_dtype=pl.Utf8).fill_null('').alias('ORIGEM'),
        pl.col('DESTINO').map_elements(_normalizar_cidade, return_dtype=pl.Utf8).fill_null('').alias('DESTINO'),
        pl.col('SERVIÇO').cast(pl.Float64, strict=False).fill_null(1).round(0).cast(pl.Int64).alias('SERVIÇO'),
        pl.col('PAX').cast(pl.Float64, strict=False).round(0).cast(pl.Int64).alias('PAX'),
        pl.col('IPV').cast(pl.Float64, strict=False).alias('IPV'),
        pl.col('Ano').cast(pl.Int64, strict=False).alias('Ano'),
    ])


def _expr_chave():
    return pl.concat_str([
        pl.col('DATA').dt.strftime('%Y-%m-%d').fill_null(''),
        pl.col('HORÁRIO').fill_null(''),
        pl.col('EMPRESA').fill_null(''),
        pl.col('ORIGEM').fill_null(''),
        pl.col('DESTINO').fill_null(''),
        pl.col('PAX').cast(pl.Utf8).fill_null(''),
        pl.col('IPV').round(4).cast(pl.Utf8).fill_null(''),
        pl.col('SERVIÇO').cast(pl.Utf8).fill_null('1'),
        pl.col('Ano').cast(pl.Utf8).fill_null(''),
    ], separator='|').alias('KEY')


def _aplicar_estilo_openpyxl(cell, col_idx, ordem_colunas):
    estilo_borda = Side(border_style="thin", color="000000")
    borda_fina = Border(top=estilo_borda, left=estilo_borda, right=estilo_borda, bottom=estilo_borda)
    alinhamento_centro = Alignment(horizontal='center', vertical='center')

    nome_coluna = ordem_colunas[col_idx - 1]
    cell.alignment = alinhamento_centro
    if nome_coluna != 'Ano':
        cell.border = borda_fina
    if nome_coluna == 'DATA':
        cell.number_format = 'dd/mmm'
    elif nome_coluna == 'HORÁRIO':
        cell.number_format = 'hh:mm'
    elif nome_coluna == 'IPV':
        cell.number_format = '0%'


def _converter_valor_planilha(nome_coluna, valor):
    if nome_coluna == 'DATA':
        return _normalizar_data_excel(valor)
    if nome_coluna == 'HORÁRIO':
        hora_txt = _normalizar_horario(valor)
        if not hora_txt:
            return None
        try:
            h, m = hora_txt.split(':')
            return time(int(h), int(m))
        except Exception:
            return None
    if nome_coluna == 'IPV':
        try:
            return float(valor)
        except Exception:
            return 0.0
    if nome_coluna in {'Nº Mês', 'PAX', 'SEMANA', 'SERVIÇO', 'Ano'}:
        try:
            return int(float(valor))
        except Exception:
            return None
    if valor is None:
        return ''
    return str(valor)


def _append_openpyxl(caminho_arquivo, nome_aba, df_saida, ordem_colunas):
    book = load_workbook(caminho_arquivo)
    if nome_aba in book.sheetnames:
        sheet = book[nome_aba]
    else:
        sheet = book.active
        sheet.title = nome_aba
        sheet.append(ordem_colunas)

    start_row = sheet.max_row + 1
    modelo_row = 2 if sheet.max_row >= 2 else None
    estilos_coluna = None
    if modelo_row:
        estilos_coluna = [sheet.cell(row=modelo_row, column=i)._style for i in range(1, len(ordem_colunas) + 1)]

    for row_idx, row_data in enumerate(df_saida.iter_rows(named=True), start=start_row):
        for col_idx, col_name in enumerate(ordem_colunas, start=1):
            valor = _converter_valor_planilha(col_name, row_data.get(col_name))
            cell = sheet.cell(row=row_idx, column=col_idx, value=valor)
            if estilos_coluna:
                # Reusa exatamente o estilo de uma linha modelo para manter o design original.
                cell._style = estilos_coluna[col_idx - 1]
            else:
                _aplicar_estilo_openpyxl(cell, col_idx, ordem_colunas)

    book.save(caminho_arquivo)


def tratar_e_consolidar_bases(caminho_base_rio=None, caminho_share=None, caminho_base_principal=None, datas_filtro=None, callback_progresso=None):
    if callback_progresso:
        callback_progresso(0.60, "Iniciando leitura dos pacotes locais e preparação de consolidação polars...")
    print("Iniciando leitura e tratamento dos dados...")

    ordem_colunas = ['DATA', 'Nº Mês', 'MÊS', 'EMPRESA', 'ORIGEM', 'DESTINO', 'SERVIÇO', 'HORÁRIO', 'PAX', 'SEMANA', 'IPV', 'GRUPO', 'MODALIDADE', 'Ano']
    dfs = []

    # 1. EXTRAÇÃO
    if caminho_base_rio and os.path.exists(caminho_base_rio):
        df1 = _ler_planilha_polars(caminho_base_rio, 'Planilha1')
        if 'PASSAGEIRO' in df1.columns and 'PAX' not in df1.columns:
            df1 = df1.rename({'PASSAGEIRO': 'PAX'})

        if 'ORIGEM' not in df1.columns:
            df1 = df1.with_columns(pl.lit('SÃO PAULO').alias('ORIGEM'))

        df1 = _garantir_colunas(df1, ordem_colunas)
        df1 = df1.with_columns(pl.lit('BASE_RIO').alias('_FONTE_SR'))
        dfs.append(df1)
        print(f"[DATA] Base RIO carregada: {len(df1)} linhas.")

    if caminho_share and os.path.exists(caminho_share):
        df2 = _ler_planilha_polars(caminho_share, 'Base Relatorio   RIO x SAO')
        if 'DATA SP' in df2.columns:
            if 'DATA' in df2.columns:
                df2 = df2.with_columns(pl.coalesce([pl.col('DATA'), pl.col('DATA SP')]).alias('DATA'))
            else:
                df2 = df2.with_columns(pl.col('DATA SP').alias('DATA'))

        df2 = _garantir_colunas(df2, ordem_colunas)
        df2 = df2.with_columns(pl.lit('SHARE').alias('_FONTE_SR'))
        dfs.append(df2)
        print(f"[DATA] Base Share carregada: {len(df2)} linhas.")

    if not dfs:
        if callback_progresso: callback_progresso(0.65, "[Aviso] Nenhum dado localizado nos arquivos de entrada.")
        print("[AVISO] Nenhum dado encontrado para as bases especificadas.")
        return

    df_consolidado = pl.concat(dfs, how='diagonal_relaxed')
    if callback_progresso:
        callback_progresso(0.65, f"Arquivos unificados com sucesso. Total bruto: {len(df_consolidado)} registros. Aplicando regras de negócio e validações...")

    # FILTRO DE DATAS
    df_consolidado = df_consolidado.with_columns(
        pl.col('DATA').map_elements(_normalizar_data_excel, return_dtype=pl.Date).alias('DATA')
    )

    if datas_filtro:
        datas_limite = [d for d in (_normalizar_data_excel(x) for x in datas_filtro) if d is not None]
        df_consolidado = df_consolidado.filter(pl.col('DATA').is_in(datas_limite))

        if len(df_consolidado) == 0:
            print("[AVISO] Nenhum dado encontrado para as datas especificadas.")
            return

    # 2. LIMPEZA E REGRAS DE NEGÓCIO
    df_consolidado = df_consolidado.with_columns([
        pl.col('EMPRESA').cast(pl.Utf8, strict=False).fill_null('').alias('EMPRESA_BRUTA'),
        pl.col('DESTINO').map_elements(_normalizar_cidade, return_dtype=pl.Utf8).alias('DESTINO'),
        pl.col('ORIGEM').map_elements(_normalizar_cidade, return_dtype=pl.Utf8).alias('ORIGEM'),
        pl.col('HORÁRIO').map_elements(_normalizar_horario, return_dtype=pl.Utf8).alias('HORÁRIO'),
        pl.col('PAX').cast(pl.Float64, strict=False).alias('PAX'),
        pl.col('SERVIÇO').cast(pl.Float64, strict=False).fill_null(1).round(0).cast(pl.Int64).alias('SERVIÇO'),
    ])

    mask_catarinense = pl.col('EMPRESA_BRUTA').str.to_uppercase().str.contains('CATARINENSE')
    mask_bf = pl.col('EMPRESA_BRUTA').str.to_uppercase().str.contains(r'\(BF\)|\(BAF\)')

    for col in ['DESTINO', 'ORIGEM']:
        df_consolidado = df_consolidado.with_columns(
            pl.when(mask_bf & (pl.col(col) == 'SÃO PAULO')).then(pl.lit('SÃO PAULO (BARRA FUNDA)')).otherwise(pl.col(col)).alias(col)
        )
        df_consolidado = df_consolidado.with_columns(
            pl.when(mask_catarinense & ~mask_bf & (pl.col(col) == 'SÃO PAULO (BARRA FUNDA)')).then(pl.lit('SÃO PAULO')).otherwise(pl.col(col)).alias(col)
        )

    cidades_sp = ['SÃO PAULO', 'SÃO PAULO (BARRA FUNDA)']
    df_consolidado = df_consolidado.filter(
        ((pl.col('ORIGEM').is_in(cidades_sp)) & (pl.col('DESTINO') == 'RIO DE JANEIRO')) |
        ((pl.col('ORIGEM') == 'RIO DE JANEIRO') & (pl.col('DESTINO').is_in(cidades_sp)))
    )

    df_consolidado = df_consolidado.with_columns(
        pl.col('EMPRESA_BRUTA').map_elements(_normalizar_empresa, return_dtype=pl.Utf8).alias('EMPRESA')
    )

    df_consolidado = df_consolidado.with_columns(
        pl.when(pl.col('EMPRESA') == 'ITAPEMIRIM').then(pl.lit('KAISSARA')).otherwise(pl.col('EMPRESA')).alias('EMPRESA')
    )

    if 'EMPRESA_BRUTA' in df_consolidado.columns:
        df_consolidado = df_consolidado.drop(['EMPRESA_BRUTA'])

    mapa_regras = {
        '1001': {'GRUPO': 'JCA', 'MODALIDADE': 'RODOVIARIO'},
        'AGUIA BRANCA': {'GRUPO': 'AGUIA BRANCA', 'MODALIDADE': 'RODOVIARIO'},
        'EXPRESSO DO SUL': {'GRUPO': 'JCA', 'MODALIDADE': 'RODOVIARIO'},
        'CATARINENSE': {'GRUPO': 'JCA', 'MODALIDADE': 'RODOVIARIO'},
        'PENHA': {'GRUPO': 'COMPORTE', 'MODALIDADE': 'RODOVIARIO'},
        'AGUIA FLEX': {'GRUPO': 'AGUIA BRANCA', 'MODALIDADE': 'DIGITAL'},
        'KAISSARA': {'GRUPO': 'SUZANTUR', 'MODALIDADE': 'RODOVIARIO'},
        'WEMOBI': {'GRUPO': 'JCA', 'MODALIDADE': 'DIGITAL'},
        'ADAMANTINA': {'GRUPO': 'ADAMANTINA', 'MODALIDADE': 'RODOVIARIO'},
        'FLIXBUS': {'GRUPO': 'FLIXBUS', 'MODALIDADE': 'DIGITAL'},
        'RIO DOCE': {'GRUPO': 'RIO DOCE', 'MODALIDADE': 'RODOVIARIO'},
        'NOTAVEL': {'GRUPO': 'NOTÁVEL', 'MODALIDADE': 'RODOVIARIO'}
    }

    df_consolidado = df_consolidado.with_columns([
        pl.col('EMPRESA').map_elements(lambda x: mapa_regras.get(x, {}).get('GRUPO', 'OUTROS'), return_dtype=pl.Utf8).alias('GRUPO'),
        pl.col('EMPRESA').map_elements(lambda x: mapa_regras.get(x, {}).get('MODALIDADE', 'OUTROS'), return_dtype=pl.Utf8).alias('MODALIDADE')
    ])

    empresas_nao_mapeadas = df_consolidado.filter(pl.col('GRUPO') == 'OUTROS').select('EMPRESA').unique().to_series().to_list()
    if len(empresas_nao_mapeadas) > 0:
        raise ValueError(f"[ERRO DE MAPEAMENTO] As seguintes empresas nao estao no sistema: {list(empresas_nao_mapeadas)}. Favor contatar o administrador para atualizar o mapa de regras.")

    df_consolidado = df_consolidado.filter(pl.col('PAX').is_not_null() & (pl.col('PAX') <= 69))
    df_consolidado = df_consolidado.with_columns(
        pl.when(pl.col('PAX') == 69).then(pl.lit(68)).otherwise(pl.col('PAX')).alias('PAX')
    )
    df_consolidado = df_consolidado.with_columns([
        pl.when(pl.col('PAX') > 54).then(pl.col('PAX') / 68).otherwise(pl.col('PAX') / 54).alias('IPV')
    ])
    df_consolidado = df_consolidado.with_columns([
        pl.when(pl.col('IPV') > 1).then(pl.lit(1.0)).otherwise(pl.col('IPV')).alias('IPV'),
        pl.col('PAX').round(0).cast(pl.Int64).alias('PAX')
    ])

    df_consolidado = df_consolidado.filter(pl.col('DATA').is_not_null())
    df_consolidado = df_consolidado.with_columns([
        pl.col('DATA').dt.month().alias('Nº Mês'),
        pl.col('DATA').dt.year().alias('Ano'),
        pl.col('DATA').dt.week().alias('SEMANA'),
        pl.col('Nº Mês').map_elements(lambda x: MESES_PT.get(int(x)) if x is not None else None, return_dtype=pl.Utf8).alias('MÊS')
    ])

    # 4. EVITA DUPLICATAS E ORDENA NOVAS LINHAS PARA ALINHAR DIAS ENTRE DOCUMENTOS
    if callback_progresso:
        callback_progresso(0.75, "Sanitização de caracteres completada. Verificando dados históricos e identificando eventuais duplicatas de viagem...")
    print("Verificando duplicatas e preparando salvamento...")

    df_novos = _garantir_colunas(df_consolidado, ordem_colunas + ['_FONTE_SR'])
    df_novos = _normalizar_df_chaves(df_novos).with_columns(_expr_chave())

    if os.path.exists(caminho_base_principal):
        try:
            df_existente = _ler_planilha_polars(caminho_base_principal, 'Base Relatorio   RIO x SAO')
            df_existente = _garantir_colunas(df_existente, ordem_colunas)
            df_existente = _normalizar_df_chaves(df_existente)
            df_existente = df_existente.filter(pl.col('DATA').is_not_null())
            df_existente = df_existente.with_columns([
                pl.col('DATA').dt.month().alias('Nº Mês'),
                pl.col('DATA').dt.year().alias('Ano'),
                pl.col('DATA').dt.week().alias('SEMANA'),
                pl.col('Nº Mês').map_elements(lambda x: MESES_PT.get(int(x)) if x is not None else None, return_dtype=pl.Utf8).alias('MÊS')
            ])

            df_existente_keys = df_existente.with_columns(_expr_chave()).select('KEY').unique()
            df_novos = df_novos.join(df_existente_keys, on='KEY', how='anti')

            if len(df_novos) == 0:
                if callback_progresso:
                    callback_progresso(0.85, "Nenhum dado novo encontrado. Nada para inserir na base histórica.")
                print("[INFO] Nenhum dado novo encontrado. Nada para inserir na base principal.")
                return
            else:
                if callback_progresso:
                    callback_progresso(0.85, f"Sucesso! Mapeadas {len(df_novos)} novas viagens de ônibus inéditas para inclusão!")
                print(f"[INFO] Adicionando {len(df_novos)} novas linhas inéditas...")

        except Exception as e:
            print(f"[AVISO] Não foi possível verificar duplicatas de forma detalhada: {e}. Prosseguindo com precaução.")

    df_novos = df_novos.with_columns([
        pl.col('DATA').dt.month().alias('Nº Mês'),
        pl.col('DATA').dt.year().alias('Ano'),
        pl.col('DATA').dt.week().alias('SEMANA'),
        pl.col('Nº Mês').map_elements(lambda x: MESES_PT.get(int(x)) if x is not None else None, return_dtype=pl.Utf8).alias('MÊS'),
        pl.when(pl.col('_FONTE_SR') == 'BASE_RIO').then(pl.lit(0)).when(pl.col('_FONTE_SR') == 'SHARE').then(pl.lit(1)).otherwise(pl.lit(9)).alias('ORDEM_FONTE'),
        pl.col('HORÁRIO').fill_null('').alias('HORARIO_SORT')
    ])
    df_novos = df_novos.sort(['DATA', 'ORDEM_FONTE', 'HORARIO_SORT', 'EMPRESA', 'ORIGEM', 'DESTINO', 'SERVIÇO'])
    df_novos = df_novos.drop(['ORDEM_FONTE', 'HORARIO_SORT'])
    if '_FONTE_SR' in df_novos.columns:
        df_novos = df_novos.drop(['_FONTE_SR'])
    if 'KEY' in df_novos.columns:
        df_novos = df_novos.drop(['KEY'])
    df_novos = _garantir_colunas(df_novos, ordem_colunas)

    if len(df_novos) == 0:
        if callback_progresso:
            callback_progresso(0.85, "Nenhum dado novo encontrado. Nada para inserir na base histórica.")
        print("[INFO] Nenhum dado novo encontrado. Nada para inserir na base principal.")
        return

    if not os.path.exists(caminho_base_principal):
        raise FileNotFoundError(f"Base principal não encontrada para escrita: {caminho_base_principal}")

    if callback_progresso: callback_progresso(0.92, "Descarregando células limpas no documento Excel (Aplicando Borders e Formatadores de Data)...")
    _append_openpyxl(caminho_base_principal, 'Base Relatorio   RIO x SAO', df_novos, ordem_colunas)
    print(f"[OK] Sucesso! O arquivo foi atualizado com novos dados em: {caminho_base_principal}")

# ==========================================
# INTEGRAÇÃO COM A UI
# ==========================================

def executar_sr(
    id_usuario,
    e_ini,
    e_fim,
    b_ini,
    b_fim,
    callback_progresso=None,
    hook_cancelamento=None,
    modo_execucao="completo",
    pasta_destino=None,
    arquivo_entrada=None,
    base_automacao=None,
):
    """Executa SR com suporte a modos modulares sem quebrar compatibilidade."""
    if callback_progresso:
        callback_progresso(0.01, f"Robô SR Acionado ({modo_execucao}): Construindo trilhas lógicas e mapeando diretórios...")

    modos_validos = {
        "completo",
        "download",
        "download_tratamento",
        "tratamento",
        "tratamento_envio",
        "arquivo_tratamento",
        "arquivo_envio",
        "arquivo_tratamento_envio",
    }
    if modo_execucao not in modos_validos:
        raise ValueError("Modo selecionado não é suportado para esta automação SR.")

    e_ini = normalizar_data_br(e_ini)
    e_fim = normalizar_data_br(e_fim)
    b_ini = normalizar_data_br(b_ini)
    b_fim = normalizar_data_br(b_fim)

    try:
        pasta_downloads = Path.home() / "Downloads"
        destino_envio = Path(pasta_destino) if pasta_destino else None

        def resolver_base_referencia():
            """Resolve o arquivo base de referência de forma robusta para os modos do frontend."""
            base_raw = (base_automacao or "").strip() if isinstance(base_automacao, str) else base_automacao

            if isinstance(base_raw, str):
                base_raw_lower = base_raw.lower()

                # Quando o frontend envia 'personalizada', o caminho costuma chegar em pasta_personalizada.
                # Em modos de tratamento por arquivo, arquivo_entrada representa o arquivo a tratar,
                # então não deve ser reutilizado como base de referência.
                if base_raw_lower == "personalizada":
                    if arquivo_entrada and modo_execucao not in {"tratamento", "tratamento_envio", "arquivo_tratamento", "arquivo_tratamento_envio"}:
                        base_raw = arquivo_entrada
                    else:
                        base_raw = ""
                elif base_raw_lower in {"padrao", "sem_base", "none", "null"}:
                    base_raw = ""

            # Modo padrão/sem base no frontend: usa Base RIO x SAO.xlsx da pasta Downloads.
            if not base_raw:
                return pasta_downloads / "Base RIO x SAO.xlsx"

            caminho_base = Path(base_raw)
            if caminho_base.suffix.lower() in {".xlsx", ".xls", ".xlsm"}:
                caminho_base.parent.mkdir(parents=True, exist_ok=True)
                return caminho_base

            caminho_base.mkdir(parents=True, exist_ok=True)
            return caminho_base / "Base RIO x SAO.xlsx"

        base_referencia_arquivo = resolver_base_referencia()

        pasta_saida_final = destino_envio or base_referencia_arquivo.parent
        pasta_saida_final.mkdir(parents=True, exist_ok=True)
        arquivo_saida_principal = pasta_saida_final / base_referencia_arquivo.name

        def montar_lista_datas_base():
            d_ini = datetime.strptime(b_ini, "%d/%m/%Y")
            d_fim = datetime.strptime(b_fim, "%d/%m/%Y")
            delta = d_fim - d_ini
            return [(d_ini + timedelta(days=i)).strftime('%Y-%m-%d') for i in range(delta.days + 1)]

        def criar_base_vazia(caminho_arquivo: Path):
            """Cria uma base mínima do SR quando o arquivo de referência ainda não existe."""
            caminho_arquivo.parent.mkdir(parents=True, exist_ok=True)

            ordem_colunas = [
                'DATA', 'Nº Mês', 'MÊS', 'EMPRESA', 'ORIGEM', 'DESTINO', 'SERVIÇO',
                'HORÁRIO', 'PAX', 'SEMANA', 'IPV', 'GRUPO', 'MODALIDADE', 'Ano'
            ]

            wb = Workbook()
            ws = wb.active
            ws.title = 'Base Relatorio   RIO x SAO'
            ws.append(ordem_colunas)
            wb.save(str(caminho_arquivo))
            print(f"[INFO] Base SR criada automaticamente em: {caminho_arquivo}")

        def preparar_base_saida():
            if arquivo_saida_principal.exists():
                return
            if base_referencia_arquivo.exists() and base_referencia_arquivo.resolve() != arquivo_saida_principal.resolve():
                shutil.copy2(str(base_referencia_arquivo), str(arquivo_saida_principal))
                return
            criar_base_vazia(arquivo_saida_principal)

        def localizar_arquivos_locais_tratamento():
            """Localiza os anexos SR já baixados sem depender de nova busca no e-mail."""
            termos = ["BASE RIO", "Share - Mercado RIO"]
            extensoes_validas = {".xlsx", ".xls", ".xlsm"}

            if arquivo_entrada:
                caminho_entrada = Path(arquivo_entrada)
                if caminho_entrada.exists() and caminho_entrada.is_file():
                    if caminho_entrada.name.startswith("~$"):
                        raise ValueError(f"Arquivo temporário do Excel não é válido para tratamento: {caminho_entrada.name}")

                    nome = caminho_entrada.name.lower()
                    encontrados_diretos = {}

                    if "base rio" in nome:
                        encontrados_diretos["BASE RIO"] = str(caminho_entrada)
                    if "share - mercado rio" in nome or "share mercado rio" in nome:
                        encontrados_diretos["Share - Mercado RIO"] = str(caminho_entrada)

                    # Fallback: se o arquivo manual não tiver nome esperado, trata como BASE RIO.
                    if not encontrados_diretos:
                        encontrados_diretos["BASE RIO"] = str(caminho_entrada)

                    print(f"[INFO] Arquivo manual informado para tratamento SR: {caminho_entrada}")
                    return encontrados_diretos

            pastas_busca = []
            if arquivo_entrada:
                caminho_entrada = Path(arquivo_entrada)
                if caminho_entrada.exists() and caminho_entrada.is_dir():
                    pastas_busca.append(caminho_entrada)
            if destino_envio:
                pastas_busca.append(destino_envio)
            pastas_busca.append(pasta_downloads)

            pastas_unicas = []
            for pasta in pastas_busca:
                if pasta and pasta.exists() and pasta not in pastas_unicas:
                    pastas_unicas.append(pasta)

            encontrados = {}
            for pasta in pastas_unicas:
                arquivos_validos = [
                    f for f in pasta.iterdir()
                    if f.is_file() and f.suffix.lower() in extensoes_validas and not f.name.startswith("~$")
                ]

                for termo in termos:
                    if termo in encontrados:
                        continue

                    candidatos = [f for f in arquivos_validos if termo.lower() in f.name.lower()]
                    if termo == "BASE RIO":
                        candidatos = [f for f in candidatos if "x sao" not in f.name.lower() and "rio x sao" not in f.name.lower()]
                    if not candidatos:
                        continue

                    candidatos.sort(key=lambda x: x.stat().st_mtime, reverse=True)
                    encontrados[termo] = str(candidatos[0])

                if len(encontrados) == len(termos):
                    break

            if encontrados:
                print(f"[INFO] Arquivos locais selecionados para tratamento SR: {encontrados}")

            return encontrados

        if modo_execucao == "arquivo_envio":
            if not arquivo_entrada:
                raise ValueError("Selecione um arquivo para envio.")
            origem_arquivo = Path(arquivo_entrada)
            if not origem_arquivo.exists():
                raise FileNotFoundError(f"Arquivo não encontrado: {origem_arquivo}")
            pasta_final = destino_envio or pasta_downloads
            pasta_final.mkdir(parents=True, exist_ok=True)
            destino_final = pasta_final / origem_arquivo.name
            if origem_arquivo.resolve() != destino_final.resolve():
                shutil.copy2(str(origem_arquivo), str(destino_final))
            if callback_progresso:
                callback_progresso(1.0, f"Arquivo enviado: {destino_final.name}")
            return {
                "arquivo_principal": str(destino_final),
                "arquivos_saida": [str(destino_final)],
                "pasta_final": str(pasta_final),
                "mensagem": "Arquivo enviado com sucesso.",
            }

        if modo_execucao in {"arquivo_tratamento", "arquivo_tratamento_envio"}:
            if not arquivo_entrada:
                raise ValueError("Selecione um arquivo já baixado para realizar o tratamento no SR.")

            caminho_arquivo = Path(arquivo_entrada)
            if not caminho_arquivo.exists():
                raise FileNotFoundError(f"Arquivo não encontrado: {caminho_arquivo}")

            if callback_progresso:
                callback_progresso(0.55, f"Iniciando tratamento da base informada: {caminho_arquivo.name}")

            preparar_base_saida()
            tratar_e_consolidar_bases(
                caminho_base_rio=str(caminho_arquivo),
                caminho_share=None,
                caminho_base_principal=str(arquivo_saida_principal),
                datas_filtro=montar_lista_datas_base(),
                callback_progresso=callback_progresso
            )

            if callback_progresso:
                callback_progresso(1.0, f"Módulo SR Finalizado: Estrutura lógica exportada para a pasta padrão sem pendências.")

            return {
                "arquivo_principal": str(arquivo_saida_principal),
                "arquivos_saida": [str(arquivo_saida_principal)],
                "pasta_final": str(pasta_saida_final),
                "mensagem": "Tratamento SR concluído com sucesso.",
            }

        if modo_execucao in {"tratamento", "tratamento_envio"}:
            if callback_progresso:
                callback_progresso(0.45, "Localizando arquivos já baixados para tratamento...")

            arquivos_locais = localizar_arquivos_locais_tratamento()
            if len(arquivos_locais) == 0:
                raise ValueError("Nenhum arquivo local foi encontrado para tratamento SR. Verifique os nomes BASE RIO e Share - Mercado RIO na pasta Downloads.")

            print(f"[INFO] Termos encontrados para tratamento SR (modo local): {list(arquivos_locais.keys())}")

            if callback_progresso:
                termos_localizados = []
                for termo, caminho in arquivos_locais.items():
                    termos_localizados.append(f"{termo}: {Path(caminho).name}")
                if termos_localizados:
                    callback_progresso(0.5, f"Arquivos locais encontrados: {' | '.join(termos_localizados)}")

            if callback_progresso:
                callback_progresso(0.55, "Iniciando tratamento dos arquivos locais...")

            preparar_base_saida()
            tratar_e_consolidar_bases(
                caminho_base_rio=arquivos_locais.get("BASE RIO"),
                caminho_share=arquivos_locais.get("Share - Mercado RIO"),
                caminho_base_principal=str(arquivo_saida_principal),
                datas_filtro=montar_lista_datas_base(),
                callback_progresso=callback_progresso
            )

            if callback_progresso:
                callback_progresso(1.0, f"Módulo SR Finalizado: Os arquivos locais foram limpos, unificados e sincronizados na base RIO X SP.")

            return {
                "arquivo_principal": str(arquivo_saida_principal),
                "arquivos_saida": [str(arquivo_saida_principal)],
                "pasta_final": str(pasta_saida_final),
                "mensagem": f"Tratamento SR concluído com sucesso ({len(arquivos_locais)} arquivo(s) de entrada).",
            }

        if callback_progresso:
            callback_progresso(0.1, "Conectando ao Gmail...")
        service, _ = obter_servico_gmail()

        if hook_cancelamento and hook_cancelamento():
            return {
                "arquivo_principal": None,
                "arquivos_saida": [],
                "pasta_final": None,
                "mensagem": "Processo cancelado pelo usuário.",
            }

        termos_esperados = ["BASE RIO", "Share - Mercado RIO"]

        data_busca_ini = datetime.strptime(e_ini, "%d/%m/%Y").strftime("%Y/%m/%d")
        data_busca_fim = datetime.strptime(e_fim, "%d/%m/%Y").strftime("%Y/%m/%d")

        if callback_progresso:
            callback_progresso(0.3, f"Buscando anexos entre {e_ini} e {e_fim}...")
        arquivos = baixar_anexos_especificos(
            service,
            pasta_downloads,
            termos_esperados,
            data_inicio=data_busca_ini,
            data_fim=data_busca_fim,
            callback_progresso=callback_progresso,
        )

        if hook_cancelamento and hook_cancelamento():
            return {
                "arquivo_principal": None,
                "arquivos_saida": [],
                "pasta_final": None,
                "mensagem": "Processo cancelado pelo usuário.",
            }

        if len(arquivos) == 0:
            raise Exception(f"Nenhum dos arquivos esperados foi encontrado no e-mail: {termos_esperados}")

        print(f"[INFO] Termos encontrados para tratamento SR: {list(arquivos.keys())}")

        if callback_progresso:
            termos_email = []
            for termo, caminho in arquivos.items():
                termos_email.append(f"{termo}: {Path(caminho).name}")
            if termos_email:
                callback_progresso(0.48, f"Arquivos encontrados no e-mail: {' | '.join(termos_email)}")

        arquivos_baixados = [str(v) for v in arquivos.values() if v]

        if modo_execucao == "download":
            if destino_envio:
                destino_envio.mkdir(parents=True, exist_ok=True)
                arquivos_copiados = []
                for item in arquivos_baixados:
                    origem_item = Path(item)
                    destino_item = destino_envio / origem_item.name
                    shutil.copy2(str(origem_item), str(destino_item))
                    arquivos_copiados.append(str(destino_item))
                arquivos_baixados = arquivos_copiados

            if callback_progresso:
                callback_progresso(1.0, "Download de anexos concluído com sucesso!")
            return {
                "arquivo_principal": arquivos_baixados[0] if arquivos_baixados else None,
                "arquivos_saida": arquivos_baixados,
                "pasta_final": str(destino_envio or pasta_downloads),
                "mensagem": "Download de anexos concluído com sucesso.",
            }

        if callback_progresso:
            callback_progresso(0.6, "Iniciando processamento interno de normalização (Lendo anexos em memória)...")

        preparar_base_saida()
        tratar_e_consolidar_bases(
            caminho_base_rio=arquivos.get("BASE RIO"),
            caminho_share=arquivos.get("Share - Mercado RIO"),
            caminho_base_principal=str(arquivo_saida_principal),
            datas_filtro=montar_lista_datas_base(),
            callback_progresso=callback_progresso
        )

        if callback_progresso:
            callback_progresso(1.0, "Fim da Operação 100%: Pipeline Gmail -> Tratamento -> Base Histórica transacionado!")

        return {
            "arquivo_principal": str(arquivo_saida_principal),
            "arquivos_saida": [str(arquivo_saida_principal)],
            "pasta_final": str(pasta_saida_final),
            "mensagem": "Processo SR concluído com sucesso.",
        }

    except Exception as e:
        raise Exception(f"Erro SR: {str(e)}")

# ==========================================
# EXECUÇÃO MANUAL (TESTE)
# ==========================================

# ==========================================
# EXECUÇÃO VIA CLI/BACKEND
# ==========================================
if __name__ == '__main__':
    # Ler parâmetros de CLI (Base64) ou STDIN
    try:
        if len(sys.argv) > 1:
            params = json.loads(base64.b64decode(sys.argv[1]))
        else:
            line = sys.stdin.readline()
            params = json.loads(line) if line else {}
    except:
        params = {}


    user_id = params.get('user_id', 1)
    e_ini = params.get('data_ini') or (datetime.now().strftime('%d/%m/%Y'))
    e_fim = params.get('data_fim') or e_ini
    b_ini = params.get('data_ini_base') or (datetime.now() - timedelta(days=1)).strftime('%d/%m/%Y')
    b_fim = params.get('data_fim_base') or b_ini
    
    # Datas podem vir em ISO do frontend, converter se necessário
    def fix_date(d):
        if not d:
            return d
        return normalizar_data_br(d)

    e_ini = fix_date(e_ini)
    e_fim = fix_date(e_fim)
    b_ini = fix_date(b_ini)
    b_fim = fix_date(b_fim)

    def progress_callback(p, m):
        # Formato que o server.js captura
        print(f'PROGRESS:{{"p": {int(p*100)}, "m": "{m}"}}', flush=True)

    try:
        resultado = executar_sr(
            id_usuario=user_id,
            e_ini=e_ini,
            e_fim=e_fim,
            b_ini=b_ini,
            b_fim=b_fim,
            callback_progresso=progress_callback,
            modo_execucao=params.get('acao', 'completo'),
            pasta_destino=params.get('pasta_saida'),
            arquivo_entrada=params.get('pasta_personalizada'),
            base_automacao=params.get('base')
        )
        # O último print deve ser o resultado para o backend capturar
        print(json.dumps(resultado))
    except Exception as e:
        print(f"ERRO: {str(e)}", file=sys.stderr)
        sys.exit(1)

