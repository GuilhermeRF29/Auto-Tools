# automações/adm_new.py
import os
import time
import shutil
import calendar
import pandas as pd
from pathlib import Path
from datetime import datetime, timedelta
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.edge.options import Options
import sys
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
try:
    from selenium_helper import get_driver_path
except ImportError:
    def get_driver_path():
        return None
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

# IMPORTAMOS O BANCO PARA PUXAR A SENHA
from core.banco import buscar_credencial_site

class CanceladoPeloUsuario(Exception):
    pass

def gerar_intervalos_mensais(data_str_inicio, data_str_fim):
    """Divide a janela de datas informada em pedaços mensais para puxar 
    no site (pois pode existir limite de dados por pesquisa no ADM)."""
    inicio = datetime.strptime(data_str_inicio, "%d/%m/%Y")
    fim_global = datetime.strptime(data_str_fim, "%d/%m/%Y")
    atual = inicio
    intervalos = []
    while atual <= fim_global:
        ultimo_dia_mes = calendar.monthrange(atual.year, atual.month)[1]
        data_ultimo_dia = datetime(atual.year, atual.month, ultimo_dia_mes)
        fim_trecho = min(data_ultimo_dia, fim_global)
        intervalos.append((atual.strftime("%d/%m/%Y"), fim_trecho.strftime("%d/%m/%Y")))
        atual = data_ultimo_dia + timedelta(days=1)
    return intervalos

def renomear_arquivo(caminho_arquivo, caminho_destino, contador):
    """Move e renomeia o arquivo que acabou de ser baixado, adicionando um sufixo numérico."""
    extensao = caminho_arquivo.suffix 
    novo_nome = f"Temp_Relatorio_{contador}{extensao}"
    novo_caminho = caminho_destino / novo_nome
    shutil.move(str(caminho_arquivo), str(novo_caminho))
    return novo_caminho

def consolidar_varios_arquivos(diretorio_origem, arquivo_saida, callback_progresso=None):
    """Junta todos os Temp_Relatorios fragmentados que o robô baixou em um Excel final,
    padronizando tipos (Linha) e datas."""
    if callback_progresso: callback_progresso(0.8, "Analisando relatórios temporários baixados...")
    caminho_origem = Path(diretorio_origem)
    lista_dfs = []
    colunas_padrao = None
    arquivos_excel = sorted([f for f in caminho_origem.glob("Temp_Relatorio_*")])

    if not arquivos_excel:
        return

    total_arquivos = len(arquivos_excel)
    for index, arquivo in enumerate(arquivos_excel, 1):
        if callback_progresso: callback_progresso(0.80 + (0.05 * (index/total_arquivos)), f"Lendo e padronizando arquivo {index} de {total_arquivos}...")
        print(f"Inspecionando: {arquivo.name}")
        xls = pd.ExcelFile(arquivo)
        for nome_aba in xls.sheet_names:
            df_aba = pd.read_excel(xls, sheet_name=nome_aba, header=None)
            df_aba = df_aba.dropna(how='all').dropna(axis=1, how='all')
            if not df_aba.empty:
                primeira_celula = str(df_aba.iloc[0, 0]).strip()
                if primeira_celula.lower() == "data":
                    if colunas_padrao is None:
                        colunas_padrao = df_aba.iloc[0].tolist()
                    df_aba = df_aba.iloc[1:].copy()
                if df_aba.empty:
                    continue
                if colunas_padrao is not None:
                    if len(df_aba.columns) == len(colunas_padrao):
                        df_aba.columns = colunas_padrao
                        lista_dfs.append(df_aba)
        xls.close()

    if lista_dfs:
        if callback_progresso: callback_progresso(0.86, "Base final montada. Concatenando e formatando dados...")
        df_final = pd.concat(lista_dfs, ignore_index=True)
        if "Data" in df_final.columns:
            df_final["Data"] = pd.to_datetime(df_final["Data"], errors='coerce').dt.strftime('%d/%m/%Y')

        df_final['Data Observação'] = datetime.now().strftime('%d/%m/%Y')
        
        # Correção do Dtype e variável (Como conversamos antes!)
        if 'Linha' in df_final.columns:
            df_final['Linha'] = df_final['Linha'].astype(str).str.strip()
            df_final['Linha'] = pd.to_numeric(df_final['Linha'], errors='coerce').astype("Int64")
        
        df_final.to_excel(arquivo_saida, index=False, sheet_name='Planilha1')
        
        destino_final = Path(r"\\172.16.98.12\Relatórios Power BI\Forecast\Forecast2")
        destino_backup = Path(r"\\172.16.98.12\Relatórios Power BI\Forecast\Forecast - Antigo")

        destino_final_arquivo = destino_final / arquivo_saida.name
        
        try:
            if callback_progresso: callback_progresso(0.9, "Gerando Base e realizando Backups na rede (Drive)...")
            shutil.move(str(arquivo_saida), str(destino_final_arquivo))
            shutil.copy2(str(destino_final_arquivo), str(destino_backup / arquivo_saida.name))
            
            if callback_progresso: callback_progresso(0.92, "Formatando as células (Openpyxl)...")
            # Formatação via openpyxl direto na rede
            wb = load_workbook(destino_final_arquivo)
            ws = wb.active
            ws.column_dimensions['D'].width = 18
            ws.column_dimensions['F'].width = 18
            ws.column_dimensions['G'].width = 18
            tamanho_fonte = Font(size=8)
            alinhamento_fonte = Alignment(wrap_text=True)
            for linha in ws.iter_rows():
                for celula in linha:
                    celula.font = tamanho_fonte
                    celula.alignment = alinhamento_fonte
            # Salva e Fecha o WorkBook para liberar o lock do Windows no arquivo da rede
            wb.save(destino_final_arquivo)
            wb.close()
            print(f"\nConsolidação e formatação concluídas na rede!")
        except Exception as e:
            print(f"Erro ao mover/formatar na rede: {e}")

        # Limpa todos os relatórios baixados via selenium da pasta temporária
        for f in arquivos_excel:
            try:
                os.remove(f)
            except:
                pass

def buscar_e_mover_arquivo_ano_passado(diretorio_origem, diretorio_destino):
    """
    Busca o relatório da mesma semana do ano anterior para ser utilizado como
    base de comparação/dado na nova dashboard do PowerBI.
    """
    hoje = datetime.now()
    try:
        data_ano_passado = hoje.replace(year=hoje.year - 1)
    except ValueError:
        data_ano_passado = hoje.replace(year=hoje.year - 1, day=28)
        
    dias_para_domingo = (data_ano_passado.weekday() + 1) % 7
    domingo_ano_passado = data_ano_passado - timedelta(days=dias_para_domingo)
    
    print(f"Buscando arquivos da semana que começou no domingo: {domingo_ano_passado.strftime('%d/%m/%Y')}")

    origem = Path(diretorio_origem)
    destino = Path(diretorio_destino)
    arquivo_encontrado = None
    
    for i in range(7):
        data_busca = domingo_ano_passado + timedelta(days=i)
        data_str = data_busca.strftime("%d-%m-%Y") 
        arquivos_possiveis = list(origem.glob(f"*{data_str}*.*"))
        
        if arquivos_possiveis:
            arquivo_encontrado = arquivos_possiveis[0]
            print(f"Sucesso! Arquivo encontrado no dia {data_str}: {arquivo_encontrado.name}")
            break
        else:
            print(f"Sem arquivo para o dia {data_str}. Tentando o próximo dia da semana...")

    if arquivo_encontrado:
        caminho_final = destino / arquivo_encontrado.name
        shutil.copy2(str(arquivo_encontrado), str(caminho_final)) 
        print(f"\nArquivo transferido para: {caminho_final}")
        return caminho_final
    else:
        print("\nPoxa, nenhum arquivo encontrado em nenhum dia dessa semana do ano passado.")
        return None

# =========================================================================
# A FUNÇÃO PRINCIPAL: É ELA QUE O MAIN.PY CHAMA
# =========================================================================
def executar_adm(id_usuario_logado, data_inicio, data_final, callback_progresso=None, hook_cancelamento=None):
    """Executa o processo do Selenium e Pandas para o ADM."""
    def checar_parada():
        if hook_cancelamento and hook_cancelamento():
            raise CanceladoPeloUsuario("Processo cancelado pelo usuário.")

    if callback_progresso: callback_progresso(0.05, "Inicializando Robô ADM...")
    
    # 1. Puxa as credenciais do cofre automaticamente!
    user, passwd = buscar_credencial_site(id_usuario_logado, "ADM de Vendas")
    
    if not user:
        if callback_progresso: callback_progresso(0.0, "ERRO: Credenciais do ADM não encontradas no Cofre!")
        return

    resultado = gerar_intervalos_mensais(data_inicio, data_final)

    home = Path.home()
    origem = home / "Downloads"
    destino = home / "Documents" / "Relatório Demanda"
    destino.mkdir(parents=True, exist_ok=True)

    opcoes = Options()
    opcoes.add_argument("--window-size=1920,1080")
    opcoes.add_argument("--headless") # Recomendo deixar comentado enquanto testa

    if callback_progresso: callback_progresso(0.1, "Abrindo Navegador Invisível...")
    driver_path = get_driver_path()
    if driver_path:
        from selenium.webdriver.edge.service import Service
        driver = webdriver.Edge(service=Service(driver_path), options=opcoes)
    else:
        driver = webdriver.Edge(options=opcoes)
    
    try:
        checar_parada()
        driver.get("http://ttadm01.jcatlm.com.br:8080/ventaboletosadm/index.zul;jsessionid=xFIW8nh_t8n9-74topChhriraeW-2Y5y-MKUCIG3.gcp-pd-ttadm-01")

        wait = WebDriverWait(driver, 1000)

        if callback_progresso: callback_progresso(0.15, "Injetando Credenciais...")
        # Login automático com as credenciais puxadas do DB
        login_element = wait.until(EC.presence_of_element_located((By.NAME, "j_username")))
        login_element.click()
        login_element.clear()
        login_element.send_keys(user)

        senha = driver.find_element(By.NAME, "j_password")
        senha.click()
        senha.clear()
        if passwd:
            senha.send_keys(passwd)

        acesso = driver.find_element(By.NAME, "btnAcessar")
        acesso.click()

        if callback_progresso: callback_progresso(0.2, "Navegando entre MENUS de Demandas...")
        # O Resto do seu código Selenium continua igual...
        relatorios = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[contains(text(), 'Relatórios')]")))
        relatorios.click()
        
        relatorios = wait.until(EC.element_to_be_clickable((By.XPATH, "//a [contains(text(), ' Relatórios')and contains(@id, 'c4')]")))
        relatorios.click()
        time.sleep(1)
        
        relatorios = wait.until(EC.element_to_be_clickable((By.XPATH, "//a [contains(text(), ' Relatórios Operacionais')and contains(@id, '25-a')]")))
        relatorios.click()
        
        relatorios = wait.until(EC.element_to_be_clickable((By.XPATH, "//a [contains(text(), ' Demandas')and contains(@id, 'a5-a')]")))
        relatorios.click()

        total_blocos = len(resultado)
        for idx, (inicial_data, final_data) in enumerate(resultado):
            checar_parada()
            porc = 0.25 + (0.45 * (idx / total_blocos))
            if callback_progresso: callback_progresso(porc, f"Lote {idx+1}/{total_blocos} - Inserindo filtro: {inicial_data} a {final_data}...")
            
            data1 = wait.until(EC.element_to_be_clickable((By.XPATH, "//input[contains (@id, 'ia-real')]")))
            data1.clear()
            data1.send_keys(inicial_data)
            
            data2 = wait.until(EC.element_to_be_clickable((By.XPATH, "//input[contains(@id, 'la-real')]")))
            data2.clear()
            data2.send_keys(final_data)

            btn_layout = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[contains(text(), ' Novo Layout')and contains(@id, 'zb')]")))
            btn_layout.click()

            btn_salvar = wait.until(EC.element_to_be_clickable((By.XPATH, "//div[contains(@id, '4c')]")))
            btn_salvar.click()

            if callback_progresso: callback_progresso(porc + 0.05, f"Lote {idx+1}/{total_blocos} - Download solicitado. Aguardando arquivo...")
            tempo_espera = 0
            while tempo_espera < 30: 
                arquivos = list(origem.rglob("RelatorioDemandaDetalhado*.xls*"))
                arquivos_validos = [arq for arq in arquivos if not arq.name.endswith(('.crdownload', '.tmp')) and "_a_" not in arq.name]

                if arquivos_validos:
                    if callback_progresso: callback_progresso(porc + 0.06, f"Lote {idx+1}/{total_blocos} - Arquivo detectado (espera: {tempo_espera}s).")
                    time.sleep(1)
                    break
                    
                time.sleep(1)
                tempo_espera += 1.5
            
            checar_parada()
            btn_voltar = wait.until(EC.element_to_be_clickable((By.XPATH, "//div[contains(@id, 'c-close')]")))
            btn_voltar.click()
            time.sleep(1)

        time.sleep(10)
        checar_parada()
        arquivos_encontrados = [arq for arq in origem.glob("RelatorioDemanda*.xls*") 
                                if not arq.name.endswith(('.crdownload', '.tmp'))]

        if callback_progresso: callback_progresso(0.75, "Agrupando planilhas temporárias...")

        if arquivos_encontrados:
            for i, arquivo in enumerate(arquivos_encontrados, 1):
                renomear_arquivo(arquivo, destino, i)

            data_execucao = datetime.now().strftime('%d-%m-%Y')
            arquivo_final_path = destino / f"{data_execucao}.xlsx"
            checar_parada()
            consolidar_varios_arquivos(destino, arquivo_final_path, callback_progresso)
            
            if callback_progresso: callback_progresso(0.98, "Baixando baseção ano passado...")
            print("Processo finalizado com sucesso.")
            origem_nova = Path(r"\\172.16.98.12\Relatórios Power BI\Forecast\Forecast - Antigo")
            destino_novo = Path(r"\\172.16.98.12\Relatórios Power BI\Forecast\Forecast2")
            buscar_e_mover_arquivo_ano_passado(origem_nova, destino_novo)
        else:
            print("Nenhum arquivo encontrado em Downloads.")

    except CanceladoPeloUsuario as erro_cancel:
        if callback_progresso: callback_progresso(0, str(erro_cancel))
        raise # Sobe o erro pra ser tratado na GUI

    finally:
        try:
            driver.quit()
        except:
            pass