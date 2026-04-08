import sys
import os
# Feedback imediato para o dashboard!
print("PROGRESS:{\"p\": 1, \"m\": \"Carregando módulos eBus...\"}", flush=True)

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import time
import shutil
import calendar
import unicodedata
import pandas as pd # type: ignore
from pathlib import Path
from datetime import datetime, timedelta
from selenium import webdriver # type: ignore
from selenium.webdriver.common.by import By # type: ignore
from selenium.webdriver.chrome.options import Options # type: ignore

try:
    from selenium_helper import get_driver_path
except ImportError:
    def get_driver_path():
        return None
from selenium.webdriver.support.ui import WebDriverWait # type: ignore
from selenium.webdriver.support import expected_conditions as EC # type: ignore
from openpyxl import load_workbook # type: ignore
from openpyxl.utils import get_column_letter # type: ignore
from openpyxl.styles import PatternFill, Alignment # type: ignore
from core.banco import buscar_credencial_site # type: ignore

class CanceladoPeloUsuario(Exception):
    pass

MESES_PT = {1: 'Janeiro', 2: 'Fevereiro', 3: 'Março', 4: 'Abril', 5: 'Maio', 
            6: 'Junho', 7: 'Julho', 8: 'Agosto', 9: 'Setembro', 10: 'Outubro', 
            11: 'Novembro', 12: 'Dezembro'}

def sanitizar_nome_arquivo(nome):
    """Remove caracteres inválidos para nome de arquivo no Windows."""
    invalidos = '<>:"/\\|?*'
    for ch in invalidos:
        nome = nome.replace(ch, "-")
    return nome

def gerar_intervalos_mensais_ebus(data_str_inicio, data_str_fim):
    meses_pt = {1: 'Janeiro', 2: 'Fevereiro', 3: 'Março', 4: 'Abril', 5: 'Maio', 
    6: 'Junho', 7: 'Julho', 8: 'Agosto', 9: 'Setembro', 10: 'Outubro', 
    11: 'Novembro', 12: 'Dezembro'}
    inicio = datetime.strptime(data_str_inicio, "%d/%m/%Y")
    fim_global = datetime.strptime(data_str_fim, "%d/%m/%Y")
    atual = inicio
    intervalos = []
    while atual <= fim_global:
        ultimo_dia_mes = calendar.monthrange(atual.year, atual.month)[1]
        data_ultimo_dia = datetime(atual.year, atual.month, ultimo_dia_mes)
        fim_trecho = min(data_ultimo_dia, fim_global)
        intervalos.append((atual.strftime("%d/%m/%Y"), fim_trecho.strftime("%d/%m/%Y"), meses_pt[atual.month], atual.year))
        atual = data_ultimo_dia + timedelta(days=1)
    return intervalos


def ler_planilha_relatorio(caminho_arquivo):
    """Lê planilhas .xls/.xlsx do EBUS usando o engine compatível com a extensão."""
    caminho = Path(caminho_arquivo)
    sufixo = caminho.suffix.lower()

    if sufixo == '.xls':
        return pd.read_excel(caminho, engine='xlrd')

    if sufixo in {'.xlsx', '.xlsm'}:
        return pd.read_excel(caminho, engine='openpyxl')

    # Fallback para extensões não previstas.
    try:
        return pd.read_excel(caminho)
    except Exception:
        return pd.read_excel(caminho, engine='openpyxl')


def encontrar_coluna_data_viagem(df: pd.DataFrame):
    """Localiza a coluna de Data Viagem mesmo com variações de grafia."""
    for col in df.columns:
        normalizado = normalizar_nome_coluna(col)
        if normalizado == 'dataviagem' or ('data' in normalizado and 'viagem' in normalizado):
            return col
    return None


def separar_arquivo_por_mes_data_viagem(arquivo_original, pasta_destino, callback_progresso=None):
    """
    Separa um relatório único em múltiplos arquivos por mês de Data Viagem.
    Mantém o arquivo original caso não seja possível separar com segurança.
    """
    arquivo_original = Path(arquivo_original)
    pasta_destino = Path(pasta_destino)
    df = ler_planilha_relatorio(arquivo_original)
    df = df.dropna(how='all')

    coluna_data = encontrar_coluna_data_viagem(df)
    if not coluna_data:
        if callback_progresso:
            callback_progresso(0.54, "Data Viagem não encontrada; arquivo seguirá sem separação mensal.")
        return [{"arquivo": arquivo_original, "nome_mes": None, "ano": None}]

    df_aux = df.copy()
    df_aux['__data_parse'] = pd.to_datetime(df_aux[coluna_data], errors='coerce', dayfirst=True)
    mascara_validas = df_aux['__data_parse'].notna()
    if not mascara_validas.any():
        if callback_progresso:
            callback_progresso(0.54, "Nenhuma Data Viagem válida; arquivo seguirá sem separação mensal.")
        return [{"arquivo": arquivo_original, "nome_mes": None, "ano": None}]

    df_validas = df_aux.loc[mascara_validas].copy()
    df_validas['__mes'] = df_validas['__data_parse'].apply(lambda dt: int(dt.month))
    df_validas['__ano'] = df_validas['__data_parse'].apply(lambda dt: int(dt.year))
    df_validas = df_validas.drop(columns=['__data_parse'])

    arquivos_saida = []
    chaves_ordenadas = sorted({(int(ano), int(mes)) for ano, mes in zip(df_validas['__ano'], df_validas['__mes'])})
    total_grupos = len(chaves_ordenadas)

    for idx, (ano, mes) in enumerate(chaves_ordenadas, 1):
        grupo = df_validas[(df_validas['__ano'] == ano) & (df_validas['__mes'] == mes)].copy()
        nome_mes = MESES_PT.get(int(mes), f"Mes {mes}")
        nome_arquivo = sanitizar_nome_arquivo(
            f"RelatorioRevenue - {nome_mes} {int(ano)} - origem_unica.xlsx"
        )
        caminho_mes = pasta_destino / nome_arquivo

        grupo_saida = grupo.drop(columns=['__mes', '__ano'])
        grupo_saida.to_excel(caminho_mes, index=False)

        arquivos_saida.append({"arquivo": caminho_mes, "nome_mes": nome_mes, "ano": int(ano)})

        if callback_progresso:
            callback_progresso(
                0.54 + (0.06 * (idx / max(total_grupos, 1))),
                f"Separação mensal: {nome_mes}/{ano} preparado ({idx}/{total_grupos}).",
            )

    linhas_sem_data = int((~mascara_validas).sum())
    if linhas_sem_data > 0 and callback_progresso:
        callback_progresso(
            0.60,
            f"Aviso: {linhas_sem_data} linhas ficaram fora da separação por mês por não terem Data Viagem válida.",
        )

    try:
        os.remove(arquivo_original)
    except OSError:
        pass

    return arquivos_saida


def limpar_cookies_ebus(driver, callback_progresso=None):
    """
    Sequência de higienização pedida para o EBUS:
    acessa > apaga tudo > recarrega > segue para login.
    """
    url_login = "http://10.61.65.84/auth/login"
    origem_ebus = "http://10.61.65.84"

    def ler_estado_storage():
        try:
            estado = driver.execute_script(
                "return {"
                "localCount: window.localStorage ? window.localStorage.length : -1,"
                "sessionCount: window.sessionStorage ? window.sessionStorage.length : -1"
                "};"
            ) or {}
            return int(estado.get("localCount", -1)), int(estado.get("sessionCount", -1))
        except Exception:
            return -1, -1

    driver.get(url_login)
    time.sleep(1)

    limpo_total = False
    for tentativa in range(1, 4):
        try:
            driver.delete_all_cookies()
        except Exception:
            pass

        # Apaga armazenamento web do domínio atual de forma resiliente.
        try:
            driver.execute_script("window.localStorage.clear();")
            driver.execute_script("window.sessionStorage.clear();")
        except Exception:
            pass

        # Em navegadores Chromium, limpa também cookies/cache/dados do site.
        try:
            driver.execute_cdp_cmd("Network.enable", {})
            driver.execute_cdp_cmd("Network.clearBrowserCookies", {})
            driver.execute_cdp_cmd("Network.clearBrowserCache", {})
            driver.execute_cdp_cmd("Storage.clearDataForOrigin", {
                "origin": origem_ebus,
                "storageTypes": "local_storage,session_storage,indexeddb,websql,cache_storage,service_workers",
            })
        except Exception:
            pass

        local_count, session_count = ler_estado_storage()
        try:
            cookies_count = len(driver.get_cookies())
        except Exception:
            cookies_count = -1

        limpo_total = cookies_count == 0 and local_count == 0 and session_count == 0
        if callback_progresso:
            callback_progresso(
                0.14,
                f"Validação de limpeza ({tentativa}/3): cookies={cookies_count}, localStorage={local_count}, sessionStorage={session_count}",
            )

        if limpo_total:
            break

        time.sleep(0.8)

    # Recarrega conforme processo validado manualmente.
    try:
        driver.refresh()
    except Exception:
        driver.get(url_login)

    time.sleep(1)
    return limpo_total

# ==========================================
# FUNÇÃO AUXILIAR DE FORMATAÇÃO VISUAL
# ==========================================
def aplicar_formatacao_excel(caminho_arquivo, nome_aba):
    """
    Abre um arquivo Excel existente, aplica cores de fundo no cabeçalho,
    ajusta larguras das colunas e centraliza os textos.
    É fundamental usar wb.close() no final para liberar o arquivo no Windows.
    """
    wb = load_workbook(caminho_arquivo)
    ws = wb[nome_aba]
    
    # Define a cor de fundo Cinza para o cabeçalho
    fundo_cabeçalho = PatternFill(start_color="BFBFBF", fill_type="solid")
    alinhamento = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Aumenta a altura da primeira linha (Cabeçalho)
    ws.row_dimensions[1].height = 25
    
    # Nomes das colunas para formatação específica (em letras minúsculas e sem espaços nas pontas)
    colunas_moeda_limpas = [
        'tarifa', 'embarque', 'pedágio', 'pedagio', 
        'sugestão', 'sugestao', 'sugestão revenue', 'sugestao revenue',
        'revenue aplicado'
    ]
    colunas_porcentagem_limpas = [
        'market share', 'perc. de diferença', 'perc. de diferenca', 'aproveitamento'
    ]

    indices_moeda = []
    indices_porcentagem = []

    # Aplica o estilo para cada coluna no cabeçalho e identifica os índices das colunas especiais
    for col in range(1, ws.max_column + 1):
        letra = get_column_letter(col)
        celula = ws.cell(row=1, column=col)
        celula.fill = fundo_cabeçalho
        celula.alignment = alinhamento
        
        # Guarda o índice para formatar o resto da coluna depois
        nome_cabecalho = str(celula.value).strip().lower() if celula.value else ""
        
        if nome_cabecalho in colunas_moeda_limpas:
            indices_moeda.append(col)
        elif nome_cabecalho in colunas_porcentagem_limpas:
            indices_porcentagem.append(col)
        
        # A coluna 'S' (Justificativa no relatorio base) recebe uma largura maior, o resto recebe 20
        ws.column_dimensions[letra].width = 75.6 if letra == 'S' else 20
        
    # Centraliza todas as células das linhas normais e aplica formatos numéricos
    for linha in ws.iter_rows(min_row=2):
        for celula in linha: 
            celula.alignment = Alignment(vertical="center", wrap_text=True)
            
            # Formata apenas se a célula contiver um número (ignora células vazias ou texto)
            if celula.column in indices_moeda:
                if isinstance(celula.value, (int, float)):
                    celula.number_format = 'R$ #,##0.00_-'
            elif celula.column in indices_porcentagem:
                if isinstance(celula.value, (int, float)):
                    celula.number_format = '0.00 %'
            
    # Salva e FECHA o arquivo para evitar erros de permissão ao mover
    wb.save(caminho_arquivo)
    wb.close() 

# ==========================================
# FUNÇÃO AUXILIAR DA JUSTIFICATIVA
# ==========================================
def padronizar_justificativa(df: pd.DataFrame) -> pd.DataFrame:
    if 'Status Revenue' not in df.columns:
        return df

    if 'Justificativa' not in df.columns:
        df['Justificativa'] = ''
    
    novos_valores = []
    for just, status in zip(df['Justificativa'], df['Status Revenue']):
        # Limpeza segura de valores nulos ou vazios
        just_str = '' if pd.isna(just) else str(just).strip().upper()
        status_str = '' if pd.isna(status) else str(status).strip().upper()
        
        eh_vazio = just_str in ('', 'NAN', 'NONE', 'NULL', 'NAT')
        
        if eh_vazio and status_str == 'APROVADO':
            novos_valores.append('Aprovado')
        elif eh_vazio and status_str in ('REPROVADO', 'RECUSADO'):
            novos_valores.append('Concorrência')
        else:
            novos_valores.append(str(just) if not pd.isna(just) else '')
            
    df['Justificativa'] = novos_valores
    return df


def normalizar_nome_coluna(nome_coluna: str) -> str:
    """Normaliza nome de coluna para comparação sem acentos/espaços/caracteres especiais."""
    nome = unicodedata.normalize('NFKD', str(nome_coluna)).encode('ascii', 'ignore').decode('ascii')
    return ''.join(ch.lower() for ch in nome if ch.isalnum())


def encontrar_coluna_data_aplicacao(df: pd.DataFrame):
    """Localiza a coluna de Data Aplicação mesmo com variações de grafia."""
    for col in df.columns:
        normalizado = normalizar_nome_coluna(col)
        if normalizado == 'dataaplicacao' or ('data' in normalizado and 'aplic' in normalizado):
            return col
    return None


def data_aplicacao_parece_timestamp_download(df: pd.DataFrame, nome_coluna: str) -> bool:
    """
    Detecta o padrão problemático em que Data Aplicação vem praticamente igual para todo o arquivo
    (ou com variação mínima de segundos), típico de timestamp de download.
    """
    if nome_coluna not in df.columns or df.empty:
        return False

    serie = pd.to_datetime(df[nome_coluna], errors='coerce', dayfirst=True)
    serie_valida = serie.dropna()
    if serie_valida.empty:
        return False

    cobertura = len(serie_valida) / max(len(df), 1)
    if cobertura < 0.85:
        # Se a coluna quase não converte para data, não tratamos como padrão de timestamp.
        return False

    serie_segundos = serie_valida.dt.floor('s').astype('int64') // 10**9
    total = len(serie_segundos)
    unicos = int(serie_segundos.nunique())
    span_segundos = int(serie_segundos.max() - serie_segundos.min()) if total > 1 else 0

    # Permite pequena variação de segundos mantendo o comportamento de "carimbo quase único".
    if total < 200:
        limite_unicos = max(3, int(total * 0.03))
    else:
        limite_unicos = max(3, int(total * 0.01))

    return unicos <= limite_unicos and span_segundos <= 180


def data_aplicacao_top10_iguais(arquivo_excel) -> bool:
    """
    Valida as 10 primeiras linhas de Data Aplicação.
    Retorna True quando todos os valores válidos são exatamente iguais.
    """
    try:
        df = ler_planilha_relatorio(arquivo_excel)
    except Exception:
        return False

    coluna_data = encontrar_coluna_data_aplicacao(df)
    if not coluna_data or coluna_data not in df.columns:
        return False

    serie = df[coluna_data].head(10)
    if serie.empty:
        return False

    valores = []
    for valor in serie:
        if pd.isna(valor):
            continue

        convertido = pd.to_datetime(valor, errors='coerce', dayfirst=True)
        if isinstance(convertido, pd.Timestamp) and not pd.isna(convertido):
            valores.append(convertido.floor('s').strftime('%Y-%m-%d %H:%M:%S'))
        else:
            valores.append(str(valor).strip())

    if len(valores) < 2:
        return False

    return len(set(valores)) == 1


def valor_para_chave_comparacao(valor):
    """Padroniza valores para composição de chave de comparação de estado."""
    if pd.isna(valor):
        return ''
    if isinstance(valor, pd.Timestamp):
        return valor.floor('s').strftime('%Y-%m-%d %H:%M:%S')
    if isinstance(valor, datetime):
        return valor.strftime('%Y-%m-%d %H:%M:%S')
    return str(valor).strip()


def montar_chaves_estado(df: pd.DataFrame, colunas_comparacao):
    """Cria chaves estáveis para casar linhas repetidas sem cruzamento cartesiano."""
    resultado = df.copy()

    if not colunas_comparacao:
        resultado['__cmp_key'] = ''
    else:
        resultado['__cmp_key'] = resultado[colunas_comparacao].apply(
            lambda row: '|'.join(valor_para_chave_comparacao(v) for v in row.values),
            axis=1,
        )

    resultado['__cmp_idx'] = resultado.groupby('__cmp_key', dropna=False).cumcount()
    return resultado



# ==========================================
# MOTOR DE CONCILIAÇÃO DOS ARQUIVOS EBUS
# ==========================================
def processar_arquivos_relatorios(arquivo_original, destino, nome_mes=None, ano_atual=None, callback_progresso=None, destino_base=None, destino_saida=None):
    """
    Função coração da regra de negócio EBUS.
    1. Lê a planilha web crua.
    2. Valida se veio vazia.
    3. Concatena colunas de Origem + Destino e aplica Justificativa.
    4. Processa a "Base Nova" (Histórico de Novo, Excluido, Presente).
    5. Processa a "Base Normal" (Mercado vazio, regra de volume).
    """
    nome_arquivo_origem = Path(arquivo_original).name
    if callback_progresso: callback_progresso(0.40, f"Lendo planilha Excel bruta: {nome_arquivo_origem}")
    
    df_novo = ler_planilha_relatorio(arquivo_original)

    if callback_progresso: callback_progresso(0.42, f"Removendo linhas fantasmas ou inválidas...")
    df_novo = df_novo.dropna(how='all')
    
    # =========================================================
    # DETECÇÃO AUTOMÁTICA DE MÊS E ANO (Via Data Viagem)
    # =========================================================
    if nome_mes is None or ano_atual is None:
        if 'Data Viagem' in df_novo.columns:
            # Tenta converter a primeira data válida para datetime com formato brasileiro (dayfirst=True)
            datas_validas = pd.to_datetime(df_novo['Data Viagem'], format='%d/%m/%Y', errors='coerce')
            if datas_validas.isna().all():
                # Plano B: se tiver formato de hora junto ou outro separador
                datas_validas = pd.to_datetime(df_novo['Data Viagem'], dayfirst=True, errors='coerce')
            
            datas_validas = datas_validas.dropna()
            if not datas_validas.empty:
                primeira_data = datas_validas.iloc[0]
                nome_mes = MESES_PT[primeira_data.month]
                ano_atual = primeira_data.year
            else:
                nome_mes = nome_mes or "Desconhecido"
                ano_atual = ano_atual or datetime.now().year
        else:
            nome_mes = nome_mes or "Desconhecido"
            ano_atual = ano_atual or datetime.now().year

    # Verifica se o relatório veio vazio direto da mensagem do site
    mensagem_vazio = 'Não foi possivel obter dados com os parâmetros informados.'
    # Procura a mensagem tanto nas colunas (se o Pandas a leu como cabeçalho) quanto dentro dos dados
    if df_novo.empty or mensagem_vazio in df_novo.columns or df_novo.apply(lambda col: col.astype(str).str.contains(mensagem_vazio, regex=False)).any().any():
        if callback_progresso: callback_progresso(0.45, f"Relatório de {nome_mes} está sem dados no sistema. Descartando e ignorando...")
        try:
            os.remove(arquivo_original)
        except OSError:
            pass
        return {
            "arquivo_principal": None,
            "arquivos_saida": [],
            "pasta_final": None,
            "mensagem": f"Relatório de {nome_mes} sem dados no sistema.",
        }

    # =========================================================
    # TRATAMENTOS UNIVERSAIS (Aplicados a todas as bases)
    # =========================================================
    if 'Origem' in df_novo.columns and 'Destino' in df_novo.columns:
        df_novo['Concatenar Origem e Destino'] = df_novo['Origem'].astype(str) + " X " + df_novo['Destino'].astype(str)

    if callback_progresso: callback_progresso(0.45, f"{nome_mes} - Executando motor de Justificativas e Status do Revenue...")
    # Aplica as regras da Justificativa antes de começar as comparações
    df_novo = padronizar_justificativa(df_novo)

    # =========================================================
    # ESTRUTURA DE PASTAS E ARQUIVOS
    # =========================================================
    nome_oficial = f"Relatorio Revenue Completo - {nome_mes.capitalize()} {ano_atual}.xlsx"

    caminho_rede = Path(r"\\172.16.98.12")
    destino_referencia = Path(destino_base) if destino_base else (caminho_rede / "Relatórios Power BI" / "DASH REVENUE APPLICATION")
    destino_saida_final = Path(destino_saida) if destino_saida else destino_referencia

    pasta_base_ref = destino_referencia / "BASE"
    pasta_base_nova_ref = destino_referencia / "BASE NOVA"

    pasta_base = destino_saida_final / "BASE"
    pasta_backup = destino_saida_final / "BACKUP"
    pasta_base_nova = destino_saida_final / "BASE NOVA"
    pasta_backup_base_nova = destino_saida_final / "BACKUP NOVO"

    # Criando todas as pastas de saída de forma limpa
    for pasta in [pasta_base, pasta_backup, pasta_base_nova, pasta_backup_base_nova]:
        pasta.mkdir(parents=True, exist_ok=True)

    caminho_base_ref = pasta_base_ref / nome_oficial
    caminho_base_nova_ref = pasta_base_nova_ref / nome_oficial
    caminho_base = pasta_base / nome_oficial
    caminho_base_nova = pasta_base_nova / nome_oficial

    # =========================================================
    # PARTE 1: BASE NOVA (Mapeamento de Novo / Excluido / Presente)
    # =========================================================
    if callback_progresso: callback_progresso(0.50, f"{nome_mes} - Preparando comparação histórica de Base Nova...")
    houve_alteracao_estado = True

    if caminho_base_nova_ref.exists():
        df_base_estado_antiga = pd.read_excel(caminho_base_nova_ref)
        
        if 'Estado' in df_base_estado_antiga.columns:
            df_old_active = df_base_estado_antiga[df_base_estado_antiga['Estado'] != 'Excluido'].copy()
            df_previously_excluded = df_base_estado_antiga[df_base_estado_antiga['Estado'] == 'Excluido'].copy()
        else:
            df_old_active = df_base_estado_antiga.copy()
            df_previously_excluded = pd.DataFrame()
            
        df_old_comp = df_old_active.drop(columns=['Estado'], errors='ignore')
        
        df_new_comp = df_novo.copy()
        
        df_old_comp = padronizar_justificativa(df_old_comp)

        colunas_comuns = [c for c in df_new_comp.columns if c in df_old_comp.columns]

        # Harmoniza tipos das colunas comuns para evitar divergências artificiais na chave.
        for col in colunas_comuns:
            if df_old_comp[col].dtype != df_new_comp[col].dtype:
                df_old_comp[col] = df_old_comp[col].astype(str)
                df_new_comp[col] = df_new_comp[col].astype(str)

        coluna_data_old = encontrar_coluna_data_aplicacao(df_old_comp)
        coluna_data_new = encontrar_coluna_data_aplicacao(df_new_comp)

        ignorar_data_aplicacao = False
        coluna_data_validacao = None

        if coluna_data_new and coluna_data_new in colunas_comuns:
            coluna_data_validacao = coluna_data_new
        elif coluna_data_old and coluna_data_old in colunas_comuns:
            coluna_data_validacao = coluna_data_old

        if coluna_data_validacao:
            flag_old = coluna_data_old and data_aplicacao_parece_timestamp_download(df_old_comp, coluna_data_old)
            flag_new = coluna_data_new and data_aplicacao_parece_timestamp_download(df_new_comp, coluna_data_new)
            ignorar_data_aplicacao = bool(flag_old or flag_new)

        if ignorar_data_aplicacao and coluna_data_validacao:
            colunas_comparacao = [c for c in colunas_comuns if c != coluna_data_validacao]
            print(f"[INFO] Modo adaptativo: '{coluna_data_validacao}' ignorada na validação de estado (padrão de timestamp de download detectado).")
            if callback_progresso:
                callback_progresso(0.52, f"{nome_mes} - Ajuste inteligente: Data Aplicação ignorada na validação de estado.")
        else:
            colunas_comparacao = colunas_comuns
            if coluna_data_validacao and callback_progresso:
                callback_progresso(0.52, f"{nome_mes} - Data Aplicação mantida na validação de estado.")

        df_old_comp_keyed = montar_chaves_estado(df_old_comp, colunas_comparacao)
        df_new_comp_keyed = montar_chaves_estado(df_new_comp, colunas_comparacao)

        df_merge_estado = pd.merge(
            df_old_comp_keyed,
            df_new_comp_keyed,
            on=['__cmp_key', '__cmp_idx'],
            how='outer',
            indicator=True,
            suffixes=('_old', '_new'),
        )

        todas_colunas = list(df_new_comp.columns)
        for col in df_old_comp.columns:
            if col not in todas_colunas:
                todas_colunas.append(col)

        df_comparado = pd.DataFrame()
        for col in todas_colunas:
            col_new = f"{col}_new"
            col_old = f"{col}_old"

            if col_new in df_merge_estado.columns and col_old in df_merge_estado.columns:
                df_comparado[col] = df_merge_estado[col_new].combine_first(df_merge_estado[col_old])
            elif col_new in df_merge_estado.columns:
                df_comparado[col] = df_merge_estado[col_new]
            elif col_old in df_merge_estado.columns:
                df_comparado[col] = df_merge_estado[col_old]
        
        novas_mudancas = len(df_merge_estado[df_merge_estado['_merge'] == 'right_only']) + len(df_merge_estado[df_merge_estado['_merge'] == 'left_only'])
        
        mapa_estado = {'left_only': 'Excluido', 'right_only': 'Novo', 'both': 'Manteve'}
        
        df_comparado['Estado'] = df_merge_estado['_merge'].map(mapa_estado)
        
        if not df_previously_excluded.empty:
            df_comparado = pd.concat([df_comparado, df_previously_excluded], ignore_index=True)
            
        if novas_mudancas == 0:
            houve_alteracao_estado = False
            
    else:
        df_comparado = df_novo.copy()
        df_comparado['Estado'] = 'Novo'
        houve_alteracao_estado = True

    if houve_alteracao_estado:
        arquivo_temp_estado = destino / f"temp_estado_{nome_oficial}"
        aba_estado = "Relatorio Comparado"
        
        # Garante que a coluna Mercado não vai sujar a base nova
        if 'Mercado' in df_comparado.columns:
            df_comparado = df_comparado.drop(columns=['Mercado'])

        df_comparado.to_excel(arquivo_temp_estado, index=False, sheet_name=aba_estado)
        if callback_progresso: callback_progresso(0.55, f"{nome_mes} - Aplicando estilos na Base Nova...")
        aplicar_formatacao_excel(arquivo_temp_estado, aba_estado)
        
        if caminho_base_nova.exists():
            if callback_progresso:
                callback_progresso(0.75, f"Fazendo backup da Base Nova: {caminho_base_nova.name}")
            
            # Mantendo histórico também para a Base Nova
            mtime_nova = os.path.getmtime(caminho_base_nova)
            data_mod_nova = datetime.fromtimestamp(mtime_nova).strftime('%d.%m.%Y_%H-%M-%S')
            nome_backup_nova_com_data = sanitizar_nome_arquivo(
                f"{caminho_base_nova.stem} - Backup ({data_mod_nova}){caminho_base_nova.suffix}"
            )
            
            shutil.copy2(str(caminho_base_nova), str(pasta_backup_base_nova / nome_backup_nova_com_data))
            caminho_base_nova.unlink()
            
        shutil.move(str(arquivo_temp_estado), str(caminho_base_nova))
        if callback_progresso:
            callback_progresso(0.78, f"Base Nova atualizada: {caminho_base_nova.name}")
        
        # =========================================================================
        # CRIAÇÃO OU ATUALIZAÇÃO DO ARQUIVO CONSOLIDADO CSV (BASE NOVA)
        # =========================================================================
        csv_consol = pasta_base_nova / 'base_nova_consolidada.csv'
        df_csv_novo = df_comparado.copy()
        
        if csv_consol.exists():
            df_csv_existente = pd.read_csv(csv_consol, sep=';', encoding='utf-8-sig')
            df_final_csv = pd.concat([df_csv_existente, df_csv_novo], ignore_index=True)
            
            # Removemos duplicatas reais baseadas nas colunas chaves
            colunas_unicas = [
                'Origem', 'Destino', 'Data Viagem', 
                'Sugestão Revenue', 'Revenue Aplicado'
            ]
            subset_existente = [c for c in colunas_unicas if c in df_final_csv.columns]
            
            if subset_existente:
                df_final_csv = df_final_csv.drop_duplicates(subset=subset_existente, keep='last')
            else:
                df_final_csv = df_final_csv.drop_duplicates()
        else:
            df_final_csv = df_csv_novo
        
        df_final_csv.to_csv(csv_consol, sep=';', index=False, encoding='utf-8-sig')
        # =========================================================================

    # =========================================================
    # PARTE 2: BASE NORMAL (Com "Mercado" em branco)
    # =========================================================
    if callback_progresso: callback_progresso(0.60, f"{nome_mes} - Avaliando regras de proteção de volume para Base Normal...")
    
    
    df_novo_base = df_novo.copy()
    
    # Prevenção: Garante que a coluna Estado não vaze para a base normal
    if 'Estado' in df_novo_base.columns:
        df_novo_base = df_novo_base.drop(columns=['Estado'])
        
    df_novo_base['Mercado'] = ""

    substituir_base = True
    if caminho_base_ref.exists():
        df_base_antiga = pd.read_excel(caminho_base_ref)
        if len(df_novo_base) <= len(df_base_antiga):
            substituir_base = False

    if substituir_base:
        arquivo_temp_base = destino / f"temp_base_{nome_oficial}"
        aba_normal = "Relatorio Revenue Sistema"
        
        df_novo_base.to_excel(arquivo_temp_base, index=False, sheet_name=aba_normal)
        if callback_progresso: callback_progresso(0.65, f"{nome_mes} - Salvando Base Normal formatada...")
        aplicar_formatacao_excel(arquivo_temp_base, aba_normal)
        
        if caminho_base.exists():
            if callback_progresso:
                callback_progresso(0.85, f"Fazendo backup da Base: {caminho_base.name}")
            
            # Obtém a data de modificação do arquivo atual para compor o nome do backup
            mtime = os.path.getmtime(caminho_base)
            data_mod = datetime.fromtimestamp(mtime).strftime('%d.%m.%Y_%H-%M-%S')
            nome_backup_com_data = sanitizar_nome_arquivo(
                f"{caminho_base.stem} - Backup ({data_mod}){caminho_base.suffix}"
            )
            
            shutil.copy2(str(caminho_base), str(pasta_backup / nome_backup_com_data))
            caminho_base.unlink()
            
        shutil.move(str(arquivo_temp_base), str(caminho_base))
        if callback_progresso:
            callback_progresso(0.88, f"Base principal atualizada: {caminho_base.name}")

    # =========================================================
    # LIMPEZA FINAL
    # =========================================================
    if os.path.exists(arquivo_original):
        os.remove(arquivo_original)
    
    if callback_progresso: callback_progresso(0.9, f"Arquivo processado com sucesso: {nome_arquivo_origem}")

    return {
        "arquivo_principal": str(caminho_base),
        "arquivos_saida": [str(caminho_base), str(caminho_base_nova)],
        "pasta_final": str(destino_saida_final),
        "mensagem": "Processamento EBUS concluído com sucesso.",
    }

# ==========================================
# GESTOR DE NAVEGAÇÃO WEB EBUS (Selenium)
# ==========================================
def executar_ebus(
    id_usuario_logado,
    data_inicio,
    data_final,
    callback_progresso=None,
    hook_cancelamento=None,
    modo_execucao="completo",
    pasta_destino=None,
    arquivo_entrada=None,
    base_automacao=None,
):
    def checar_parada():
        if hook_cancelamento and hook_cancelamento():
            raise CanceladoPeloUsuario("Processo cancelado pelo usuário.")

    modos_validos = {
        "download",
        "download_tratamento",
        "tratamento",
        "tratamento_envio",
        "completo",
        "arquivo_tratamento",
        "arquivo_envio",
        "arquivo_tratamento_envio",
    }
    if modo_execucao not in modos_validos:
        raise ValueError(f"Modo de execução inválido para EBUS: {modo_execucao}")

    destino_envio = Path(pasta_destino) if pasta_destino else None
    destino_padrao_ebus = Path(r"\\172.16.98.12") / "Relatórios Power BI" / "DASH REVENUE"
    base_referencia = Path(base_automacao) if base_automacao else destino_padrao_ebus
    destino_final = destino_envio or base_referencia
    origem_download = Path.home() / "Downloads"
    pasta_trabalho = Path.home() / "Documents" / "Relatórios Revenue"
    pasta_trabalho.mkdir(parents=True, exist_ok=True)

    precisa_download = modo_execucao in {"download", "download_tratamento", "completo"}
    precisa_tratamento = modo_execucao in {
        "download_tratamento",
        "tratamento",
        "tratamento_envio",
        "completo",
        "arquivo_tratamento",
        "arquivo_tratamento_envio",
    }

    if modo_execucao.startswith("arquivo_") and not arquivo_entrada:
        raise ValueError("Selecione um arquivo de entrada para este modo.")

    if not precisa_download and precisa_tratamento and not arquivo_entrada:
        raise ValueError("Tratamento sem download exige um arquivo já baixado.")

    if modo_execucao == "arquivo_envio":
        origem_arquivo = Path(arquivo_entrada)
        if not origem_arquivo.exists():
            raise FileNotFoundError(f"Arquivo de entrada não encontrado: {origem_arquivo}")
        pasta_final = destino_envio or pasta_trabalho
        pasta_final.mkdir(parents=True, exist_ok=True)
        destino_final = pasta_final / origem_arquivo.name
        shutil.copy2(str(origem_arquivo), str(destino_final))
        if callback_progresso:
            callback_progresso(1.0, f"Arquivo enviado: {destino_final.name}")
        return {
            "arquivo_principal": str(destino_final),
            "arquivos_saida": [str(destino_final)],
            "pasta_final": str(pasta_final),
            "mensagem": "Arquivo enviado com sucesso.",
        }

    arquivos_baixados = []
    resultados_processamento = []
    driver = None

    try:
        if precisa_download:
            username, senha_user = buscar_credencial_site(id_usuario_logado, "EBUS")
            if not username or not senha_user:
                if callback_progresso:
                    callback_progresso(0.0, "ERRO: Credenciais do EBUS não encontradas no Cofre!")
                return {
                    "arquivo_principal": None,
                    "arquivos_saida": [],
                    "pasta_final": None,
                    "mensagem": "Credenciais do EBUS não encontradas no Cofre.",
                }

            if callback_progresso:
                callback_progresso(0.1, "Abrindo Navegador Invisível...")

            opcoes = Options()
            opcoes.add_argument("--window-size=1920,1080")
            opcoes.add_argument("--headless")

            driver_path = get_driver_path()
            if driver_path and "chrome" in Path(driver_path).name.lower():
                from selenium.webdriver.chrome.service import Service
                driver = webdriver.Chrome(service=Service(driver_path), options=opcoes)
            else:
                driver = webdriver.Chrome(options=opcoes)

            max_tentativas_download = 2
            baixou_arquivo_valido = False

            for tentativa_download in range(1, max_tentativas_download + 1):
                checar_parada()
                if callback_progresso:
                    callback_progresso(
                        0.12,
                        f"Tentativa {tentativa_download}/{max_tentativas_download}: acessando e higienizando sessão EBUS...",
                    )

                limpar_cookies_ebus(driver, callback_progresso=callback_progresso)
                wait = WebDriverWait(driver, 60)

                if callback_progresso:
                    callback_progresso(0.2, "Fazendo login no EBUS...")

                login = wait.until(EC.presence_of_element_located((By.XPATH, "//input[contains (@id, 'input-usuario')]")))
                login.send_keys(username)
                senha = driver.find_element(By.XPATH, "//input[contains (@id, 'input-senha')]")
                senha.send_keys(senha_user)
                wait.until(EC.element_to_be_clickable((By.XPATH, "//button[contains(text(), 'Login')]"))).click()

                if callback_progresso:
                    callback_progresso(0.3, "Navegando até aba de Revenue...")

                wait.until(EC.element_to_be_clickable((By.XPATH, "//span[contains(@class, 'menu-title ng-tns-c129-33')]"))).click()
                wait.until(EC.element_to_be_clickable((By.XPATH, "//span[contains(@class, 'menu-title ng-tns-c129-37')]"))).click()

                checar_parada()
                if callback_progresso:
                    callback_progresso(0.32, f"Aplicando filtro único no período: {data_inicio} até {data_final}...")

                data_relatorio = wait.until(EC.element_to_be_clickable((By.XPATH, "//input[contains(@placeholder, 'Data Início Viagem')]")))
                data_relatorio.clear()
                data_relatorio.send_keys(data_inicio)
                data_relatorio = wait.until(EC.element_to_be_clickable((By.XPATH, "//input[contains(@placeholder, 'Data Fim Viagem')]")))
                data_relatorio.clear()
                data_relatorio.send_keys(data_final)

                if callback_progresso:
                    callback_progresso(0.35, "Pesquisando período completo... Aguardando tabela.")
                wait.until(EC.element_to_be_clickable((By.XPATH, "//button[contains(text(), ' Pesquisar ')]"))).click()

                try:
                    wait.until(EC.presence_of_element_located((By.XPATH, "//nb-card[contains(@class, 'nb-spinner-container')]")))
                except Exception:
                    pass

                def spinner_disappeared(driver_ref):
                    spinners = driver_ref.find_elements(By.XPATH, "//nb-card[contains(@class, 'nb-spinner-container')]")
                    for spinner in spinners:
                        if spinner.get_attribute("ng-reflect-nb-spinner") == "true":
                            return False
                    return True

                wait.until(spinner_disappeared)
                checar_parada()

                arquivos_antes = {str(p.resolve()) for p in origem_download.rglob("*RelatorioRevenue*.xls*")}

                if callback_progresso:
                    callback_progresso(0.40, "Tabela carregada. Disparando download único do período...")
                wait.until(EC.element_to_be_clickable((By.XPATH, "//button[contains(text(), 'Gerar EXCEL')]"))).click()

                if callback_progresso:
                    callback_progresso(0.45, "Aguardando arquivo do período completo...")

                arquivo_encontrado = None
                limite_espera = time.time() + 120
                while time.time() < limite_espera:
                    checar_parada()
                    candidatos = [
                        p for p in origem_download.rglob("*RelatorioRevenue*.xls*")
                        if str(p.resolve()) not in arquivos_antes
                    ]
                    if candidatos:
                        candidatos.sort(key=os.path.getmtime, reverse=True)
                        arquivo_encontrado = candidatos[0]
                        break
                    time.sleep(1)

                if arquivo_encontrado is None:
                    if callback_progresso:
                        callback_progresso(0.52, "AVISO: arquivo do período completo não foi encontrado.")
                    continue

                if callback_progresso:
                    callback_progresso(0.50, f"Arquivo encontrado: {arquivo_encontrado.name}")

                novo_nome = pasta_trabalho / f"RelatorioRevenue - {datetime.now().strftime('%d.%m.%Y_%H%M%S')}{arquivo_encontrado.suffix}"
                shutil.move(str(arquivo_encontrado), str(novo_nome))

                if data_aplicacao_top10_iguais(novo_nome):
                    if callback_progresso:
                        callback_progresso(
                            0.53,
                            "Validação detectou Data Aplicação repetida nas 10 primeiras linhas. Excluindo e refazendo download com nova limpeza.",
                        )
                    try:
                        os.remove(novo_nome)
                    except OSError:
                        pass
                    continue

                arquivos_baixados.append(novo_nome)
                baixou_arquivo_valido = True
                if callback_progresso:
                    callback_progresso(0.55, f"Arquivo renomeado: {arquivo_encontrado.name} -> {novo_nome.name}")
                break

            if not baixou_arquivo_valido and callback_progresso:
                callback_progresso(0.56, "Nenhum arquivo válido foi obtido após as tentativas automáticas de limpeza e novo download.")

            if modo_execucao == "download":
                arquivos_saida = [str(p) for p in arquivos_baixados]
                pasta_final = pasta_trabalho
                if destino_envio:
                    destino_envio.mkdir(parents=True, exist_ok=True)
                    arquivos_saida = []
                    total_baixados = len(arquivos_baixados)
                    for idx_item, item in enumerate(arquivos_baixados, 1):
                        destino_item = destino_envio / item.name
                        shutil.copy2(str(item), str(destino_item))
                        arquivos_saida.append(str(destino_item))
                        if callback_progresso:
                            callback_progresso(
                                0.9 + (0.09 * (idx_item / max(total_baixados, 1))),
                                f"Arquivo copiado para saída ({idx_item}/{total_baixados}): {destino_item.name}",
                            )
                    pasta_final = destino_envio

                if callback_progresso:
                    callback_progresso(1.0, "Download EBUS concluído com sucesso!")
                return {
                    "arquivo_principal": arquivos_saida[0] if arquivos_saida else None,
                    "arquivos_saida": arquivos_saida,
                    "pasta_final": str(pasta_final),
                    "mensagem": "Download EBUS concluído com sucesso.",
                }

        if not precisa_download and arquivo_entrada:
            origem_manual = Path(arquivo_entrada)
            if not origem_manual.exists():
                raise FileNotFoundError(f"Arquivo de entrada não encontrado: {origem_manual}")
            copia_manual = pasta_trabalho / f"manual_{datetime.now().strftime('%Y%m%d_%H%M%S')}{origem_manual.suffix}"
            shutil.copy2(str(origem_manual), str(copia_manual))
            arquivos_baixados = [copia_manual]
            if callback_progresso:
                callback_progresso(0.5, f"Arquivo manual localizado para tratamento: {origem_manual.name}")

        if precisa_tratamento:
            arquivos_para_processamento = []
            for idx, arquivo in enumerate(arquivos_baixados, 1):
                checar_parada()
                if callback_progresso:
                    callback_progresso(0.55, f"Separando por mês o arquivo {idx}/{len(arquivos_baixados)}: {Path(arquivo).name}")

                arquivos_mes = separar_arquivo_por_mes_data_viagem(
                    arquivo,
                    pasta_trabalho,
                    callback_progresso=callback_progresso,
                )
                arquivos_para_processamento.extend(arquivos_mes)

            for idx, item in enumerate(arquivos_para_processamento, 1):
                checar_parada()
                caminho_arquivo = item["arquivo"]
                nome_mes_item = item.get("nome_mes")
                ano_item = item.get("ano")

                if callback_progresso:
                    callback_progresso(
                        0.62,
                        f"Processando recorte mensal {idx}/{len(arquivos_para_processamento)}: {Path(caminho_arquivo).name}",
                    )

                resultado_proc = processar_arquivos_relatorios(
                    caminho_arquivo,
                    pasta_trabalho,
                    nome_mes=nome_mes_item,
                    ano_atual=ano_item,
                    callback_progresso=callback_progresso,
                    destino_base=str(base_referencia),
                    destino_saida=str(destino_final),
                )
                if resultado_proc:
                    resultados_processamento.append(resultado_proc)

            if not resultados_processamento:
                return {
                    "arquivo_principal": None,
                    "arquivos_saida": [],
                    "pasta_final": str(destino_final),
                    "mensagem": "Nenhum arquivo válido foi processado no EBUS.",
                }

            arquivos_saida = []
            for resultado_proc in resultados_processamento:
                arquivos_saida.extend(resultado_proc.get("arquivos_saida", []))

            arquivo_principal = resultados_processamento[-1].get("arquivo_principal")
            pasta_final = resultados_processamento[-1].get("pasta_final")

            if callback_progresso:
                callback_progresso(1.0, "Processo EBUS concluído com sucesso!")

            return {
                "arquivo_principal": arquivo_principal,
                "arquivos_saida": arquivos_saida,
                "pasta_final": pasta_final,
                "mensagem": "Processo EBUS concluído com sucesso.",
            }

        if callback_progresso:
            callback_progresso(1.0, "Nenhuma ação executada.")
        return {
            "arquivo_principal": None,
            "arquivos_saida": [],
            "pasta_final": None,
            "mensagem": "Nada para executar com a combinação selecionada.",
        }

    except CanceladoPeloUsuario as erro_cancel:
        if callback_progresso:
            callback_progresso(0, str(erro_cancel))
        raise

    finally:
        if driver is not None:
            try:
                driver.quit()
            except Exception:
                pass

# ==========================================
# EXECUÇÃO VIA CLI/BACKEND
# ==========================================
if __name__ == '__main__':
    import json
    import sys
    import base64

    # Ler parâmetros de CLI (Base64 via argv) ou fallback para STDIN
    try:
        if len(sys.argv) > 1:
            params = json.loads(base64.b64decode(sys.argv[1]))
        else:
            line = sys.stdin.readline()
            params = json.loads(line) if line else {}
    except:
        params = {}

    user_id = params.get('user_id', 1)
    
    # Formatos de data
    def fix_date(d):
        if not d: return d
        if 'T' in d: # ISO
            return datetime.fromisoformat(d.replace('Z', '')).strftime('%d/%m/%Y')
        return d

    data_ini = fix_date(params.get('data_ini')) or (datetime.now().strftime('%d/%m/%Y'))
    data_fim = fix_date(params.get('data_fim')) or data_ini

    def progress_callback(p, m):
        print(f'PROGRESS:{{"p": {int(p*100)}, "m": "{m}"}}', flush=True)

    try:
        resultado = executar_ebus(
            id_usuario_logado=user_id,
            data_inicio=data_ini,
            data_final=data_fim,
            callback_progresso=progress_callback,
            modo_execucao=params.get('acao', 'completo'),
            pasta_destino=params.get('pasta_saida'),
            arquivo_entrada=params.get('pasta_personalizada'),
            base_automacao=params.get('base')
        )
        print(json.dumps(resultado))
    except Exception as e:
        print(f"ERRO: {str(e)}", file=sys.stderr)
        sys.exit(1)
