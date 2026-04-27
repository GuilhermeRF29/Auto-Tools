import os
import sys
import json
import base64
import shutil
from playwright.sync_api import sync_playwright
import time
import re
import pandas as pd
import pprint
import openpyxl
from datetime import datetime
from datetime import timedelta
from copy import copy
from pathlib import Path

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from core.banco import buscar_credencial_site

print("PROGRESS:{\"p\": 1, \"m\": \"Carregando Modulos Busca Dados...\"}", flush=True)

DEFAULT_BASE_DIR = Path(r"G:\Drives compartilhados\Cometa   Comercial\COM_Comercial\05 - RM\Share Canais\Performance de Canais - Atualização")
DEFAULT_TEMPLATE_FILE = "PERFORMANCE DE CANAIS_COMPARATIVO YoY.xlsx"
DEFAULT_REPORT_NAME = "Busca Dados BI"
DEFAULT_CRED_SERVICE = "Busca Dados BI"
DEFAULT_TEMP_DIR = Path.home() / "Downloads" / "tmp_busca_dados"

MESES_PTBR = {
    1: "01-Janeiro", 2: "02-Fevereiro", 3: "03-Março", 4: "04-Abril",
    5: "05-Maio", 6: "06-Junho", 7: "07-Julho", 8: "08-Agosto",
    9: "09-Setembro", 10: "10-Outubro", 11: "11-Novembro", 12: "12-Dezembro",
}

SIGLAS_PTBR = {
    1: "JAN", 2: "FEV", 3: "MAR", 4: "ABR", 5: "MAI", 6: "JUN",
    7: "JUL", 8: "AGO", 9: "SET", 10: "OUT", 11: "NOV", 12: "DEZ",
}


def parse_frontend_date(valor):
    if not valor:
        return None

    if isinstance(valor, datetime):
        return valor

    texto = str(valor).strip()
    if not texto:
        return None

    formatos = [
        "%d/%m/%Y",
        "%Y-%m-%d",
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%dT%H:%M:%S.%f",
    ]

    for fmt in formatos:
        try:
            return datetime.strptime(texto.replace("Z", ""), fmt)
        except Exception:
            continue

    try:
        return datetime.fromisoformat(texto.replace("Z", ""))
    except Exception:
        return None


def resolver_diretorio_base(base_automacao=None, pasta_personalizada=None):
    base_raw = (str(base_automacao).strip() if base_automacao is not None else "")
    base_lower = base_raw.lower()

    if base_lower == "personalizada" and pasta_personalizada:
        caminho = Path(str(pasta_personalizada).strip())
        return caminho if caminho.is_dir() else caminho.parent

    if base_raw and base_lower not in {"padrao", "sem_base", "none", "null", "personalizada"}:
        caminho = Path(base_raw)
        return caminho if caminho.is_dir() else caminho.parent

    return DEFAULT_BASE_DIR


def resolver_diretorio_saida(saida_automacao=None, pasta_saida=None, base_dir=None):
    saida_raw = (str(saida_automacao).strip() if saida_automacao is not None else "")
    saida_lower = saida_raw.lower()

    if saida_lower == "personalizada" and pasta_saida:
        return Path(str(pasta_saida).strip())

    if base_dir is not None:
        return Path(base_dir)

    return DEFAULT_BASE_DIR


def gerar_segmentos_mensais(data_inicio: datetime, data_fim: datetime):
    atual = datetime(data_inicio.year, data_inicio.month, 1)
    limite = datetime(data_fim.year, data_fim.month, 1)
    segmentos = []

    while atual <= limite:
        prox_mes = datetime(atual.year + (1 if atual.month == 12 else 0), 1 if atual.month == 12 else atual.month + 1, 1)
        fim_mes = prox_mes - timedelta(days=1)

        inicio_segmento = max(data_inicio, atual)
        fim_segmento = min(data_fim, fim_mes)

        if inicio_segmento <= fim_segmento:
            dias = [str(dia) for dia in range(inicio_segmento.day, fim_segmento.day + 1)]
            segmentos.append({
                "ano": inicio_segmento.year,
                "mes": inicio_segmento.month,
                "mes_label": MESES_PTBR[inicio_segmento.month],
                "dias": dias,
                "ultimo_dia": str(fim_segmento.day).zfill(2),
            })

        atual = prox_mes

    return segmentos

def atualizar_planilha_comparativa(caminho_extracao, caminho_template, mes_sigla="ABR", ultimo_dia="01", ano_referencia=2026):
    
    if not os.path.exists(caminho_extracao) or not os.path.exists(caminho_template):
        raise FileNotFoundError("Arquivos base (Extracao ou Gabarito YoY) nao encontrados.")
        
    print(f"Lendo extrato da automação: {caminho_extracao}...")
    df = pd.read_excel(caminho_extracao)
    df['Ano'] = df['Ano'].astype(str)
    
    print(f"Acessando Motor de Gravação Segura do Excel para: {caminho_template}...")
    wb = openpyxl.load_workbook(caminho_template, data_only=False)
    
    ano_ref = int(ano_referencia)
    nome_aba_alvo = f"COMAPRATIVO {str(ano_ref - 1)[-2:]}X{str(ano_ref)[-2:]} {mes_sigla.upper()}"
    
    if nome_aba_alvo not in wb.sheetnames:
        print(f"Mês base {nome_aba_alvo} detectado ausente. Duplicando histórico p/ criação...")
        aba_referencia = wb[wb.sheetnames[-1]]
        nova_aba = wb.copy_worksheet(aba_referencia)
        nova_aba.title = nome_aba_alvo
        ws = nova_aba
    else:
        ws = wb[nome_aba_alvo]
        print(f"Aba do Mês encontrada. Injetando em {nome_aba_alvo}...")

    mapa_meses = {"JAN": "01", "FEV": "02", "MAR": "03", "ABR": "04", 
                  "MAI": "05", "JUN": "06", "JUL": "07", "AGO": "08", 
                  "SET": "09", "OUT": "10", "NOV": "11", "DEZ": "12"}
    num_mes = mapa_meses.get(mes_sigla.upper(), "01")
    
    # Data ancorada no filtro PBI invés do relógio do Windows
    data_alvo_str = f"{ultimo_dia}/{num_mes}/{ano_ref}"
    ws['B2'] = f"Atualizado até dia - {data_alvo_str}"

    def extrair_v_base(df_slice, nome_categoria):
        reg = df_slice[df_slice["Categoria"] == nome_categoria]
        if len(reg) > 0:
            return reg.iloc[0]
        return None

    mapa_geral_linhas_canais = {
        "Guichê próprio": (5, 25, 44), "Guichê terceiro": (6, 26, 45), 
        "Venda corporativa": (7, 27, 46), "Venda embarcada": (8, 28, 47), 
        "Sites próprios": (9, 29, 48), "Sites terceiros": (10, 30, 49), 
        "Outlet de passagens": (11, 31, 50), "Clube giro": (12, 32, 51), 
        "Wemobi": (13, 33, 52), "Totem": (14, 34, 53), "Gipsyy": (15, 35, 54)
    }

    ano_passado = str(ano_ref - 1)
    ano_atual = str(ano_ref)
    df_ano_passado = df[df["Ano"] == ano_passado]
    df_ano_atual = df[df["Ano"] == ano_atual]

    print("Iniciando Transplante Reverso de Receitas, TM e Volume...")
    for categ, coordenadas in mapa_geral_linhas_canais.items():
        linha_fin = coordenadas[0]; linha_tm = coordenadas[1]; linha_pax = coordenadas[2]
        
        dado_passado = extrair_v_base(df_ano_passado, categ)
        dado_atual = extrair_v_base(df_ano_atual, categ)
        
        if dado_passado is not None:
            ws[f'C{linha_fin}'].value = dado_passado['Realizado Geral']
            ws[f'C{linha_tm}'].value = dado_passado['TM Geral']
            ws[f'C{linha_pax}'].value = dado_passado['Passagens/Mix']
            
        if dado_atual is not None:
            ws[f'D{linha_fin}'].value = dado_atual['Orçado (Fórmula)']
            ws[f'E{linha_fin}'].value = dado_atual['Realizado Geral']
            ws[f'D{linha_tm}'].value = dado_atual['TM Geral']
            ws[f'D{linha_pax}'].value = dado_atual['Passagens/Mix']

        for célula_pax in [ws[f'C{linha_pax}'], ws[f'D{linha_pax}']]:
            if célula_pax.font:
                fnt = copy(célula_pax.font)
                fnt.underline = None
                célula_pax.font = fnt

    linha_global_passado = extrair_v_base(df_ano_passado, "Mês Global Padrão")
    linha_global_atual = extrair_v_base(df_ano_atual, "Mês Global Padrão")
    
    if linha_global_passado is not None:
        ws['C36'].value = linha_global_passado['TM Geral']; ws['C39'].value = linha_global_passado['TM Offline']; ws['C40'].value = linha_global_passado['TM Online']
        
    if linha_global_atual is not None:
        ws['D36'].value = linha_global_atual['TM Geral']; ws['D39'].value = linha_global_atual['TM Offline']; ws['D40'].value = linha_global_atual['TM Online']

    linhas_varredura = list(range(5, 17)) + list(range(24, 42)) + list(range(43, 58))
    for linha in linhas_varredura:
        for col_idx in range(3, 18):
            celula = ws.cell(row=linha, column=col_idx)
            fmt = celula.number_format
            if fmt and '[Red]' not in fmt:
                if '%' in fmt.lower() or 'p' in fmt.lower() or 'percent' in str(type(celula.value)).lower():
                    celula.number_format = '0.00%;[Red]-0.00%'
                elif isinstance(celula.value, (int, float)) or celula.data_type == 'f':
                    if linha >= 43:
                        celula.number_format = '#,##0;[Red]-#,##0'
                    else:
                        celula.number_format = '#,##0.00;[Red]-#,##0.00'

    print("Campos preenchidos. Salvando sistema reativo de Layout...")
    try:
        wb.save(caminho_template)
        print(f" >>> [CONCLUÍDO] Sucesso absoluto! Verifique o arquivo: {caminho_template}")
    except PermissionError:
        print("\n [ERRO] O Excel alvo recusa modificações. Ele está ABERTO em seu computador? Feche a janela e tente de novo.")

def limpar_valor(texto):
    """Remove espaços, R$ e caracteres não-numéricos para comparação segura."""
    if not texto: return "N/A"
    t_lower = texto.lower()
    if "n/a" in t_lower or "branco" in t_lower or "blank" in t_lower: 
        return "N/A"
    val = re.sub(r'[^0-9,.-]', '', texto).strip()
    return val if val else "N/A"

def extrair_numero(valor_texto):
    """Converte 'R$ 25.543.070,50' para 25543070.50 como float real."""
    v = limpar_valor(valor_texto)
    if v in ["N/A", ""]: return 0.0
    # Remove pontos de milhar, troca vírgula decimal por ponto
    v = v.replace('.', '').replace(',', '.')
    try:
        return float(v)
    except:
        return 0.0

def run(
    mes_para_selecionar="04-Abril",
    dias_para_selecionar=None,
    anos_para_processar=None,
    caminho_extracao=None,
    caminho_template=None,
    ano_referencia=2026,
    credenciais=None,
    callback_progresso=None,
):
    # --- CONFIGURAÇÃO DE INPUTS ---
    dias_para_selecionar = dias_para_selecionar or [str(i) for i in range(1, 13)]
    anos_para_processar = anos_para_processar or ["2026", "2025"]
    caminho_extracao = str(caminho_extracao or (Path.cwd() / "analise_relatorio_performance.xlsx"))

    login_usuario = None
    senha_usuario = None
    if credenciais and isinstance(credenciais, (list, tuple)) and len(credenciais) == 2:
        login_usuario, senha_usuario = credenciais

    if not login_usuario or not senha_usuario:
        raise ValueError("Credenciais de acesso ao BI nao foram informadas.")

    if callback_progresso:
        callback_progresso(0.02, f"Preparando extracao BI: {mes_para_selecionar} ({dias_para_selecionar[0]}-{dias_para_selecionar[-1]})")
    
    # 11 Grupos Mapeados (Série Histórica e Fechamento)
    grupos_canais = {
        "Guichê próprio": ["AG. PRÓPRIA", "AGÊNCIA", "AGENCIA MOVEL", "CALL CENTER", "GARAGEM", "RELACIONAMENTO COM O CLIENTE", "TERMINAL URBANO"],
        "Guichê terceiro": ["AG. TERCEIRA", "ND"],
        "Venda corporativa": ["VENDA CORPORATIVA"],
        "Venda embarcada": ["VENDA EMBARCADA"],
        "Sites próprios": ["INTERNET", "INTERNET PRÓPRIA"],
        "Sites terceiros": ["J3"],
        "Outlet de passagens": ["OUTLET"],
        "Clube giro": ["CLUBE GIRO"],
        "Wemobi": ["APLICATIVO WEMOBI", "WEMOBI"],
        "Totem": ["TOTEM"],
        "Gipsyy": ["GIPSYY"]
    }

    # Proporções Matemáticas Dinâmicas por Mês para o Orçado (Rateio)
    percentuais_orcado_matriz = {
        "Guichê próprio": {
            "01-Janeiro": 0.297898068, 
            "02-Fevereiro": 0.306609382, 
            "03-Março": 0.305845086, 
            "04-Abril": 0.29477968403785700, 
            "05-Maio": 0.272309885, 
            "06-Junho": 0.275790669, 
            "07-Julho": 0.279996161, 
            "08-Agosto": 0.2723831, 
            "09-Setembro": 0.268500247, 
            "10-Outubro": 0.264597326, 
            "11-Novembro": 0.250145235, 
            "12-Dezembro": 0.2584493},
        "Guichê terceiro": {
            "01-Janeiro": 0.094497577, 
            "02-Fevereiro": 0.093297763, 
            "03-Março": 0.090158333, 
            "04-Abril": 0.08636660324305720, 
            "05-Maio": 0.08924162, 
            "06-Junho": 0.087848823, 
            "07-Julho": 0.090925786, 
            "08-Agosto": 0.0888922, 
            "09-Setembro": 0.08474676, 
            "10-Outubro": 0.083753732, 
            "11-Novembro": 0.080417732, 
            "12-Dezembro": 0.0765662},
        "Venda corporativa": {
            "01-Janeiro": 0.004354097, 
            "02-Fevereiro": 0.005123234, 
            "03-Março": 0.005620111, 
            "04-Abril": 0.00440764246674112, 
            "05-Maio": 0.005096693, 
            "06-Junho": 0.005541518, 
            "07-Julho": 0.004232944, 
            "08-Agosto": 0.005904, 
            "09-Setembro": 0.004992371, 
            "10-Outubro": 0.004845715, 
            "11-Novembro": 0.004845715, 
            "12-Dezembro": 0.0048457},
        "Venda embarcada": {
            "01-Janeiro": 0.003601542, 
            "02-Fevereiro": 0.004563394, 
            "03-Março": 0.004527487, 
            "04-Abril": 0.00397734645879112, 
            "05-Maio": 0.005460637, 
            "06-Junho": 0.005131232, 
            "07-Julho": 0.004657636, 
            "08-Agosto": 0.0060163, 
            "09-Setembro": 0.005915509, 
            "10-Outubro": 0.00541492, 
            "11-Novembro": 0.00541492, 
            "12-Dezembro": 0.0054149},
        "Sites próprios": {
            "01-Janeiro": 0.297, 
            "02-Fevereiro": 0.274, 
            "03-Março": 0.285, 
            "04-Abril": 0.29000000000000000, 
            "05-Maio": 0.303, 
            "06-Junho": 0.305, 
            "07-Julho": 0.303, 
            "08-Agosto": 0.301, 
            "09-Setembro": 0.306, 
            "10-Outubro": 0.312, 
            "11-Novembro": 0.323, 
            "12-Dezembro": 0.326},
        "Sites terceiros": {
            "01-Janeiro": 0.230432094, 
            "02-Fevereiro": 0.231403385, 
            "03-Março": 0.232196137, 
            "04-Abril": 0.23485758280823100, 
            "05-Maio": 0.23439076, 
            "06-Junho": 0.231335214, 
            "07-Julho": 0.22282375, 
            "08-Agosto": 0.2288473, 
            "09-Setembro": 0.232761443, 
            "10-Outubro": 0.239734765, 
            "11-Novembro": 0.237625658, 
            "12-Dezembro": 0.2391684},
        "Outlet de passagens": {
            "01-Janeiro": 0.014067286, 
            "02-Fevereiro": 0.019094536, 
            "03-Março": 0.01005431, 
            "04-Abril": 0.01752322486302950, 
            "05-Maio": 0.013866524, 
            "06-Junho": 0.015263575, 
            "07-Julho": 0.017029231, 
            "08-Agosto": 0.0138705, 
            "09-Setembro": 0.01406217, 
            "10-Outubro": 0.012236937, 
            "11-Novembro": 0.015396373, 
            "12-Dezembro": 0.0133792},
        "Clube giro": {
            "01-Janeiro": 0.024493415, 
            "02-Fevereiro": 0.030049285, 
            "03-Março": 0.034017978, 
            "04-Abril": 0.03292920606589090,
            "05-Maio": 0.039388321, 
            "06-Junho": 0.035232906, 
            "07-Julho": 0.031205395, 
            "08-Agosto": 0.0390799, 
            "09-Setembro": 0.039597813, 
            "10-Outubro": 0.032004985, 
            "11-Novembro": 0.036895973, 
            "12-Dezembro": 0.0315744},
        "Wemobi": {
            "01-Janeiro": 0.015437723, 
            "02-Fevereiro": 0.017640825, 
            "03-Março": 0.014332593, 
            "04-Abril": 0.01440171569479450, 
            "05-Maio": 0.016469063, 
            "06-Junho": 0.018076838, 
            "07-Julho": 0.02285262, 
            "08-Agosto": 0.0207286, 
            "09-Setembro": 0.020131599, 
            "10-Outubro": 0.019618475, 
            "11-Novembro": 0.020442873, 
            "12-Dezembro": 0.018764},
        "Totem": {
            "01-Janeiro": 0.0175, 
            "02-Fevereiro": 0.0175, 
            "03-Março": 0.0175, 
            "04-Abril": 0.02000000000000000, 
            "05-Maio": 0.02, 
            "06-Junho": 0.02, 
            "07-Julho": 0.0225, 
            "08-Agosto": 0.0225, 
            "09-Setembro": 0.0225, 
            "10-Outubro": 0.025, 
            "11-Novembro": 0.025, 
            "12-Dezembro": 0.025},
        "Gipsyy": {
            "01-Janeiro": 0.000718197, 
            "02-Fevereiro": 0.000718197, 
            "03-Março": 0.000747965, 
            "04-Abril": 0.00075699436160802, 
            "05-Maio": 0.000776497, 
            "06-Junho": 0.000779224, 
            "07-Julho": 0.000776478, 
            "08-Agosto": 0.000778, 
            "09-Setembro": 0.000792088, 
            "10-Outubro": 0.000793145, 
            "11-Novembro": 0.000815521, 
            "12-Dezembro": 0.0008379}
    }
    
    # Lista global que alimentará o arquivo Excel final
    dados_tabela_excel = []
    
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False, args=["--start-maximized"])
        context = browser.new_context(no_viewport=True)
        page = context.new_page()

        # 2. Acessar o site
        url = "https://app.powerinsight.com.br/relatorios/groups/0a7a5306-246f-45ea-81ac-3f3b81d7eb10/reports/fa328b2d-c4a1-47d7-a2a2-1e52954bca95?navpath=5778&repBdId=5113"
        print(f"Acessando: {url}")
        page.goto(url)

        # --- FUNÇÕES DE SUPORTE (MOTOR DO PLAYWRIGHT) ---

        def esperar_carregamento(frame):
            try:
                spinner = frame.locator(".powerbi-spinner.shown, .powerbi-spinner.xsmall.shown")
                spinner.first.wait_for(state="hidden", timeout=15000)
            except:
                pass

        def force_click(locator, modifiers=None):
            is_ctrl = 'true' if (modifiers and "Control" in modifiers) else 'false'
            # Avalia JS injetando Ctrl se necessário e dispara o evento ignorando saltos de canvas
            locator.evaluate(f"el => {{"
                             f"el.dispatchEvent(new PointerEvent('pointerdown', {{bubbles: true, ctrlKey: {is_ctrl}}}));"
                             f"el.dispatchEvent(new PointerEvent('pointerup', {{bubbles: true, ctrlKey: {is_ctrl}}}));"
                             f"el.dispatchEvent(new MouseEvent('click', {{bubbles: true, ctrlKey: {is_ctrl}}}));"
                             f"}}")

        def clicar_com_scroll(frame, texto, modifiers=None):
            print(f"Buscando item: {texto}...")
            # Encontra apenas spans visíveis com o texto exato
            target = frame.locator(f'span:text-is("{texto}") >> visible=true').first
            
            # Reposiciona o mouse fisicamente no centro do Menu Suspenso SEM forçar "ScrollIntoView" 
            # para evitar que o Power BI dê aquele pulo lateral
            try:
                caixa = frame.locator('.slicerBody >> visible=true').first
                box = caixa.bounding_box()
                if box:
                    page.mouse.move(box["x"] + box["width"]/2, box["y"] + box["height"]/2)
                time.sleep(0.3)
            except:
                pass
            
            for i in range(12): 
                if target.is_visible():
                    print(f"Item '{texto}' encontrado! Clicando...")
                    force_click(target, modifiers)
                    return True
                for _ in range(3):
                    page.mouse.wheel(0, 60)
                    time.sleep(0.1)
                time.sleep(0.4)
            
            print(f"Item '{texto}' não encontrado descendo. Resetando scroll para o topo...")
            for _ in range(5):
                 page.mouse.wheel(0, -1200)
                 time.sleep(0.1)
            time.sleep(0.5)
            
            for i in range(20): 
                if target.is_visible():
                    print(f"Item '{texto}' encontrado após reset! Clicando...")
                    force_click(target, modifiers)
                    return True
                for _ in range(3):
                    page.mouse.wheel(0, 60)
                    time.sleep(0.1)
                time.sleep(0.4)
            
            try:
                force_click(target, modifiers)
                return True
            except:
                print(f"Falha ao forçar clique em {texto}")
                return False

        def resetar_multiselecao(frame, descricao_slicer):
            """Normaliza a multiseleção com 'Selecionar tudo' (marca e desmarca)."""
            print(f"Normalizando seleção do slicer '{descricao_slicer}' (Selecionar tudo 2x)...")
            opcoes_select_all = [
                "Selecionar tudo",
                "Selecionar Tudo",
                "SELECIONAR TUDO",
                "Selecionar todos",
                "Selecionar Todos",
                "SELECIONAR TODOS",
            ]

            for etapa in range(2):
                clicou = False
                for texto in opcoes_select_all:
                    if clicar_com_scroll(frame, texto, modifiers=None):
                        clicou = True
                        break

                if not clicou:
                    print(
                        f"[WARN] Não foi possível acionar 'Selecionar tudo' no slicer '{descricao_slicer}' "
                        f"(etapa {etapa + 1}/2). Prosseguindo com o fluxo atual."
                    )
                    return False

                time.sleep(0.35)

            return True

        def visual_click(locator):
            """Mantido para cliques de gráficos de pizza."""
            force_click(locator, modifiers=None)

        def extrair_valor_bi(frame, label_parcial):
            try:
                seletor_svg = f'svg[aria-label*="{label_parcial}"]'
                valor_elemento = frame.locator(seletor_svg).locator('tspan').first
                val = valor_elemento.text_content().strip()
                return val if val else "N/A"
            except:
                return "N/A"

        def obter_valor_base_valido(frame, label):
            """Garante um start point limpo não-nulo para monitorar atualizações numéricas de dashboards lentas."""
            for _ in range(60): # Aumentado timeout global para 60 segundos 
                val = extrair_valor_bi(frame, label)
                if limpar_valor(val) not in ["N/A", ""]:
                    return val
                time.sleep(1)
            return extrair_valor_bi(frame, label)

        def extrair_valor_garantido(frame, label, valor_base):
            """Validação Inteligente contra Falsos-Positivos Temporais."""
            v_base_clean = limpar_valor(valor_base)
            print(f"[{label}] Aguardando processamento. Valor de referência: '{valor_base}'")
            
            time.sleep(1.5) 
            
            for i in range(120): # Aumentado de 30 para 120 repetições garantindo suportes a demoras massivas  (aprox 60 seg)
                novo_valor = extrair_valor_bi(frame, label)
                v_novo_clean = limpar_valor(novo_valor)
                if v_novo_clean != "N/A" and v_novo_clean != "":
                    if v_novo_clean != v_base_clean:
                        print(f"[{label}] Sucesso! Atualizou no servidor para: {novo_valor}")
                        return novo_valor
                time.sleep(0.5)
   
            print(f"[{label}] ALERTA: O tempo limite expirou. Retornando valor: {novo_valor}")
            return extrair_valor_bi(frame, label)


        # --- EXECUÇÃO PRINCIPAL ---

        try:
            # Login
            print("Realizando login...")
            if callback_progresso:
                callback_progresso(0.08, "Realizando autenticacao no BI...")
            time.sleep(2)
            page.wait_for_selector("//input[contains(@id, 'mat-input-0')]", state="visible", timeout=30000)
            page.fill("//input[contains(@id, 'mat-input-0')]", login_usuario)
            page.wait_for_selector("//input[contains(@id, 'mat-input-1')]", state="visible")
            page.fill("//input[contains(@id, 'mat-input-1')]", senha_usuario)
            page.click("//span[contains(text(), ' Entrar ')]")
            print("Login realizado!")
            if callback_progresso:
                callback_progresso(0.12, "Login concluido. Carregando dashboard...")

        except Exception as e:
            raise RuntimeError(f"Erro ao realizar login: {e}")

        # Entrando na Sessão do Power BI
        try:
            print("Aguardando montagem do Dashboard inicial...")
            dashboard_frame = page.frame_locator("iframe").first
            time.sleep(15) 

            # --- NAVEGAÇÃO PARA ABA R$ HORA ---
            print("Navegando para a aba 'R$ Hora'...")
            
            # 1. Clicar na seta esquerda até o início do carrossel
            seletor_seta_esq = 'button[data-testid="carousel-previous-page"]'
            for _ in range(10): # Limite de segurança
                seta_esq = dashboard_frame.locator(seletor_seta_esq)
                if seta_esq.is_visible() and not seta_esq.is_disabled():
                    print("Voltando carrossel de abas...")
                    seta_esq.click()
                    time.sleep(0.8)
                else:
                    break
            
            # 2. Selecionar a aba "R$ Hora"
            seletor_dashboard_rhoras = 'div[data-testid="section"]'
            selecao_dashboard_rhoras = dashboard_frame.locator(seletor_dashboard_rhoras).filter(has_text="R$ Hora")
            
            if selecao_dashboard_rhoras.count() > 0:
                print("Clicando na aba 'R$ Hora'...")
                selecao_dashboard_rhoras.first.click()
            else:
                print("Aba não encontrada pelo filtro principal, tentando via label de texto...")
                try:
                    dashboard_frame.locator('div.textLabel:has-text("R$ Hora")').first.click()
                except:
                    print("Aviso: Não foi possível clicar na aba 'R$ Hora'.")
            
            time.sleep(3) 
            esperar_carregamento(dashboard_frame)
            
            # O Snapshot virgem será tirado agora na aba correta.

            # Pega Snapshot do Painel Virgem (antes de tocar em nada)
            # Dessa forma podemos forçar o sistema a sentir a Primeira filtragem de forma rastreável!
            valor_virgem_orcado = obter_valor_base_valido(dashboard_frame, "Valor Orçado")
            valor_virgem_realizado = obter_valor_base_valido(dashboard_frame, "Valor Realizado")
            print(f"Estado INICIAL do Dashboard (Pré-Filtros) -> Orçado: {valor_virgem_orcado} | Realizado: {valor_virgem_realizado}")

            # Variáveis Cascata entre Anos
            referencia_ano_anterior_orcado = "N/A"
            referencia_ano_anterior_realizado = "N/A"
            dias_ja_configurados = False

            # LOOP GIGANTE POR ANO (Ciclando 2X = Rodada 1 para Caching, Rodada 2 Extrativa Real)
            ciclo_de_anos = anos_para_processar + anos_para_processar
            for iteracao, ano_atual in enumerate(ciclo_de_anos):
                if iteracao == len(anos_para_processar):
                    print("\n" + "*"*50)
                    print("INICIANDO 2ª RODADA (UTILIZANDO CACHE DO POWER BI FORÇADO)")
                    print("Limpando dados provisórios da Rodada 1 para extração real...")
                    print("*"*50)
                    dados_tabela_excel.clear() # Descarta o 1º ciclo
                    
                print(f"\n==========================================")
                print(f"INICIANDO EXTRAÇÃO DE DADOS PARA O ANO: {ano_atual} (Passo {iteracao+1}/{len(ciclo_de_anos)})")
                print(f"==========================================")
                
                # Seleção de Ano
                menu_ano_dropdown = dashboard_frame.locator('div[role="combobox"][aria-label="Ano"]')
                texto_ano_atual = menu_ano_dropdown.text_content() or ""
                
                if ano_atual not in texto_ano_atual:
                    print(f"Ajustando Ano para: {ano_atual}...")
                    menu_ano_dropdown.click()
                    time.sleep(0.5)
                    
                    try: 
                        dashboard_frame.locator("span").get_by_text(ano_atual, exact=True).click()
                    except:
                        clicar_com_scroll(dashboard_frame, ano_atual, modifiers=None)
                        
                    time.sleep(0.5)
                    menu_ano_dropdown.click() 
                    esperar_carregamento(dashboard_frame)
                else:
                    print(f"O Ano {ano_atual} já consta como selecionado!")

                # Seleção de Mês Variável
                menu_mes_dropdown = dashboard_frame.locator('div[role="combobox"][aria-label="Mês Formatado"]')
                texto_mes_atual = menu_mes_dropdown.text_content() or ""
                
                if mes_para_selecionar not in texto_mes_atual:
                    print(f"Limpando/Selecionando Mês ({mes_para_selecionar})...")
                    menu_mes_dropdown.click()
                    time.sleep(0.5)
                    resetar_multiselecao(dashboard_frame, "Mês Formatado")
                    time.sleep(0.3)
                    clicar_com_scroll(dashboard_frame, mes_para_selecionar, modifiers=None)
                    time.sleep(0.5)
                    menu_mes_dropdown.click() # Fecha menu
                    esperar_carregamento(dashboard_frame)
                else:
                    print(f"Mês {mes_para_selecionar} já consta como selecionado!")

                # Seleção de Dias (Até ontem)
                if not dias_ja_configurados:
                    print(f"Configurando escopo de dias de fechamento: {dias_para_selecionar}")
                    menu_dia = dashboard_frame.locator('div[role="combobox"][aria-label="Dia"]')
                    menu_dia.wait_for(state="visible", timeout=30000)
                    menu_dia.click()
                    time.sleep(1.5) # Respiro extra para garantir que a animação da caixa completou
                    resetar_multiselecao(dashboard_frame, "Dia")
                    time.sleep(0.5)
                    
                    # Com a seleção limpa, marca o primeiro e soma os demais com Ctrl
                    clicar_com_scroll(dashboard_frame, dias_para_selecionar[0], modifiers=None) 
                    time.sleep(0.8) # Delay estendido apenas no primeiro
                    for dia in dias_para_selecionar[1:]:
                        clicar_com_scroll(dashboard_frame, dia, modifiers=["Control"])
                        time.sleep(0.3)
                    menu_dia.click() # Recolhe
                    esperar_carregamento(dashboard_frame)
                    dias_ja_configurados = True
                else:
                    print(f"Dias base globais {dias_para_selecionar} mantidos ativamente da interação anterior!")

                # -----------------------------------------------------
                # COLETA 1: TOTAL GLOBAL & TICKET MÉDIO (SEM FILTROS CANAIS)
                print(f"\n[{ano_atual}] LENDO TOTAIS FINANCEIROS...")
                
                if referencia_ano_anterior_orcado == "N/A":
                    # Primeiro Ano: Garante que os números MUDARAM em relação à tela inicial "Virgem" do Power BI!
                    print(f"Validando Filtro Inicial: Forçando mutação do Orç. original {valor_virgem_orcado}")
                    taxa_str_orcado = extrair_valor_garantido(dashboard_frame, "Valor Orçado", valor_virgem_orcado)
                    
                    print(f"Validando Filtro Inicial: Forçando mutação do Real. original {valor_virgem_realizado}")
                    taxa_str_realizado = extrair_valor_garantido(dashboard_frame, "Valor Realizado", valor_virgem_realizado)
                else:
                    # Ano(s) Seguinte(s): Garante que O VALOR MUDOU em relação ao Ano Anterior 
                    # Isso cria uma trava de segurança absoluta contra a lentidão da virada de Ano!
                    print(f"Validando Virada de Ano: Orçado anterior era {referencia_ano_anterior_orcado}")
                    taxa_str_orcado = extrair_valor_garantido(dashboard_frame, "Valor Orçado", referencia_ano_anterior_orcado)
                    
                    print(f"Validando Virada de Ano: Realizado anterior era {referencia_ano_anterior_realizado}")
                    taxa_str_realizado = extrair_valor_garantido(dashboard_frame, "Valor Realizado", referencia_ano_anterior_realizado)
                    
                # Substitui as referências para o próximo ciclo de loop enxergar como base
                referencia_ano_anterior_orcado = taxa_str_orcado
                referencia_ano_anterior_realizado = taxa_str_realizado
                
                time.sleep(1) # Carga paralela de outros Cards geralmente termina quase instantânea a essa altura
                
                taxa_str_tm_total = obter_valor_base_valido(dashboard_frame, "Fin Ticket Médio")
                bilhetes_total = obter_valor_base_valido(dashboard_frame, "Total Passagens")
                
                v_total_orcado_num = extrair_numero(taxa_str_orcado)
                v_total_real_num = extrair_numero(taxa_str_realizado)
                
                # Ticket Médio Pivot Actions (Pizza O/O)
                print(f"\n[{ano_atual}] ISOLANDO TICKET MÉDIO DA PIZZA O/O...")
                grafico_online = dashboard_frame.locator('path[aria-label*="Venda Online."]')
                grafico_offline = dashboard_frame.locator('path[aria-label*="Venda OffLine."]')

                # Extract Online
                grafico_online.wait_for(state="visible")
                visual_click(grafico_online)
                val_tm_online = extrair_valor_garantido(dashboard_frame, "Fin Ticket Médio", taxa_str_tm_total)

                # Extract Offline
                grafico_offline.wait_for(state="visible")
                visual_click(grafico_offline)
                val_tm_offline = extrair_valor_garantido(dashboard_frame, "Fin Ticket Médio", val_tm_online)

                # Limpa Filtro O/O
                visual_click(grafico_offline)
                time.sleep(1.5)
                esperar_carregamento(dashboard_frame)
                
                # Salvando o Total Resumo Historico
                dados_tabela_excel.append({
                    "Ano": ano_atual,
                    "Visão": "RESUMO TOTAL",
                    "Categoria": "Mês Global Padrão",
                    "Orçado (Fórmula)": v_total_orcado_num,
                    "Realizado Geral": v_total_real_num,
                    "Passagens/Mix": extrair_numero(bilhetes_total),
                    "TM Geral": extrair_numero(taxa_str_tm_total),
                    "TM Online": extrair_numero(val_tm_online),
                    "TM Offline": extrair_numero(val_tm_offline)
                })

                # -----------------------------------------------------
                # COLETA 2: LOOP DOS 11 GRUPOS DE CANAIS
                print(f"\n[{ano_atual}] INICIANDO VARREDURA PROFUNDA DE CANAIS ({len(grupos_canais)} Mapeados)...")
                menu_canal = dashboard_frame.locator('div[role="combobox"][aria-label="Canal"]')
                
                # Vamos usar o valor total (Capturado lá atrás) como primeira base de comparação para os canais!
                referencia_valida_realizado = taxa_str_realizado
                
                for nome_grupo, canais_selecionaveis in grupos_canais.items():
                    print(f"\n--- Extraindo Categoria: {nome_grupo} ---")
                    menu_canal.click()
                    time.sleep(0.5)

                    try:
                        # O primeiro sem modificador reseta os checkbox anteriores
                        clicar_com_scroll(dashboard_frame, canais_selecionaveis[0], modifiers=None)
                        time.sleep(0.3)
                        # Combina com Control para compor a matriz
                        for canal_extra in canais_selecionaveis[1:]:
                            clicar_com_scroll(dashboard_frame, canal_extra, modifiers=["Control"])
                            time.sleep(0.3)
                    except Exception as e:
                        print(f"Atenção: Falha na filtragem da sub-matriz {nome_grupo}. ({e})")
                        
                    time.sleep(0.5)
                    menu_canal.click() # Fecha para a tela atualizar
                    # Ponto de checagem. Como a label pode piscar:
                    esperar_carregamento(dashboard_frame)
                    
                    # Usa a validação ativa baseada no valor processado no loop anterior para garantir a carga
                    r_realizado = extrair_valor_garantido(dashboard_frame, "Valor Realizado", referencia_valida_realizado)
                    referencia_valida_realizado = r_realizado # Atualiza a referência em formato Cascata
                    
                    r_bilhetes = extrair_valor_bi(dashboard_frame, "Total Passagens")
                    r_tm_canal = extrair_valor_bi(dashboard_frame, "Fin Ticket Médio")
                    
                    # Cálculo matemático local vinculado à matriz mensal
                    perc_do_mes = percentuais_orcado_matriz[nome_grupo].get(mes_para_selecionar, 0.0)
                    orcamento_fatiado = v_total_orcado_num * perc_do_mes

                    print(f"Coleta do Grupo {nome_grupo}: Realizado={r_realizado} | Matriz Orcado={orcamento_fatiado:.2f}")

                    dados_tabela_excel.append({
                        "Ano": ano_atual,
                        "Visão": "DESMEMBRAMENTO DE CANAIS",
                        "Categoria": nome_grupo,
                        "Orçado (Fórmula)": orcamento_fatiado,
                        "Realizado Geral": extrair_numero(r_realizado),
                        "Passagens/Mix": extrair_numero(r_bilhetes),
                        "TM Geral": extrair_numero(r_tm_canal),
                        "TM Online": "", # Não exigido por grupo
                        "TM Offline": "" # Não exigido por grupo
                    })

                # Limpar filtros da caixa suspensa (Selecionar Tudo) 
                print(f"[{ano_atual}] Varredura concluída. Restaurando Filtros de Canal (Duplo Clique no Selecionar tudo)...")
                menu_canal.click()
                time.sleep(0.5)
                # Selecionar tudo roda 2 vezes para garantir toggle correto (marcar e desmarcar de fato)
                clicar_com_scroll(dashboard_frame, "Selecionar tudo", modifiers=None)
                time.sleep(0.5)
                clicar_com_scroll(dashboard_frame, "Selecionar tudo", modifiers=None)
                time.sleep(0.5)
                menu_canal.click()
                esperar_carregamento(dashboard_frame)

            # FIM DO LOOP ANUAL! A MAGIA DO PANDAS ENTRA AQUI
            print("\n" + "="*50)
            print("EXTRAÇÃO CONCLUÍDA - PREPARANDO MICRO SERVIÇO DE EXCEL...")

            pprint.pprint(dados_tabela_excel)
            
            df = pd.DataFrame(dados_tabela_excel)
            caminho_xlsx = caminho_extracao
            caminho_xlsx_path = Path(caminho_xlsx)
            caminho_xlsx_path.parent.mkdir(parents=True, exist_ok=True)
            
            # Usando openpyxl na exportação de formatação via writer
            with pd.ExcelWriter(caminho_xlsx, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name="Extrato Automático")
                # Estilização básica pra números
                worksheet = writer.sheets["Extrato Automático"]
                for row in worksheet.iter_rows(min_col=4, max_col=9, min_row=2):
                    for cell in row:
                        if isinstance(cell.value, (int, float)):
                            cell.number_format = '#,##0.00' # Formato Numérico Limpo

            print(f"SUCESSO TOTAL! Planilha compilada para números brutos (calculável) e salva em:")
            print(f">>> {caminho_xlsx}")
            print("="*50 + "\n")
            
            # --- START DA MIGRACÃO AUTOMÁTICA GABARITO ---
            mes_sigla = mes_para_selecionar.split('-')[1][:3].upper()
            ultimo_dia_selecionado = dias_para_selecionar[-1].zfill(2)
            print(f"Iniciando Micro-Serviço integrador do Excel para {mes_sigla} (Data Máx: {ultimo_dia_selecionado})...")
            if not caminho_template:
                raise ValueError("Caminho do template comparativo nao foi informado.")
            atualizar_planilha_comparativa(
                caminho_extracao=caminho_xlsx,
                caminho_template=str(caminho_template),
                mes_sigla=mes_sigla,
                ultimo_dia=ultimo_dia_selecionado,
                ano_referencia=int(ano_referencia),
            )

            if callback_progresso:
                callback_progresso(0.95, f"Mes {mes_sigla} processado e aplicado no comparativo.")

            resultado = {
                "arquivo_principal": str(caminho_template),
                "arquivos_saida": [str(caminho_template)],
                "pasta_final": str(Path(caminho_template).parent),
                "mensagem": f"Extracao do mes {mes_para_selecionar} concluida.",
            }

        except Exception as e:
            raise RuntimeError(f"Erro Master na Dashboard: {e}")

        print("Encerramento seguro acionado...")
        browser.close()
        return resultado

def executar_busca_dados(
    user_id,
    data_ini,
    data_fim,
    base_automacao=None,
    pasta_personalizada=None,
    saida_automacao=None,
    pasta_saida=None,
    callback_progresso=None,
    servico_credencial=DEFAULT_CRED_SERVICE,
):
    data_ini_dt = parse_frontend_date(data_ini)
    data_fim_dt = parse_frontend_date(data_fim)

    hoje = datetime.now()
    ultimo_dia_fechado = datetime(hoje.year, hoje.month, hoje.day) - timedelta(days=1)
    if data_ini_dt is None:
        data_ini_dt = datetime(ultimo_dia_fechado.year, ultimo_dia_fechado.month, 1)
    if data_fim_dt is None:
        data_fim_dt = ultimo_dia_fechado

    data_ini_dt = datetime(data_ini_dt.year, data_ini_dt.month, data_ini_dt.day)
    data_fim_dt = datetime(data_fim_dt.year, data_fim_dt.month, data_fim_dt.day)

    if data_ini_dt > data_fim_dt:
        raise ValueError("Data inicial nao pode ser maior que a data final.")

    base_dir = resolver_diretorio_base(base_automacao=base_automacao, pasta_personalizada=pasta_personalizada)
    base_dir.mkdir(parents=True, exist_ok=True)
    caminho_template = base_dir / DEFAULT_TEMPLATE_FILE
    if not caminho_template.exists():
        raise FileNotFoundError(f"Template nao encontrado: {caminho_template}")

    login_usuario, senha_usuario = buscar_credencial_site(int(user_id), servico_credencial)
    if not login_usuario or not senha_usuario:
        raise ValueError(f"Credenciais '{servico_credencial}' nao encontradas no Cofre.")

    segmentos = gerar_segmentos_mensais(data_ini_dt, data_fim_dt)
    if not segmentos:
        raise ValueError("Nenhum segmento mensal foi gerado para o periodo informado.")

    DEFAULT_TEMP_DIR.mkdir(parents=True, exist_ok=True)
    arquivos_temporarios = []

    if callback_progresso:
        callback_progresso(0.03, f"Periodo dividido em {len(segmentos)} mes(es).")

    try:
        for idx, segmento in enumerate(segmentos, start=1):
            mes_label = segmento["mes_label"]
            ano_atual = int(segmento["ano"])
            anos_processar = [str(ano_atual), str(ano_atual - 1)]
            temp_xlsx = DEFAULT_TEMP_DIR / f"analise_performance_{ano_atual}_{int(segmento['mes']):02d}.xlsx"
            arquivos_temporarios.append(temp_xlsx)

            inicio_faixa = 0.08 + (idx - 1) * (0.82 / len(segmentos))
            fim_faixa = 0.08 + idx * (0.82 / len(segmentos))

            if callback_progresso:
                callback_progresso(
                    inicio_faixa,
                    f"Processando mes {idx}/{len(segmentos)}: {mes_label} ({segmento['dias'][0]}-{segmento['dias'][-1]})",
                )

            def cb_segmento(p, m):
                if not callback_progresso:
                    return
                progresso_local = inicio_faixa + ((fim_faixa - inicio_faixa) * max(0.0, min(1.0, p)))
                callback_progresso(progresso_local, m)

            run(
                mes_para_selecionar=mes_label,
                dias_para_selecionar=segmento["dias"],
                anos_para_processar=anos_processar,
                caminho_extracao=str(temp_xlsx),
                caminho_template=str(caminho_template),
                ano_referencia=ano_atual,
                credenciais=(login_usuario, senha_usuario),
                callback_progresso=cb_segmento,
            )
    finally:
        for arq in arquivos_temporarios:
            try:
                if arq.exists():
                    arq.unlink()
            except Exception:
                pass

    if callback_progresso:
        callback_progresso(1.0, "Relatorio Performance de Canais finalizado com sucesso.")

    pasta_saida_final = resolver_diretorio_saida(
        saida_automacao=saida_automacao,
        pasta_saida=pasta_saida,
        base_dir=base_dir,
    )
    pasta_saida_final.mkdir(parents=True, exist_ok=True)

    arquivo_principal = caminho_template
    if pasta_saida_final.resolve() != base_dir.resolve():
        arquivo_principal = pasta_saida_final / caminho_template.name
        shutil.copy2(str(caminho_template), str(arquivo_principal))

    return {
        "arquivo_principal": str(arquivo_principal),
        "arquivos_saida": [str(arquivo_principal)],
        "pasta_final": str(pasta_saida_final),
        "mensagem": f"Busca de dados concluida para {len(segmentos)} mes(es).",
    }


if __name__ == "__main__":
    try:
        if len(sys.argv) > 1:
            params = json.loads(base64.b64decode(sys.argv[1]))
        else:
            line = sys.stdin.readline()
            params = json.loads(line) if line else {}
    except Exception:
        params = {}

    def progress_callback(p, m):
        texto = str(m or "").replace('"', "'")
        print(f'PROGRESS:{{"p": {int(max(0, min(100, round(p * 100))))}, "m": "{texto}"}}', flush=True)

    try:
        resultado = executar_busca_dados(
            user_id=params.get("user_id", 1),
            data_ini=params.get("data_ini"),
            data_fim=params.get("data_fim"),
            base_automacao=params.get("base"),
            pasta_personalizada=params.get("pasta_personalizada"),
            saida_automacao=params.get("saida"),
            pasta_saida=params.get("pasta_saida"),
            callback_progresso=progress_callback,
            servico_credencial=params.get("servico_credencial") or DEFAULT_CRED_SERVICE,
        )
        print(json.dumps(resultado, ensure_ascii=False))
    except Exception as e:
        print(f"ERRO: {str(e)}", file=sys.stderr)
        sys.exit(1)
