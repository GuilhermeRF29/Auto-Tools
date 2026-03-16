import os
import shutil
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Alignment

# ==========================================
# FUNÇÃO AUXILIAR DE FORMATAÇÃO VISUAL
# ==========================================
def aplicar_formatacao_excel(caminho_arquivo, nome_aba):
    """
    Abre um arquivo Excel existente, aplica cores de fundo no cabeçalho,
    ajusta larguras das colunas e centraliza os textos.
    """
    wb = load_workbook(caminho_arquivo)
    ws = wb[nome_aba]
    
    # Define a cor de fundo Cinza para o cabeçalho
    fundo_cabeçalho = PatternFill(start_color="BFBFBF", fill_type="solid")
    alinhamento = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Aumenta a altura da primeira linha (Cabeçalho)
    ws.row_dimensions[1].height = 25
    
    # Nomes das colunas para formatação específica (em letras minúsculas e sem espaços nas pontas)
    colunas_moeda_limpas = [
        'tarifa', 'embarque', 'pedágio', 'pedagio', 
        'sugestão', 'sugestao', 'sugestão revenue', 'sugestao revenue',
        'revenue aplicado'
    ]
    colunas_porcentagem_limpas = [
        'market share', 'perc. de diferença', 'perc. de diferenca', 'aproveitamento'
    ]

    indices_moeda = []
    indices_porcentagem = []

    # Aplica o estilo para cada coluna no cabeçalho e identifica os índices das colunas especiais
    for col in range(1, ws.max_column + 1):
        letra = get_column_letter(col)
        celula = ws.cell(row=1, column=col)
        celula.fill = fundo_cabeçalho
        celula.alignment = alinhamento
        
        # Guarda o índice para formatar o resto da coluna depois
        nome_cabecalho = str(celula.value).strip().lower() if celula.value else ""
        
        if nome_cabecalho in colunas_moeda_limpas:
            indices_moeda.append(col)
        elif nome_cabecalho in colunas_porcentagem_limpas:
            indices_porcentagem.append(col)
        
        # A coluna 'S' (Justificativa no relatorio base) recebe uma largura maior, o resto recebe 20
        ws.column_dimensions[letra].width = 75.6 if letra == 'S' else 20
        
    # Centraliza todas as células das linhas normais e aplica formatos numéricos
    for linha in ws.iter_rows(min_row=2):
        for celula in linha: 
            celula.alignment = Alignment(vertical="center", wrap_text=True)
            
            # Formata apenas se a célula contiver um número (ignora células vazias ou texto)
            if celula.column in indices_moeda:
                if isinstance(celula.value, (int, float)):
                    celula.number_format = 'R$ #,##0.00_-'
            elif celula.column in indices_porcentagem:
                if isinstance(celula.value, (int, float)):
                    celula.number_format = '0.00 %'
            
    # Salva e FECHA o arquivo
    wb.save(caminho_arquivo)
    wb.close()

# ==========================================
# FUNÇÃO AUXILIAR DA JUSTIFICATIVA
# ==========================================
def padronizar_justificativa(df: pd.DataFrame) -> pd.DataFrame:
    if 'Status Revenue' not in df.columns:
        return df

    if 'Justificativa' not in df.columns:
        df['Justificativa'] = ''
    
    novos_valores = []
    for just, status in zip(df['Justificativa'], df['Status Revenue']):
        just_str = '' if pd.isna(just) else str(just).strip().upper()
        status_str = '' if pd.isna(status) else str(status).strip().upper()
        
        eh_vazio = just_str in ('', 'NAN', 'NONE', 'NULL', 'NAT')
        
        if eh_vazio and status_str == 'APROVADO':
            novos_valores.append('Aprovada')
        elif eh_vazio and status_str in ('REPROVADO', 'RECUSADO'):
            novos_valores.append('Concorrência')
        else:
            novos_valores.append(str(just) if not pd.isna(just) else '')
            
    df['Justificativa'] = novos_valores
    return df

# ==========================================
# MOTOR DE CONCILIAÇÃO DOS ARQUIVOS EBUS (MINIATURA PARA TESTE)
# ==========================================
def processar_arquivos_relatorios(arquivo_original, destino, nome_mes, ano_atual):
    print(f"[{nome_mes}] - Lendo planilha bruta (xlrd)...")
    df_novo = pd.read_excel(arquivo_original, engine='xlrd')
    df_novo = df_novo.dropna(how='all')

    if 'Não foi possivel obter dados com os parâmetros informados.' in df_novo.columns:
        if not df_novo.empty and df_novo['Não foi possivel obter dados com os parâmetros informados.'].iloc[0] is True:
            print("Planilha original identificada como vazia sem dados.")
            return
        elif df_novo.empty:
            print("Planilha original identificada como totalmente vazia.")
            return

    if 'Origem' in df_novo.columns and 'Destino' in df_novo.columns:
        df_novo['Concatenar Origem e Destino'] = df_novo['Origem'].astype(str) + " - " + df_novo['Destino'].astype(str)

    print(f"[{nome_mes}] - Padronizando Justificativas...")
    df_novo = padronizar_justificativa(df_novo)

    nome_oficial = f"Relatorio Revenue Completo - {nome_mes.capitalize()} {ano_atual}.xlsx"
    pasta_base = destino / "BASE"
    pasta_base_nova = destino / "BASE NOVA"
    
    for pasta in [pasta_base, pasta_base_nova]:
        pasta.mkdir(parents=True, exist_ok=True)

    caminho_base = pasta_base / nome_oficial
    caminho_base_nova = pasta_base_nova / nome_oficial

    # --- BASE NOVA ---
    print(f"[{nome_mes}] - Criando/Atualizando Base Nova...")
    # (Para o teste isolado, vamos apenas salvar e formatar a Base Nova diretamente, ignorando o estado antigo para focar na formatação monetária)
    df_novo_base_nova = df_novo.copy()
    if 'Estado' not in df_novo_base_nova.columns:
        df_novo_base_nova['Estado'] = 'Novo'
        
    arquivo_temp_estado = destino / f"temp_estado_{nome_oficial}"
    aba_estado = "Relatorio Comparado"
    
    if 'Mercado' in df_novo_base_nova.columns:
        df_novo_base_nova = df_novo_base_nova.drop(columns=['Mercado'])

    df_novo_base_nova.to_excel(arquivo_temp_estado, index=False, sheet_name=aba_estado)
    
    print(f"[{nome_mes}] - Aplicando formatação na Base Nova...")
    aplicar_formatacao_excel(arquivo_temp_estado, aba_estado)
    shutil.move(str(arquivo_temp_estado), str(caminho_base_nova))

    # --- BASE NORMAL ---
    print(f"[{nome_mes}] - Criando/Atualizando Base Normal...")
    df_novo_base = df_novo.copy()
    if 'Estado' in df_novo_base.columns:
        df_novo_base = df_novo_base.drop(columns=['Estado'])
    df_novo_base['Mercado'] = ""

    arquivo_temp_base = destino / f"temp_base_{nome_oficial}"
    aba_normal = "Relatorio Revenue Sistema"
    
    df_novo_base.to_excel(arquivo_temp_base, index=False, sheet_name=aba_normal)
    
    print(f"[{nome_mes}] - Aplicando formatação na Base Normal...")
    aplicar_formatacao_excel(arquivo_temp_base, aba_normal)
    shutil.move(str(arquivo_temp_base), str(caminho_base))

    print(f"[{nome_mes}] - Concluído! Verifique os arquivos na pasta {destino}")

# ==========================================
# PARA EXECUTAR O TESTE VÁ NESTA PARTE:
# ==========================================
if __name__ == '__main__':
    # 1. Defina o caminho do arquivo .xls que você baixou
    # Exemplo: ARQUIVO_TESTE = Path(r"C:\Users\SEU_USUARIO\Downloads\RelatorioRevenue.xls")
    
    # Substitua este caminho pelo caminho real do seu arquivo
    caminho = Path.home()
    ARQUIVO_TESTE = Path(caminho / "Downloads" / "RelatorioRevenue2026-03-03151814.xls")
    
    # 2. Defina onde ele vai salvar as pastas "BASE" e "BASE NOVA"
    PASTA_DESTINO = Path(caminho / "Relatórios Revenue" / "Teste_ebus")
    
    # 3. Informações do mês para gerar o nome do arquivo
    NOME_MES = "Teste"
    ANO = 2026
    
    if ARQUIVO_TESTE.exists():
        PASTA_DESTINO.mkdir(parents=True, exist_ok=True)
        print("Iniciando Teste...")
        processar_arquivos_relatorios(ARQUIVO_TESTE, PASTA_DESTINO, NOME_MES, ANO)
    else:
        print(f"Arquivo não encontrado: {ARQUIVO_TESTE}")
        print("Por favor, atualize o caminho ARQUIVO_TESTE no final do script.")
