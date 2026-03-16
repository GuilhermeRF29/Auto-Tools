import getpass
import selenium
import time
import os
import shutil
from pathlib import Path
from datetime import datetime, timedelta
import calendar
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.edge.options import Options
from selenium.webdriver.edge.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Alignment, Font
import sys

MESES_PT = {1: 'Janeiro', 2: 'Fevereiro', 3: 'Março', 4: 'Abril', 5: 'Maio', 
            6: 'Junho', 7: 'Julho', 8: 'Agosto', 9: 'Setembro', 10: 'Outubro', 
            11: 'Novembro', 12: 'Dezembro'}

def aplicar_formatacao_excel(caminho_arquivo, nome_aba):
    """
    Abre um arquivo Excel existente, aplica cores de fundo no cabeçalho,
    ajusta larguras das colunas e centraliza os textos.
    É fundamental usar wb.close() no final para liberar o arquivo no Windows.
    """
    wb = load_workbook(caminho_arquivo)
    ws = wb[nome_aba]
    
    # Define a cor de fundo Cinza para o cabeçalho
    fundo_cabeçalho = PatternFill(start_color="BFBFBF", fill_type="solid")
    alinhamento = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Aumenta a altura da primeira linha (Cabeçalho)
    ws.row_dimensions[1].height = 25
    
    # Nomes das colunas para formatação específica (em letras minúsculas e sem espaços nas pontas)
    colunas_moeda_limpas = [
        'tarifa', 'embarque', 'pedágio', 'pedagio', 
        'sugestão', 'sugestao', 'sugestão revenue', 'sugestao revenue',
        'revenue aplicado'
    ]
    colunas_porcentagem_limpas = [
        'market share', 'perc. de diferença', 'perc. de diferenca', 'aproveitamento'
    ]

    indices_moeda = []
    indices_porcentagem = []

    # Aplica o estilo para cada coluna no cabeçalho e identifica os índices das colunas especiais
    for col in range(1, ws.max_column + 1):
        letra = get_column_letter(col)
        celula = ws.cell(row=1, column=col)
        celula.fill = fundo_cabeçalho
        celula.alignment = alinhamento
        
        # Guarda o índice para formatar o resto da coluna depois
        nome_cabecalho = str(celula.value).strip().lower() if celula.value else ""
        
        if nome_cabecalho in colunas_moeda_limpas:
            indices_moeda.append(col)
        elif nome_cabecalho in colunas_porcentagem_limpas:
            indices_porcentagem.append(col)
        
        # A coluna 'S' recebe uma largura maior, o resto recebe tamanho padrão 20
        ws.column_dimensions[letra].width = 75.6 if letra == 'S' else 20
        
    # Centraliza todas as células das linhas normais e aplica formatos numéricos
    for linha in ws.iter_rows(min_row=2):
        for celula in linha: 
            celula.alignment = Alignment(vertical="center", wrap_text=True)
            
            # Formata apenas se a célula contiver um número (ignora células vazias ou texto)
            if celula.column in indices_moeda:
                if isinstance(celula.value, (int, float)):
                    celula.number_format = 'R$ #,##0.00_-'
            elif celula.column in indices_porcentagem:
                if isinstance(celula.value, (int, float)):
                    celula.number_format = '0.00 %'
            
    # Salva e FECHA o arquivo para evitar erros de permissão ao mover
    wb.save(caminho_arquivo)
    wb.close() 

# ==========================================
# FUNÇÃO AUXILIAR DA JUSTIFICATIVA
# ==========================================
def padronizar_justificativa(df):
    if 'Status Revenue' not in df.columns:
        return df

    if 'Justificativa' not in df.columns:
        df['Justificativa'] = ''
    
    novos_valores = []
    for just, status in zip(df['Justificativa'], df['Status Revenue']):
        # Limpeza segura de valores nulos ou vazios
        just_str = '' if pd.isna(just) else str(just).strip().upper()
        status_str = '' if pd.isna(status) else str(status).strip().upper()
        
        eh_vazio = just_str in ('', 'NAN', 'NONE', 'NULL')
        
        if eh_vazio and status_str == 'APROVADO':
            novos_valores.append('Aprovada')
        elif eh_vazio and status_str == 'REPROVADO': # ou 'RECUSADO' dependendo do site
            novos_valores.append('Concorrência')
        else:
            novos_valores.append(str(just) if not pd.isna(just) else '')
            
    df['Justificativa'] = novos_valores
    return df


# ==========================================
# MOTOR DE CONCILIAÇÃO DOS ARQUIVOS EBUS
# ==========================================
def processar_arquivos_relatorios(arquivo_original, destino, nome_mes=None, ano_atual=None, callback_progresso=None):
    """
    Função coração da regra de negócio EBUS.
    1. Lê a planilha web crua.
    2. Valida se veio vazia.
    3. Concatena colunas de Origem + Destino e aplica Justificativa.
    4. Processa a "Base Nova" (Histórico de Novo, Excluido, Presente).
    5. Processa a "Base Normal" (Mercado vazio, regra de volume).
    """
    
    if callback_progresso: callback_progresso(0.6, "Lendo dados originais...")
    df_novo = pd.read_excel(arquivo_original, engine='xlrd')

    df_novo = df_novo.dropna(how='all')
    
    # =========================================================
    # DETECÇÃO AUTOMÁTICA DE MÊS E ANO (Via Data Aplicação)
    # =========================================================
    if nome_mes is None or ano_atual is None:
        if 'Data Viagem' in df_novo.columns:
            # Tenta converter a primeira data válida para datetime com formato brasileiro (dayfirst=True)
            datas_validas = pd.to_datetime(df_novo['Data Viagem'], format='%d/%m/%Y', errors='coerce')
            if datas_validas.isna().all():
                # Plano B: se tiver formato de hora junto ou outro separador
                datas_validas = pd.to_datetime(df_novo['Data Viagem'], dayfirst=True, errors='coerce')
            
            datas_validas = datas_validas.dropna()
            if not datas_validas.empty:
                primeira_data = datas_validas.iloc[0]
                nome_mes = MESES_PT[primeira_data.month]
                ano_atual = primeira_data.year
            else:
                nome_mes = nome_mes or "Desconhecido"
                ano_atual = ano_atual or datetime.now().year
        else:
            nome_mes = nome_mes or "Desconhecido"
            ano_atual = ano_atual or datetime.now().year

    # Faz a validação final se o arquivo está vazio ou se contém a mensagem de erro do sistema
    mensagem_vazio = 'Não foi possivel obter dados com os parâmetros informados.'
    if df_novo.empty or mensagem_vazio in df_novo.columns or df_novo.apply(lambda col: col.astype(str).str.contains(mensagem_vazio, regex=False)).any().any():
        if callback_progresso: callback_progresso(0.45, f"Relatório de {nome_mes} está sem dados no sistema. Descartando e ignorando...")
        print(f"Relatório vazio ignorado: {arquivo_original}")
        try:
            os.remove(arquivo_original)
        except OSError:
            pass
        return

    # =========================================================
    # TRATAMENTOS UNIVERSAIS (Aplicados a todas as bases)
    # =========================================================
    if 'Origem' in df_novo.columns and 'Destino' in df_novo.columns:
        df_novo['Concatenar Origem e Destino'] = df_novo['Origem'].astype(str) + " - " + df_novo['Destino'].astype(str)

    # Aplica as regras da Justificativa antes de começar as comparações
    df_novo = padronizar_justificativa(df_novo)

    # =========================================================
    # ESTRUTURA DE PASTAS E ARQUIVOS
    # =========================================================
    nome_oficial = f"Relatorio Revenue Completo - {nome_mes.capitalize()} {ano_atual}.xlsx"
    
    pasta_base = destino / "BASE"
    pasta_backup = destino / "BACKUP"
    pasta_base_nova = destino / "BASE NOVA"
    pasta_backup_base_nova = destino / "BACKUP NOVO"
    
    # Criando todas as pastas de uma vez de forma limpa
    for pasta in [pasta_base, pasta_backup, pasta_base_nova, pasta_backup_base_nova]:
        pasta.mkdir(parents=True, exist_ok=True)

    caminho_base = pasta_base / nome_oficial
    caminho_backup = pasta_backup / nome_oficial
    caminho_base_nova = pasta_base_nova / nome_oficial
    caminho_backup_base_nova = pasta_backup_base_nova / nome_oficial

    # =========================================================
    # PARTE 1: BASE NOVA (Mapeamento de Novo / Excluido / Presente)
    # =========================================================
    if callback_progresso: callback_progresso(0.7, "Processando Base Nova...")
    houve_alteracao_estado = True

    if caminho_base_nova.exists():
        df_base_estado_antiga = pd.read_excel(caminho_base_nova)
        
        if 'Estado' in df_base_estado_antiga.columns:
            df_old_active = df_base_estado_antiga[df_base_estado_antiga['Estado'] != 'Excluido'].copy()
            df_previously_excluded = df_base_estado_antiga[df_base_estado_antiga['Estado'] == 'Excluido'].copy()
        else:
            df_old_active = df_base_estado_antiga.copy()
            df_previously_excluded = pd.DataFrame()
            
        df_old_comp = df_old_active.drop(columns=['Estado'], errors='ignore')
        df_new_comp = df_novo.copy()
        
        df_old_comp = padronizar_justificativa(df_old_comp)
        
        df_old_comp_filled = df_old_comp.fillna('')
        df_new_comp_filled = df_new_comp.fillna('')
        
        # Garantindo que as colunas comuns tenham o mesmo tipo para o merge (evita object vs float64)
        for col in df_old_comp_filled.columns.intersection(df_new_comp_filled.columns):
            if df_old_comp_filled[col].dtype != df_new_comp_filled[col].dtype:
                df_old_comp_filled[col] = df_old_comp_filled[col].astype(str)
                df_new_comp_filled[col] = df_new_comp_filled[col].astype(str)
        
        df_comparado = pd.merge(df_old_comp_filled, df_new_comp_filled, how='outer', indicator=True)
        
        novas_mudancas = len(df_comparado[df_comparado['_merge'] == 'right_only']) + len(df_comparado[df_comparado['_merge'] == 'left_only'])
        
        mapa_estado = {'left_only': 'Excluido', 'right_only': 'Novo', 'both': 'Presente'}
        
        df_comparado['Estado'] = df_comparado['_merge'].map(mapa_estado)
        df_comparado = df_comparado.drop(columns=['_merge'])
        
        if not df_previously_excluded.empty:
            df_comparado = pd.concat([df_comparado, df_previously_excluded], ignore_index=True)
            
        if novas_mudancas == 0:
            houve_alteracao_estado = False
            
    else:
        df_comparado = df_novo.copy()
        df_comparado['Estado'] = 'Novo'
        houve_alteracao_estado = True

    if houve_alteracao_estado:
        arquivo_temp_estado = destino / f"temp_estado_{nome_oficial}"
        aba_estado = "Relatorio Comparado"
        
        # Garante que a coluna Mercado não vai sujar a base nova
        if 'Mercado' in df_comparado.columns:
            df_comparado = df_comparado.drop(columns=['Mercado'])

        df_comparado.to_excel(arquivo_temp_estado, index=False, sheet_name=aba_estado)
        aplicar_formatacao_excel(arquivo_temp_estado, aba_estado)
        
        if caminho_base_nova.exists():
            shutil.copy2(str(caminho_base_nova), str(caminho_backup_base_nova))
            caminho_base_nova.unlink()
            
        shutil.move(str(arquivo_temp_estado), str(caminho_base_nova))

        # =========================================================================
        # CRIAÇÃO OU ATUALIZAÇÃO DO ARQUIVO CONSOLIDADO CSV (BASE NOVA)
        # =========================================================================
        
        # 1. Definimos o caminho onde o CSV consolidado ficará armazenado. 
        # Ele fica junto aos arquivos da BASE NOVA
        csv_consol = pasta_base_nova / 'base_nova_consolidada.csv'
        
        # 2. Fazemos uma cópia do dataframe comparado atual (contendo Novo, Presente, Excluido, etc.)
        df_csv_novo = df_comparado.copy()
        
        if csv_consol.exists():
            # 3. Caso o arquivo de consolidação já exista (de meses anteriores), 
            # fazemos a leitura dele. Usamos separador ';' e utf-8-sig para garantir a compatibilidade de acentos.
            df_csv_existente = pd.read_csv(csv_consol, sep=';', encoding='utf-8-sig')
            
            # 4. Agrupamos (concatenamos) os dados do mês atual recém-processados
            # com os dados históricos armazenados no CSV, empilhando as informações, sem sobrescrevê-las.
            # ignore_index=True garante que os índices da tabela inteira sejam mantidos na ordem e recriados corretamente.
            df_final_csv = pd.concat([df_csv_existente, df_csv_novo], ignore_index=True)
            
            # 5. Aplicamos o drop_duplicates() para remover duplicatas reais.
            # Usamos as colunas sugeridas como identificadores únicos para garantir que, 
            # se o mesmo relatório for rodado duas vezes, o CSV não infle.
            colunas_unicas = [
                'Origem', 'Destino', 'Data Viagem', 
                'Sugestão Revenue', 'Revenue Aplicado'
            ]
            
            # Filtramos apenas as colunas que realmente existem no DataFrame para evitar erros
            subset_existente = [c for c in colunas_unicas if c in df_final_csv.columns]
            
            if subset_existente:
                # Convertemos para string temporariamente para garantir uma comparação de duplicatas precisa
                # (evita problemas de 10.0 vs 10 no Excel)
                df_final_csv = df_final_csv.drop_duplicates(subset=subset_existente, keep='last')
            else:
                df_final_csv = df_final_csv.drop_duplicates()
        else:
            # 6. Se o arquivo consolidado CSV ainda não existir, esta será a "primeira vez" e os dados atuais
            # formarão a base inicial dele.
            df_final_csv = df_csv_novo
        
        # 7. Finalmente, exportamos a versão acumulada para o arquivo físico base_nova_consolidada.csv .
        df_final_csv.to_csv(csv_consol, sep=';', index=False, encoding='utf-8-sig')
        # =========================================================================


    # =========================================================
    # PARTE 2: BASE NORMAL (Com "Mercado" em branco)
    # =========================================================
    if callback_progresso: callback_progresso(0.8, "Processando Base Normal...")
    
    df_novo_base = df_novo.copy()
    
    # Prevenção: Garante que a coluna Estado não vaze para a base normal
    if 'Estado' in df_novo_base.columns:
        df_novo_base = df_novo_base.drop(columns=['Estado'])
        
    df_novo_base['Mercado'] = ""

    substituir_base = True
    if caminho_base.exists():
        df_base_antiga = pd.read_excel(caminho_base)
        if len(df_novo_base) <= len(df_base_antiga):
            substituir_base = False

    if substituir_base:
        arquivo_temp_base = destino / f"temp_base_{nome_oficial}"
        aba_normal = "Relatorio Revenue Sistema"
        
        df_novo_base.to_excel(arquivo_temp_base, index=False, sheet_name=aba_normal)
        aplicar_formatacao_excel(arquivo_temp_base, aba_normal)
        
        if caminho_base.exists():
            shutil.copy2(str(caminho_base), str(caminho_backup))
            caminho_base.unlink()
            
        shutil.move(str(arquivo_temp_base), str(caminho_base))

    # =========================================================
    # LIMPEZA FINAL
    # =========================================================
    if os.path.exists(arquivo_original):
        os.remove(arquivo_original)
    
    if callback_progresso: callback_progresso(0.9, "Arquivos processados com sucesso!")

if __name__ == '__main__':
    # =========================================================
    # EXECUÇÃO DO SCRIPT - MAPEAMENTO E CHAMADA PRINCIPAL
    # =========================================================
    origem = Path.home() / "Downloads"
    destino = Path.home() / "Documents" / "Relatórios Revenue" / "Teste_ebus"

    arquivos_encontrados = list(origem.rglob("RelatorioRevenue*.xls"))
    if arquivos_encontrados:
        # Ordenamos do mais antigo para o mais novo para que a conciliação incremental siga a ordem cronológica correta
        arquivos_encontrados.sort(key=os.path.getmtime)
        
        total = len(arquivos_encontrados)
        print(f"Encontrados {total} arquivos para processar.")
        
        # Cria a pasta destino se não existir
        destino.mkdir(parents=True, exist_ok=True)
        
        for i, arquivo_path in enumerate(arquivos_encontrados, 1):
            nome_arquivo = arquivo_path.name
            novo_nome = destino / f"Processando_{i}_{nome_arquivo}"
            
            print(f"[{i}/{total}] Movendo e processando: {nome_arquivo}...")
            
            # Movemos para uma pasta temporária de processamento (destino) para evitar conflitos
            shutil.move(str(arquivo_path), str(novo_nome))
            
            try:
                # O mês e o ano agora são detectados automaticamente dentro da função
                processar_arquivos_relatorios(novo_nome, destino)
                print(f"    OK: Arquivo {i} processado.")
            except Exception as e:
                print(f"    ERRO ao processar {nome_arquivo}: {e}")
                
        print("\nProcessamento total concluído.")
    else:
        print(f"Nenhum arquivo 'RelatorioRevenue*.xls' encontrado em {origem}")
