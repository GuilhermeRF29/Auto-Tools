# automações/adm_new.py
import os
import time
import shutil
import calendar
import pandas as pd
from pathlib import Path
from datetime import datetime, timedelta
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.edge.options import Options
import sys
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
try:
    from selenium_helper import get_driver_path
except ImportError:
    def get_driver_path():
        return None
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

# IMPORTAMOS O BANCO PARA PUXAR A SENHA
from core.banco import buscar_credencial_site

class CanceladoPeloUsuario(Exception):
    pass

def gerar_intervalos_mensais(data_str_inicio, data_str_fim):
    """Divide a janela de datas informada em pedaços mensais para puxar 
    no site (pois pode existir limite de dados por pesquisa no ADM)."""
    inicio = datetime.strptime(data_str_inicio, "%d/%m/%Y")
    fim_global = datetime.strptime(data_str_fim, "%d/%m/%Y")
    atual = inicio
    intervalos = []
    while atual <= fim_global:
        ultimo_dia_mes = calendar.monthrange(atual.year, atual.month)[1]
        data_ultimo_dia = datetime(atual.year, atual.month, ultimo_dia_mes)
        fim_trecho = min(data_ultimo_dia, fim_global)
        intervalos.append((atual.strftime("%d/%m/%Y"), fim_trecho.strftime("%d/%m/%Y")))
        atual = data_ultimo_dia + timedelta(days=1)
    return intervalos

def renomear_arquivo(caminho_arquivo, caminho_destino, contador):
    """Move e renomeia o arquivo que acabou de ser baixado, adicionando um sufixo numérico."""
    extensao = caminho_arquivo.suffix 
    novo_nome = f"Temp_Relatorio_{contador}{extensao}"
    novo_caminho = caminho_destino / novo_nome
    shutil.move(str(caminho_arquivo), str(novo_caminho))
    return novo_caminho

def consolidar_varios_arquivos(
    diretorio_origem,
    arquivo_saida,
    callback_progresso=None,
    enviar_para=None,
    criar_backup=False,
    limpar_temporarios=True,
):
    """Consolida os arquivos temporários e opcionalmente envia o resultado para outro local."""
    if callback_progresso:
        callback_progresso(0.8, "Analisando relatórios temporários baixados...")

    caminho_origem = Path(diretorio_origem)
    caminho_saida_local = Path(arquivo_saida)
    caminho_saida_local.parent.mkdir(parents=True, exist_ok=True)

    lista_dfs = []
    colunas_padrao = None
    arquivos_excel = sorted([f for f in caminho_origem.glob("Temp_Relatorio_*")])

    if not arquivos_excel:
        return {
            "arquivo_principal": None,
            "arquivos_saida": [],
            "pasta_final": None,
            "mensagem": "Nenhum arquivo temporário encontrado para consolidar.",
        }

    total_arquivos = len(arquivos_excel)
    for index, arquivo in enumerate(arquivos_excel, 1):
        if callback_progresso:
            callback_progresso(0.80 + (0.05 * (index / total_arquivos)), f"Lendo e padronizando arquivo {index} de {total_arquivos}...")
        print(f"Inspecionando: {arquivo.name}")
        xls = pd.ExcelFile(arquivo)
        for nome_aba in xls.sheet_names:
            df_aba = pd.read_excel(xls, sheet_name=nome_aba, header=None)
            df_aba = df_aba.dropna(how='all').dropna(axis=1, how='all')
            if not df_aba.empty:
                primeira_celula = str(df_aba.iloc[0, 0]).strip()
                if primeira_celula.lower() == "data":
                    if colunas_padrao is None:
                        colunas_padrao = df_aba.iloc[0].tolist()
                    df_aba = df_aba.iloc[1:].copy()
                if df_aba.empty:
                    continue
                if colunas_padrao is not None and len(df_aba.columns) == len(colunas_padrao):
                    df_aba.columns = colunas_padrao
                    lista_dfs.append(df_aba)
        xls.close()

    if not lista_dfs:
        return {
            "arquivo_principal": None,
            "arquivos_saida": [],
            "pasta_final": None,
            "mensagem": "Os arquivos temporários não continham dados válidos para consolidação.",
        }

    if callback_progresso:
        callback_progresso(0.86, "Base final montada. Concatenando e formatando dados...")

    df_final = pd.concat(lista_dfs, ignore_index=True)
    if "Data" in df_final.columns:
        df_final["Data"] = pd.to_datetime(df_final["Data"], errors='coerce').dt.strftime('%d/%m/%Y')

    df_final['Data Observação'] = datetime.now().strftime('%d/%m/%Y')

    if 'Linha' in df_final.columns:
        df_final['Linha'] = df_final['Linha'].astype(str).str.strip()
        df_final['Linha'] = pd.to_numeric(df_final['Linha'], errors='coerce').astype("Int64")

    df_final.to_excel(caminho_saida_local, index=False, sheet_name='Planilha1')
    arquivo_final = caminho_saida_local

    if enviar_para:
        pasta_envio = Path(enviar_para)
        pasta_envio.mkdir(parents=True, exist_ok=True)
        destino_final_arquivo = pasta_envio / caminho_saida_local.name

        if callback_progresso:
            callback_progresso(0.9, "Enviando arquivo consolidado para pasta final...")

        shutil.move(str(caminho_saida_local), str(destino_final_arquivo))
        arquivo_final = destino_final_arquivo

        if criar_backup:
            try:
                destino_backup = Path(r"\\172.16.98.12\Relatórios Power BI\Forecast\Forecast - Antigo")
                destino_backup.mkdir(parents=True, exist_ok=True)
                shutil.copy2(str(destino_final_arquivo), str(destino_backup / destino_final_arquivo.name))
            except Exception as e:
                print(f"Erro ao gerar backup: {e}")

    if callback_progresso:
        callback_progresso(0.92, "Formatando as células (Openpyxl)...")

    wb = load_workbook(arquivo_final)
    ws = wb.active
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 18
    tamanho_fonte = Font(size=8)
    alinhamento_fonte = Alignment(wrap_text=True)
    for linha in ws.iter_rows():
        for celula in linha:
            celula.font = tamanho_fonte
            celula.alignment = alinhamento_fonte
    wb.save(arquivo_final)
    wb.close()

    if limpar_temporarios:
        for f in arquivos_excel:
            try:
                os.remove(f)
            except Exception:
                pass

    return {
        "arquivo_principal": str(arquivo_final),
        "arquivos_saida": [str(arquivo_final)],
        "pasta_final": str(Path(arquivo_final).parent),
        "mensagem": "Consolidação concluída com sucesso.",
    }

def buscar_e_mover_arquivo_ano_passado(diretorio_origem, diretorio_destino):
    """
    Busca o relatório da mesma semana do ano anterior para ser utilizado como
    base de comparação/dado na nova dashboard do PowerBI.
    """
    hoje = datetime.now()
    try:
        data_ano_passado = hoje.replace(year=hoje.year - 1)
    except ValueError:
        data_ano_passado = hoje.replace(year=hoje.year - 1, day=28)
        
    dias_para_domingo = (data_ano_passado.weekday() + 1) % 7
    domingo_ano_passado = data_ano_passado - timedelta(days=dias_para_domingo)
    
    print(f"Buscando arquivos da semana que começou no domingo: {domingo_ano_passado.strftime('%d/%m/%Y')}")

    origem = Path(diretorio_origem)
    destino = Path(diretorio_destino)
    arquivo_encontrado = None
    
    for i in range(7):
        data_busca = domingo_ano_passado + timedelta(days=i)
        data_str = data_busca.strftime("%d-%m-%Y") 
        arquivos_possiveis = list(origem.glob(f"*{data_str}*.*"))
        
        if arquivos_possiveis:
            arquivo_encontrado = arquivos_possiveis[0]
            print(f"Sucesso! Arquivo encontrado no dia {data_str}: {arquivo_encontrado.name}")
            break
        else:
            print(f"Sem arquivo para o dia {data_str}. Tentando o próximo dia da semana...")

    if arquivo_encontrado:
        caminho_final = destino / arquivo_encontrado.name
        shutil.copy2(str(arquivo_encontrado), str(caminho_final)) 
        print(f"\nArquivo transferido para: {caminho_final}")
        return caminho_final
    else:
        print("\nPoxa, nenhum arquivo encontrado em nenhum dia dessa semana do ano passado.")
        return None

# =========================================================================
# A FUNÇÃO PRINCIPAL: É ELA QUE O MAIN.PY CHAMA
# =========================================================================
def executar_adm(
    id_usuario_logado,
    data_inicio,
    data_final,
    callback_progresso=None,
    hook_cancelamento=None,
    modo_execucao="completo",
    pasta_destino=None,
    arquivo_entrada=None,
    base_automacao=None,
):
    """Executa ADM em modo modular sem quebrar compatibilidade com chamadas antigas."""

    modos_validos = {
        "download",
        "download_tratamento",
        "tratamento",
        "tratamento_envio",
        "completo",
        "arquivo_tratamento",
        "arquivo_envio",
        "arquivo_tratamento_envio",
    }
    if modo_execucao not in modos_validos:
        raise ValueError(f"Modo de execução inválido para ADM: {modo_execucao}")

    def checar_parada():
        if hook_cancelamento and hook_cancelamento():
            raise CanceladoPeloUsuario("Processo cancelado pelo usuário.")

    if callback_progresso:
        callback_progresso(0.05, "Inicializando Robô ADM...")

    home = Path.home()
    origem_download = home / "Downloads"
    pasta_trabalho = home / "Documents" / "Relatório Demanda"
    pasta_trabalho.mkdir(parents=True, exist_ok=True)

    destino_padrao = Path(base_automacao) if base_automacao else Path(r"\\172.16.98.12\Relatórios Power BI\Forecast\Forecast2")
    destino_envio = Path(pasta_destino) if pasta_destino else None

    precisa_download = modo_execucao in {"download", "download_tratamento", "completo"}
    precisa_tratamento = modo_execucao in {
        "download_tratamento",
        "tratamento",
        "tratamento_envio",
        "completo",
        "arquivo_tratamento",
        "arquivo_tratamento_envio",
    }

    if modo_execucao.startswith("arquivo_") and not arquivo_entrada:
        raise ValueError("Selecione um arquivo de entrada para este modo.")

    if not precisa_download and precisa_tratamento and not arquivo_entrada:
        raise ValueError("Tratamento sem download exige um arquivo já baixado.")

    if modo_execucao == "arquivo_envio":
        origem_arquivo = Path(arquivo_entrada)
        if not origem_arquivo.exists():
            raise FileNotFoundError(f"Arquivo de entrada não encontrado: {origem_arquivo}")
        pasta_final = destino_envio or destino_padrao
        pasta_final.mkdir(parents=True, exist_ok=True)
        destino_final = pasta_final / origem_arquivo.name
        shutil.copy2(str(origem_arquivo), str(destino_final))
        return {
            "arquivo_principal": str(destino_final),
            "arquivos_saida": [str(destino_final)],
            "pasta_final": str(pasta_final),
            "mensagem": "Arquivo enviado com sucesso.",
        }

    driver = None

    try:
        if precisa_download:
            user, passwd = buscar_credencial_site(id_usuario_logado, "ADM de Vendas")
            if not user:
                if callback_progresso:
                    callback_progresso(0.0, "ERRO: Credenciais do ADM não encontradas no Cofre!")
                return {
                    "arquivo_principal": None,
                    "arquivos_saida": [],
                    "pasta_final": None,
                    "mensagem": "Credenciais do ADM não encontradas no Cofre.",
                }

            resultado = gerar_intervalos_mensais(data_inicio, data_final)
            opcoes = Options()
            opcoes.add_argument("--window-size=1920,1080")
            opcoes.add_argument("--headless")

            if callback_progresso:
                callback_progresso(0.1, "Abrindo Navegador Invisível...")

            driver_path = get_driver_path()
            if driver_path:
                from selenium.webdriver.edge.service import Service
                driver = webdriver.Edge(service=Service(driver_path), options=opcoes)
            else:
                driver = webdriver.Edge(options=opcoes)

            checar_parada()
            driver.get("http://ttadm01.jcatlm.com.br:8080/ventaboletosadm/index.zul;jsessionid=xFIW8nh_t8n9-74topChhriraeW-2Y5y-MKUCIG3.gcp-pd-ttadm-01")
            wait = WebDriverWait(driver, 1000)

            if callback_progresso:
                callback_progresso(0.15, "Injetando Credenciais...")

            login_element = wait.until(EC.presence_of_element_located((By.NAME, "j_username")))
            login_element.click()
            login_element.clear()
            login_element.send_keys(user)

            senha = driver.find_element(By.NAME, "j_password")
            senha.click()
            senha.clear()
            if passwd:
                senha.send_keys(passwd)

            driver.find_element(By.NAME, "btnAcessar").click()

            if callback_progresso:
                callback_progresso(0.2, "Navegando entre MENUS de Demandas...")

            wait.until(EC.element_to_be_clickable((By.XPATH, "//button[contains(text(), 'Relatórios')]"))).click()
            wait.until(EC.element_to_be_clickable((By.XPATH, "//a [contains(text(), ' Relatórios')and contains(@id, 'c4')]"))).click()
            time.sleep(1)
            wait.until(EC.element_to_be_clickable((By.XPATH, "//a [contains(text(), ' Relatórios Operacionais')and contains(@id, '25-a')]"))).click()
            wait.until(EC.element_to_be_clickable((By.XPATH, "//a [contains(text(), ' Demandas')and contains(@id, 'a5-a')]"))).click()

            total_blocos = len(resultado)
            for idx, (inicial_data, final_data) in enumerate(resultado):
                checar_parada()
                porc = 0.25 + (0.45 * (idx / total_blocos))
                if callback_progresso:
                    callback_progresso(porc, f"Lote {idx+1}/{total_blocos} - Inserindo filtro: {inicial_data} a {final_data}...")

                data1 = wait.until(EC.element_to_be_clickable((By.XPATH, "//input[contains (@id, 'ia-real')]")))
                data1.clear()
                data1.send_keys(inicial_data)

                data2 = wait.until(EC.element_to_be_clickable((By.XPATH, "//input[contains(@id, 'la-real')]")))
                data2.clear()
                data2.send_keys(final_data)

                wait.until(EC.element_to_be_clickable((By.XPATH, "//button[contains(text(), ' Novo Layout')and contains(@id, 'zb')]"))).click()
                wait.until(EC.element_to_be_clickable((By.XPATH, "//div[contains(@id, '4c')]"))).click()

                if callback_progresso:
                    callback_progresso(porc + 0.05, f"Lote {idx+1}/{total_blocos} - Download solicitado. Aguardando arquivo...")

                tempo_espera = 0
                while tempo_espera < 30:
                    arquivos = list(origem_download.rglob("RelatorioDemandaDetalhado*.xls*"))
                    arquivos_validos = [
                        arq for arq in arquivos
                        if not arq.name.endswith((".crdownload", ".tmp")) and "_a_" not in arq.name
                    ]

                    if arquivos_validos:
                        if callback_progresso:
                            callback_progresso(porc + 0.06, f"Lote {idx+1}/{total_blocos} - Arquivo detectado (espera: {tempo_espera}s).")
                        time.sleep(1)
                        break

                    time.sleep(1)
                    tempo_espera += 1.5

                checar_parada()
                wait.until(EC.element_to_be_clickable((By.XPATH, "//div[contains(@id, 'c-close')]"))).click()
                time.sleep(1)

            time.sleep(10)
            checar_parada()

            arquivos_encontrados = [
                arq for arq in origem_download.glob("RelatorioDemanda*.xls*")
                if not arq.name.endswith((".crdownload", ".tmp"))
            ]

            if not arquivos_encontrados:
                raise FileNotFoundError("Nenhum arquivo do ADM foi encontrado na pasta Downloads.")

            if callback_progresso:
                callback_progresso(0.75, "Organizando arquivos baixados...")

            for i, arquivo in enumerate(arquivos_encontrados, 1):
                renomear_arquivo(arquivo, pasta_trabalho, i)

            if modo_execucao == "download":
                arquivos_temp = sorted(pasta_trabalho.glob("Temp_Relatorio_*"))
                arquivos_saida = [str(a) for a in arquivos_temp]
                pasta_final = pasta_trabalho

                if destino_envio:
                    destino_envio.mkdir(parents=True, exist_ok=True)
                    arquivos_saida = []
                    for arquivo_tmp in arquivos_temp:
                        destino_tmp = destino_envio / arquivo_tmp.name
                        shutil.copy2(str(arquivo_tmp), str(destino_tmp))
                        arquivos_saida.append(str(destino_tmp))
                    pasta_final = destino_envio

                return {
                    "arquivo_principal": arquivos_saida[0] if arquivos_saida else None,
                    "arquivos_saida": arquivos_saida,
                    "pasta_final": str(pasta_final),
                    "mensagem": "Download concluído com sucesso.",
                }

        if arquivo_entrada and precisa_tratamento and not precisa_download:
            origem_manual = Path(arquivo_entrada)
            if not origem_manual.exists():
                raise FileNotFoundError(f"Arquivo de entrada não encontrado: {origem_manual}")
            nome_temp = f"Temp_Relatorio_manual{origem_manual.suffix}"
            destino_temp = pasta_trabalho / nome_temp
            shutil.copy2(str(origem_manual), str(destino_temp))

        if precisa_tratamento:
            data_execucao = datetime.now().strftime("%d-%m-%Y")
            arquivo_final_path = pasta_trabalho / f"{data_execucao}.xlsx"

            enviar_para = None
            criar_backup = False
            if destino_envio:
                enviar_para = str(destino_envio)
            elif modo_execucao in {"tratamento_envio", "arquivo_tratamento_envio", "completo"}:
                enviar_para = str(destino_padrao)
                criar_backup = modo_execucao == "completo" and not base_automacao

            checar_parada()
            resultado_consolidacao = consolidar_varios_arquivos(
                pasta_trabalho,
                arquivo_final_path,
                callback_progresso=callback_progresso,
                enviar_para=enviar_para,
                criar_backup=criar_backup,
                limpar_temporarios=True,
            )

            if modo_execucao == "completo":
                if callback_progresso:
                    callback_progresso(0.98, "Baixando base do ano passado...")
                origem_nova = Path(r"\\172.16.98.12\Relatórios Power BI\Forecast\Forecast - Antigo")
                destino_novo = destino_envio or destino_padrao
                destino_novo.mkdir(parents=True, exist_ok=True)
                buscar_e_mover_arquivo_ano_passado(origem_nova, destino_novo)

            return resultado_consolidacao

        return {
            "arquivo_principal": None,
            "arquivos_saida": [],
            "pasta_final": None,
            "mensagem": "Nada para executar com a combinação selecionada.",
        }

    except CanceladoPeloUsuario as erro_cancel:
        if callback_progresso:
            callback_progresso(0, str(erro_cancel))
        raise

    finally:
        if driver is not None:
            try:
                driver.quit()
            except Exception:
                pass