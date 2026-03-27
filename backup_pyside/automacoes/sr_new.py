import os
import base64
import shutil
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from pathlib import Path
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
import sys
from core.banco import BASE_DIR, ASSETS_DIR
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side

# ==========================================
# CONFIGURAÇÕES GMAIL
# ==========================================
SCOPES = [
    'https://www.googleapis.com/auth/gmail.readonly',
    'https://www.googleapis.com/auth/userinfo.email',
    'openid'
]

def obter_servico_gmail():
    """Realiza a autenticação e retorna o serviço da API do Gmail."""
    # O token.json fica na pasta do executável (para persistir entre as execuções)
    TOKEN_PATH = BASE_DIR / "token.json"
    # O credentials.json foi embutido no .exe, então buscamos nos ASSETS
    CREDS_PATH = ASSETS_DIR / "credentials.json"

    creds = None
    if TOKEN_PATH.exists():
        creds = Credentials.from_authorized_user_file(str(TOKEN_PATH), SCOPES)

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            if not CREDS_PATH.exists():
                raise FileNotFoundError(f"Arquivo credentials.json não encontrado em: {CREDS_PATH}")
            
            flow = InstalledAppFlow.from_client_secrets_file(str(CREDS_PATH), SCOPES)
            creds = flow.run_local_server(
                port=0, 
                success_message='Autenticação OK! Pode fechar esta aba e voltar ao Python.'
            )
        with open(TOKEN_PATH, 'w') as token:
            token.write(creds.to_json())

    service = build('gmail', 'v1', credentials=creds)
    info_service = build('oauth2', 'v2', credentials=creds)
    user_info = info_service.userinfo().get().execute()
    
    return service, user_info.get('email')

def baixar_anexos_especificos(service, pasta_alvo, nomes_procurados, data_busca=None):
    """
    Busca e baixa anexos que contenham as palavras-chave no nome.
    data_busca: string no formato 'YYYY/MM/DD' ou None (hoje).
    """
    if not os.path.exists(pasta_alvo):
        os.makedirs(pasta_alvo)

    if not data_busca:
        data_busca = datetime.now().strftime('%Y/%m/%d')
    
    query = f'has:attachment after:{data_busca}'
    
    resultado = service.users().messages().list(userId='me', q=query).execute()
    mensagens = resultado.get('messages', [])

    arquivos_baixados = {}

    for m in mensagens:
        msg_completa = service.users().messages().get(userId='me', id=m['id']).execute()
        partes = msg_completa.get('payload', {}).get('parts', [])

        for parte in partes:
            nome_arquivo = parte.get('filename')
            if nome_arquivo:
                for termo in nomes_procurados:
                    if termo.lower() in nome_arquivo.lower() and termo not in arquivos_baixados:
                        if 'body' in parte and 'attachmentId' in parte['body']:
                            anexo_id = parte['body']['attachmentId']
                            anexo = service.users().messages().attachments().get(
                                userId='me', messageId=m['id'], id=anexo_id).execute()
                            
                            dados_decodificados = base64.urlsafe_b64decode(anexo['data'])
                            
                            caminho_final = os.path.join(pasta_alvo, nome_arquivo)
                            with open(caminho_final, 'wb') as f:
                                f.write(dados_decodificados)
                            
                            print(f"✅ Arquivo encontrado e salvo: {nome_arquivo}")
                            arquivos_baixados[termo] = caminho_final
                            break
        
        if len(arquivos_baixados) == len(nomes_procurados):
            break

    return arquivos_baixados

# ==========================================
# TRATAMENTO DE DADOS
# ==========================================

def tratar_e_consolidar_bases(caminho_base_rio=None, caminho_share=None, caminho_base_principal=None, datas_filtro=None):
    print("Iniciando leitura e tratamento dos dados...")

    dfs = []
    
    # 1. EXTRAÇÃO
    if caminho_base_rio and os.path.exists(caminho_base_rio):
        df1 = pd.read_excel(caminho_base_rio, sheet_name='Planilha1')
        df1.rename(columns={'PASSAGEIRO': 'PAX'}, inplace=True)
        
        # Normalização prévia para filtro de destino (para não perder variações como 'RIO')
        df1['DESTINO'] = df1['DESTINO'].astype(str).str.strip().str.upper()
        rio_variants = ['RIO', 'RJ', 'RIO JANEIRO', 'RIO DE JANEIRO', 'RIO DE JANERIO']
        df1 = df1[df1['DESTINO'].isin(rio_variants)].copy()
        
        if 'ORIGEM' not in df1.columns:
            df1['ORIGEM'] = 'SÃO PAULO'
        dfs.append(df1)
        print(f"✅ Base RIO carregada: {len(df1)} linhas.")

    if caminho_share and os.path.exists(caminho_share):
        df2 = pd.read_excel(caminho_share, sheet_name='Base Relatorio   RIO x SAO')
        if 'DATA SP' in df2.columns:
            df2['DATA'] = df2['DATA SP']
        
        cols_df2 = ['Nº Mês', 'MÊS', 'EMPRESA', 'ORIGEM', 'DESTINO', 'SERVIÇO', 'HORÁRIO', 'PAX', 'DATA']
        df2 = df2[[c for c in cols_df2 if c in df2.columns]]
        dfs.append(df2)
        print(f"✅ Base Share carregada: {len(df2)} linhas.")

    if not dfs:
        print("⚠️ Nenhum dado encontrado para as bases especificadas.")
        return

    df_consolidado = pd.concat(dfs, ignore_index=True)

    # FILTRO DE DATAS
    if datas_filtro:
        df_consolidado['DATA'] = pd.to_datetime(df_consolidado['DATA'], errors='coerce')
        datas_limite = pd.to_datetime(datas_filtro)
        df_consolidado = df_consolidado[df_consolidado['DATA'].isin(datas_limite)].copy()
        
        if df_consolidado.empty:
            print("⚠️ Nenhum dado encontrado para as datas especificadas.")
            return

    # 2. LIMPEZA E REGRAS DE NEGÓCIO
    df_consolidado['EMPRESA'] = df_consolidado['EMPRESA'].astype(str)
    
    # Normalização básica de cidades antes do tratamento das tags
    for col in ['DESTINO', 'ORIGEM']:
        df_consolidado[col] = df_consolidado[col].str.strip().str.upper()
        # Padroniza variações de Rio de Janeiro
        df_consolidado.loc[df_consolidado[col].isin(['RIO', 'RJ', 'RIO JANEIRO', 'RIO DE JANERIO']), col] = 'RIO DE JANEIRO'
        # Padroniza variações de São Paulo
        df_consolidado.loc[df_consolidado[col].isin(['SP', 'SAO PAULO', 'SÃO PAULO']), col] = 'SÃO PAULO'

    # --- REGRAS ESPECÍFICAS DE LOCALIZAÇÃO (BARRA FUNDA E CATARINENSE) ---
    mask_catarinense = df_consolidado['EMPRESA'].str.contains('CATARINENSE', case=False, na=False)
    mask_bf = df_consolidado['EMPRESA'].str.contains(r'\(BF\)|\(BAF\)', na=False, regex=True, case=False)
    
    # Regra Barra Funda: Se houver tag (BF) ou (BAF) e a cidade for SÃO PAULO, expande para BARRA FUNDA
    for col in ['DESTINO', 'ORIGEM']:
        df_consolidado.loc[mask_bf & (df_consolidado[col] == 'SÃO PAULO'), col] = 'SÃO PAULO (BARRA FUNDA)'

    # Regra Específica CATARINENSE: Se não tiver tag BF/BAF, garantir que seja SÃO PAULO limpo
    for col in ['DESTINO', 'ORIGEM']:
        df_consolidado.loc[mask_catarinense & ~mask_bf & (df_consolidado[col] == 'SÃO PAULO (BARRA FUNDA)'), col] = 'SÃO PAULO'
    
    # Correção específica para Itapemirim -> Kaissara
    df_consolidado.loc[(df_consolidado['EMPRESA'].str.strip().str.upper() == 'ITAPEMIRIM'), 'EMPRESA'] = 'KAISSARA'

    # Limpeza dos nomes das empresas (remove tags entre parênteses)
    df_consolidado['EMPRESA'] = df_consolidado['EMPRESA'].str.replace(r'\s*\(.*?\)\s*', '', regex=True)
    df_consolidado['EMPRESA'] = df_consolidado['EMPRESA'].str.normalize('NFKD').str.encode('ascii', errors='ignore').str.decode('utf-8').str.upper().str.strip()

    mapa_regras = {
        '1001': {'GRUPO': 'JCA', 'MODALIDADE': 'RODOVIARIO'},
        'AGUIA BRANCA': {'GRUPO': 'AGUIA BRANCA', 'MODALIDADE': 'RODOVIARIO'},
        'EXPRESSO DO SUL': {'GRUPO': 'JCA', 'MODALIDADE': 'RODOVIARIO'},
        'CATARINENSE': {'GRUPO': 'JCA', 'MODALIDADE': 'RODOVIARIO'},
        'PENHA': {'GRUPO': 'COMPORTE', 'MODALIDADE': 'RODOVIARIO'},
        'AGUIA FLEX': {'GRUPO': 'AGUIA BRANCA', 'MODALIDADE': 'DIGITAL'},
        'KAISSARA': {'GRUPO': 'SUZANTUR', 'MODALIDADE': 'RODOVIARIO'},
        'WEMOBI': {'GRUPO': 'JCA', 'MODALIDADE': 'DIGITAL'},
        'ADAMANTINA': {'GRUPO': 'ADAMANTINA', 'MODALIDADE': 'RODOVIARIO'},
        'FLIXBUS': {'GRUPO': 'FLIXBUS', 'MODALIDADE': 'DIGITAL'},
        'RIO DOCE': {'GRUPO': 'RIO DOCE', 'MODALIDADE': 'RODOVIARIO'},
        'NOTAVEL': {'GRUPO': 'NOTÁVEL', 'MODALIDADE': 'RODOVIARIO'}
    }
    
    df_consolidado['GRUPO'] = df_consolidado['EMPRESA'].map(lambda x: mapa_regras.get(x, {}).get('GRUPO', 'OUTROS'))
    df_consolidado['MODALIDADE'] = df_consolidado['EMPRESA'].map(lambda x: mapa_regras.get(x, {}).get('MODALIDADE', 'OUTROS'))

    empresas_nao_mapeadas = df_consolidado[df_consolidado['GRUPO'] == 'OUTROS']['EMPRESA'].unique()
    if len(empresas_nao_mapeadas) > 0:
        raise ValueError(f"🚨 ERRO CRÍTICO: As seguintes empresas não estão no mapa: {empresas_nao_mapeadas}")

    df_consolidado = df_consolidado[df_consolidado['PAX'] <= 69].copy()
    df_consolidado.loc[df_consolidado['PAX'] == 69, 'PAX'] = 68

    df_consolidado['IPV'] = np.where(df_consolidado['PAX'] > 54, 
                                     df_consolidado['PAX'] / 68, 
                                     df_consolidado['PAX'] / 54)
    df_consolidado['IPV'] = df_consolidado['IPV'].clip(upper=1.0)

    df_consolidado['DATA'] = pd.to_datetime(df_consolidado['DATA'])
    df_consolidado['HORÁRIO'] = pd.to_datetime(df_consolidado['HORÁRIO'].astype(str), errors='coerce').apply(lambda x: x.time() if pd.notnull(x) else None)
    
    df_consolidado['Nº Mês'] = df_consolidado['DATA'].dt.month
    df_consolidado['Ano'] = df_consolidado['DATA'].dt.year
    df_consolidado['SEMANA'] = df_consolidado['DATA'].dt.isocalendar().week 
    
    meses_pt = {1:'01-JANEIRO', 2:'02-FEVEREIRO', 3:'03-MARÇO', 4:'04-ABRIL', 
                5:'05-MAIO', 6:'06-JUNHO', 7:'07-JULHO', 8:'08-AGOSTO', 
                9:'09-SETEMBRO', 10:'10-OUTUBRO', 11:'11-NOVEMBRO', 12:'12-DEZEMBRO'}
    df_consolidado['MÊS'] = df_consolidado['Nº Mês'].map(meses_pt)
    df_consolidado['DATA'] = df_consolidado['DATA'].dt.date

    # 4. EXPORTAÇÃO E FORMATAÇÃO (EVITANDO DUPLICATAS)
    print("Verificando duplicatas e preparando salvamento...")
    ordem_colunas = ['DATA', 'Nº Mês', 'MÊS', 'EMPRESA', 'ORIGEM', 'DESTINO', 'SERVIÇO', 'HORÁRIO', 'PAX', 'SEMANA', 'IPV', 'GRUPO', 'MODALIDADE', 'Ano']
    
    if 'SERVIÇO' not in df_consolidado.columns:
        df_consolidado['SERVIÇO'] = 1
    
    # Garante que SERVIÇO seja inteiro (sem decimais)
    df_consolidado['SERVIÇO'] = pd.to_numeric(df_consolidado['SERVIÇO'], errors='coerce').fillna(1).astype(int)

    df_final = df_consolidado[ordem_colunas]

    # --- LÓGICA DE PREVENÇÃO DE DUPLICATAS ---
    try:
        colunas_checar = ['DATA', 'HORÁRIO', 'EMPRESA', 'ORIGEM', 'DESTINO', 'PAX', 'IPV', 'SERVIÇO', 'Ano']
        df_existente = pd.read_excel(caminho_base_principal, sheet_name='Base Relatorio   RIO x SAO', usecols=lambda x: x in colunas_checar)
        
        # Função para gerar chave única de forma robusta
        def gerar_chave(row):
            d = str(row['DATA'])
            h = str(row['HORÁRIO']).split('.')[0] if pd.notnull(row['HORÁRIO']) else "None"
            e = str(row['EMPRESA']).strip().upper()
            o = str(row['ORIGEM']).strip().upper()
            dest = str(row['DESTINO']).strip().upper()
            # Garante que PAX e IPV sejam strings normalizadas
            try:
                p = str(int(float(row['PAX'])))
            except:
                p = str(row['PAX'])
            try:
                ipv = str(round(float(row['IPV']), 4))
            except:
                ipv = str(row['IPV'])
            
            # Normalização de SERVIÇO na chave para evitar 1 vs 1.0
            try:
                srv = str(int(float(row.get('SERVIÇO', 1))))
            except:
                srv = str(row.get('SERVIÇO', '1')).strip().upper()
                
            ano = str(row.get('Ano', '')).strip()
            return f"{d}|{h}|{e}|{o}|{dest}|{p}|{ipv}|{srv}|{ano}"

        # Lemos e padronizamos a base existente
        df_existente['DATA'] = pd.to_datetime(df_existente['DATA']).dt.date
        chaves_existentes = set(df_existente.apply(gerar_chave, axis=1))

        # Criamos a mesma chave para o df_final
        df_final_temp = df_final.copy()
        df_final_temp['KEY'] = df_final_temp.apply(gerar_chave, axis=1)
        
        # Filtramos o df_final removendo o que já existe na base
        df_final = df_final[~df_final_temp['KEY'].isin(chaves_existentes)].copy()
        
        if df_final.empty:
            print("ℹ️ Todos os dados processados já constam na base principal. Nada a adicionar.")
            return
        else:
            print(f"📌 Adicionando {len(df_final)} novas linhas inéditas...")
            
    except Exception as e:
        print(f"⚠️ Não foi possível verificar duplicas de forma detalhada: {e}. Prosseguindo com precaução.")

    book = load_workbook(caminho_base_principal)
    sheet = book['Base Relatorio   RIO x SAO']
    start_row = sheet.max_row + 1

    estilo_borda = Side(border_style="thin", color="000000")
    borda_fina = Border(top=estilo_borda, left=estilo_borda, right=estilo_borda, bottom=estilo_borda)
    alinhamento_centro = Alignment(horizontal='center', vertical='center')

    for row_idx, row_data in enumerate(df_final.values, start=start_row):
        for col_idx, value in enumerate(row_data, start=1):
            cell = sheet.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = alinhamento_centro
            nome_coluna_atual = ordem_colunas[col_idx-1]
            if nome_coluna_atual != 'Ano':
                cell.border = borda_fina
            if nome_coluna_atual == 'DATA':
                cell.number_format = 'dd/mmm' 
            elif nome_coluna_atual == 'HORÁRIO':
                cell.number_format = 'hh:mm'
            elif nome_coluna_atual == 'IPV':
                cell.number_format = '0%'

    book.save(caminho_base_principal)
    print(f"✅ Sucesso! O arquivo foi atualizado com novos dados em: {caminho_base_principal}")

# ==========================================
# INTEGRAÇÃO COM A UI
# ==========================================

def executar_sr(
    id_usuario,
    e_ini,
    e_fim,
    b_ini,
    b_fim,
    callback_progresso=None,
    hook_cancelamento=None,
    modo_execucao="completo",
    pasta_destino=None,
    arquivo_entrada=None,
    base_automacao=None,
):
    """Executa SR com suporte a modos modulares sem quebrar compatibilidade."""

    modos_validos = {
        "completo",
        "download",
        "download_tratamento",
        "tratamento",
        "tratamento_envio",
        "arquivo_tratamento",
        "arquivo_envio",
        "arquivo_tratamento_envio",
    }
    if modo_execucao not in modos_validos:
        raise ValueError("Modo selecionado não é suportado para esta automação SR.")

    try:
        pasta_downloads = Path.home() / "Downloads"
        destino_envio = Path(pasta_destino) if pasta_destino else None

        if base_automacao:
            caminho_base = Path(base_automacao)
            if caminho_base.suffix.lower() in {".xlsx", ".xls", ".xlsm"}:
                caminho_base.parent.mkdir(parents=True, exist_ok=True)
                base_referencia_arquivo = caminho_base
            else:
                caminho_base.mkdir(parents=True, exist_ok=True)
                base_referencia_arquivo = caminho_base / "Base RIO x SAO.xlsx"
        else:
            base_referencia_arquivo = pasta_downloads / "Base RIO x SAO.xlsx"

        pasta_saida_final = destino_envio or base_referencia_arquivo.parent
        pasta_saida_final.mkdir(parents=True, exist_ok=True)
        arquivo_saida_principal = pasta_saida_final / base_referencia_arquivo.name

        def montar_lista_datas_base():
            d_ini = datetime.strptime(b_ini, "%d/%m/%Y")
            d_fim = datetime.strptime(b_fim, "%d/%m/%Y")
            delta = d_fim - d_ini
            return [(d_ini + timedelta(days=i)).strftime('%Y-%m-%d') for i in range(delta.days + 1)]

        def preparar_base_saida():
            if arquivo_saida_principal.exists():
                return
            if base_referencia_arquivo.exists() and base_referencia_arquivo.resolve() != arquivo_saida_principal.resolve():
                shutil.copy2(str(base_referencia_arquivo), str(arquivo_saida_principal))

        if modo_execucao == "arquivo_envio":
            if not arquivo_entrada:
                raise ValueError("Selecione um arquivo para envio.")
            origem_arquivo = Path(arquivo_entrada)
            if not origem_arquivo.exists():
                raise FileNotFoundError(f"Arquivo não encontrado: {origem_arquivo}")
            pasta_final = destino_envio or pasta_downloads
            pasta_final.mkdir(parents=True, exist_ok=True)
            destino_final = pasta_final / origem_arquivo.name
            if origem_arquivo.resolve() != destino_final.resolve():
                shutil.copy2(str(origem_arquivo), str(destino_final))
            return {
                "arquivo_principal": str(destino_final),
                "arquivos_saida": [str(destino_final)],
                "pasta_final": str(pasta_final),
                "mensagem": "Arquivo enviado com sucesso.",
            }

        if modo_execucao in {"tratamento", "tratamento_envio", "arquivo_tratamento", "arquivo_tratamento_envio"}:
            if not arquivo_entrada:
                raise ValueError("Selecione um arquivo já baixado para realizar o tratamento no SR.")

            caminho_arquivo = Path(arquivo_entrada)
            if not caminho_arquivo.exists():
                raise FileNotFoundError(f"Arquivo não encontrado: {caminho_arquivo}")

            if callback_progresso:
                callback_progresso(0.55, "Iniciando tratamento da base informada...")

            preparar_base_saida()
            tratar_e_consolidar_bases(
                caminho_base_rio=str(caminho_arquivo),
                caminho_share=None,
                caminho_base_principal=str(arquivo_saida_principal),
                datas_filtro=montar_lista_datas_base(),
            )

            if callback_progresso:
                callback_progresso(1.0, "Concluído com sucesso!")

            return {
                "arquivo_principal": str(arquivo_saida_principal),
                "arquivos_saida": [str(arquivo_saida_principal)],
                "pasta_final": str(pasta_saida_final),
                "mensagem": "Tratamento SR concluído com sucesso.",
            }

        if callback_progresso:
            callback_progresso(0.1, "Conectando ao Gmail...")
        service, _ = obter_servico_gmail()

        if hook_cancelamento and hook_cancelamento():
            return {
                "arquivo_principal": None,
                "arquivos_saida": [],
                "pasta_final": None,
                "mensagem": "Processo cancelado pelo usuário.",
            }

        termos_esperados = ["BASE RIO", "Share - Mercado RIO"]

        data_busca_gmail = datetime.strptime(e_ini, "%d/%m/%Y").strftime("%Y/%m/%d")

        if callback_progresso:
            callback_progresso(0.3, f"Buscando anexos desde {e_ini}...")
        arquivos = baixar_anexos_especificos(service, pasta_downloads, termos_esperados, data_busca=data_busca_gmail)

        if hook_cancelamento and hook_cancelamento():
            return {
                "arquivo_principal": None,
                "arquivos_saida": [],
                "pasta_final": None,
                "mensagem": "Processo cancelado pelo usuário.",
            }

        if len(arquivos) == 0:
            raise Exception(f"Nenhum dos arquivos esperados foi encontrado no e-mail: {termos_esperados}")

        arquivos_baixados = [str(v) for v in arquivos.values() if v]

        if modo_execucao == "download":
            if destino_envio:
                destino_envio.mkdir(parents=True, exist_ok=True)
                arquivos_copiados = []
                for item in arquivos_baixados:
                    origem_item = Path(item)
                    destino_item = destino_envio / origem_item.name
                    shutil.copy2(str(origem_item), str(destino_item))
                    arquivos_copiados.append(str(destino_item))
                arquivos_baixados = arquivos_copiados

            return {
                "arquivo_principal": arquivos_baixados[0] if arquivos_baixados else None,
                "arquivos_saida": arquivos_baixados,
                "pasta_final": str(destino_envio or pasta_downloads),
                "mensagem": "Download de anexos concluído com sucesso.",
            }

        if callback_progresso:
            callback_progresso(0.6, "Iniciando processamento da base...")

        preparar_base_saida()
        tratar_e_consolidar_bases(
            caminho_base_rio=arquivos.get("BASE RIO"),
            caminho_share=arquivos.get("Share - Mercado RIO"),
            caminho_base_principal=str(arquivo_saida_principal),
            datas_filtro=montar_lista_datas_base(),
        )

        if callback_progresso:
            callback_progresso(1.0, "Concluído com sucesso!")

        return {
            "arquivo_principal": str(arquivo_saida_principal),
            "arquivos_saida": [str(arquivo_saida_principal)],
            "pasta_final": str(pasta_saida_final),
            "mensagem": "Processo SR concluído com sucesso.",
        }

    except Exception as e:
        raise Exception(f"Erro SR: {str(e)}")

# ==========================================
# EXECUÇÃO MANUAL (TESTE)
# ==========================================
if __name__ == '__main__':
    # No manual, usamos hoje para e-mail e ontem para base
    hoje = datetime.now().strftime('%d/%m/%Y')
    ontem = (datetime.now() - timedelta(days=1)).strftime('%d/%m/%Y')
    
    executar_sr(1, hoje, hoje, ontem, ontem)
