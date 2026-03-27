"""
XLS → CSV Conversor  |  App Web Local
Rode: python xls_para_csv_app.py
Abre automaticamente no navegador em http://localhost:5000
Dependências: pip install flask openpyxl xlrd
"""

import os, csv, json, time, threading, webbrowser
from pathlib import Path
from flask import Flask, request, jsonify, send_file, Response

try:
    import openpyxl; HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import xlrd; HAS_XLRD = True
except ImportError:
    HAS_XLRD = False

app = Flask(__name__)
UPLOAD_DIR = Path("_uploads_tmp")
UPLOAD_DIR.mkdir(exist_ok=True)

# ── Progresso global (simples, single-user local app) ──────────────────────
_progress = {"pct": 0, "status": "idle", "msg": "", "output": ""}
_lock = threading.Lock()

def set_prog(pct, msg, status="running"):
    with _lock:
        _progress.update(pct=pct, msg=msg, status=status)


# ── Leitura streaming ───────────────────────────────────────────────────────

def iter_rows_xlsx(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    for row in ws.iter_rows(values_only=True):
        yield [("" if v is None else str(v)) for v in row]
    wb.close()


def iter_rows_xls(path):
    wb = xlrd.open_workbook(str(path), on_demand=True)
    ws = wb.sheet_by_index(0)
    for i in range(ws.nrows):
        row = []
        for cell in ws.row(i):
            ct = cell.ctype
            if ct == xlrd.XL_CELL_EMPTY:
                row.append("")
            elif ct == xlrd.XL_CELL_NUMBER:
                v = cell.value
                row.append(str(int(v)) if v == int(v) else str(v))
            elif ct == xlrd.XL_CELL_DATE:
                import datetime
                dt = xlrd.xldate_as_datetime(cell.value, wb.datemode)
                row.append(dt.strftime("%d/%m/%Y") if not (dt.hour or dt.minute) else dt.strftime("%d/%m/%Y %H:%M:%S"))
            elif ct == xlrd.XL_CELL_BOOLEAN:
                row.append("VERDADEIRO" if cell.value else "FALSO")
            else:
                row.append(str(cell.value))
        yield row
    wb.release_resources()


def count_rows(path):
    ext = Path(path).suffix.lower()
    try:
        if ext == ".xlsx":
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            n = wb.active.max_row or 0
            wb.close(); return n
        else:
            wb = xlrd.open_workbook(str(path), on_demand=True)
            n = wb.sheet_by_index(0).nrows
            wb.release_resources(); return n
    except Exception:
        return 0


def get_iter(path):
    ext = Path(path).suffix.lower()
    if ext == ".xlsx":
        if not HAS_OPENPYXL: raise RuntimeError("openpyxl não instalado")
        return iter_rows_xlsx(path)
    elif ext in (".xls",):
        if not HAS_XLRD: raise RuntimeError("xlrd não instalado")
        return iter_rows_xls(path)
    raise ValueError(f"Formato não suportado: {ext}")


# ── Conversão em thread ─────────────────────────────────────────────────────

def do_convert(paths, output_path, separator, bom, merge_headers):
    try:
        set_prog(2, "Calculando total de linhas…")
        total = sum(count_rows(p) for p in paths) or 1

        done = 0
        first_file = True
        encoding = "utf-8-sig" if bom else "utf-8"

        with open(output_path, "w", newline="", encoding=encoding) as fout:
            writer = csv.writer(fout, delimiter=separator, quoting=csv.QUOTE_MINIMAL)
            for i, path in enumerate(paths):
                fname = Path(path).name
                set_prog(int(done / total * 92) + 3, f"Processando: {fname}")
                skip_header = merge_headers and not first_file
                first_row = True
                for row in get_iter(path):
                    if skip_header and first_row:
                        first_row = False; done += 1; continue
                    first_row = False
                    writer.writerow(row)
                    done += 1
                    if done % 1000 == 0:
                        pct = min(int(done / total * 92) + 3, 95)
                        set_prog(pct, f"Processando: {fname} — {done:,} linhas…")
                first_file = False

        with _lock:
            _progress.update(pct=100, msg="Conversão concluída!", status="done",
                             output=str(output_path))
    except Exception as e:
        with _lock:
            _progress.update(pct=0, msg=str(e), status="error")
    finally:
        for p in paths:
            try: Path(p).unlink()
            except: pass


# ── Rotas Flask ─────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return Response(HTML, mimetype="text/html")


@app.route("/convert", methods=["POST"])
def convert():
    files = request.files.getlist("files")
    if not files:
        return jsonify(error="Nenhum arquivo enviado"), 400

    separator = request.form.get("separator", ";")
    bom = request.form.get("bom", "true") == "true"
    merge = request.form.get("merge_headers", "true") == "true"
    out_name = request.form.get("output_name", "resultado.csv")
    if not out_name.endswith(".csv"):
        out_name += ".csv"

    saved = []
    for f in files:
        dest = UPLOAD_DIR / f.filename
        f.save(dest)
        saved.append(str(dest))

    out_path = UPLOAD_DIR / out_name
    with _lock:
        _progress.update(pct=0, msg="Iniciando…", status="running", output="")

    t = threading.Thread(target=do_convert,
                         args=(saved, out_path, separator, bom, merge), daemon=True)
    t.start()
    return jsonify(ok=True)


@app.route("/progress")
def progress():
    with _lock:
        return jsonify(**_progress)


@app.route("/download")
def download():
    with _lock:
        out = _progress.get("output", "")
    if not out or not Path(out).exists():
        return "Arquivo não encontrado", 404
    return send_file(out, as_attachment=True, download_name=Path(out).name)


# ── HTML / CSS / JS embutido ────────────────────────────────────────────────

HTML = r"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>XLS → CSV</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link href="https://fonts.googleapis.com/css2?family=DM+Sans:wght@300;400;500&family=DM+Mono:wght@400;500&display=swap" rel="stylesheet">
<style>
*{box-sizing:border-box;margin:0;padding:0}
:root{
  --bg:#0e0e10;--surface:#17171a;--surface2:#1e1e22;--surface3:#26262c;
  --accent:#5b5ef4;--accent2:#7c7ff7;--green:#22c97a;--red:#f4526a;--amber:#f4a732;
  --text:#f0f0f5;--text2:#8888a0;--text3:#555568;--border:#2e2e38;--border2:#3a3a48;
  --radius:12px;--font:'DM Sans',sans-serif;--mono:'DM Mono',monospace
}
body{background:var(--bg);font-family:var(--font);color:var(--text);min-height:100vh;display:flex;align-items:flex-start;justify-content:center;padding:32px 16px 64px}
.app{width:100%;max-width:640px;display:flex;flex-direction:column;gap:16px}
.header{display:flex;align-items:center;justify-content:space-between;padding:0 2px}
.header-left{display:flex;align-items:center;gap:10px}
.logo-icon{width:32px;height:32px;background:var(--accent);border-radius:8px;display:flex;align-items:center;justify-content:center}
.logo-icon svg{width:18px;height:18px;fill:none;stroke:#fff;stroke-width:2;stroke-linecap:round;stroke-linejoin:round}
.app-title{font-size:15px;font-weight:500;letter-spacing:-.2px}
.app-sub{font-size:12px;color:var(--text2)}
.badge{font-size:11px;font-weight:500;background:var(--surface3);border:1px solid var(--border2);color:var(--text2);padding:3px 8px;border-radius:20px;font-family:var(--mono)}
.dropzone{border:1.5px dashed var(--border2);border-radius:var(--radius);background:var(--surface);padding:36px 24px;display:flex;flex-direction:column;align-items:center;gap:10px;cursor:pointer;transition:all .18s}
.dropzone:hover,.dropzone.drag{border-color:var(--accent2);background:rgba(91,94,244,.05)}
.dz-icon{width:40px;height:40px;border-radius:10px;background:var(--surface3);display:flex;align-items:center;justify-content:center;border:1px solid var(--border2)}
.dz-icon svg{width:20px;height:20px;fill:none;stroke:var(--text2);stroke-width:1.8;stroke-linecap:round;stroke-linejoin:round}
.dz-label{font-size:14px;font-weight:500}
.dz-sub{font-size:12px;color:var(--text3)}
.dz-chips{display:flex;gap:6px}
.chip{font-size:11px;font-family:var(--mono);background:var(--surface3);border:1px solid var(--border2);color:var(--text2);padding:3px 8px;border-radius:6px}
.dz-btn{margin-top:4px;font-size:13px;font-weight:500;background:transparent;border:1px solid var(--border2);color:var(--text2);padding:7px 18px;border-radius:8px;cursor:pointer;font-family:var(--font);transition:all .15s}
.dz-btn:hover{border-color:var(--accent2);color:var(--accent2)}
.section-title{font-size:11px;font-weight:500;color:var(--text3);letter-spacing:.6px;text-transform:uppercase;padding:0 2px}
.file-list{display:flex;flex-direction:column;gap:6px}
.file-item{background:var(--surface);border:1px solid var(--border);border-radius:var(--radius);padding:12px 14px;display:flex;align-items:center;gap:12px;animation:sI .2s ease;transition:opacity .2s,transform .2s}
@keyframes sI{from{opacity:0;transform:translateY(-6px)}}
.file-ext{width:34px;height:34px;border-radius:7px;display:flex;align-items:center;justify-content:center;font-size:10px;font-weight:500;font-family:var(--mono);flex-shrink:0}
.ext-xlsx{background:rgba(34,201,122,.12);color:var(--green);border:1px solid rgba(34,201,122,.25)}
.ext-xls{background:rgba(244,167,50,.12);color:var(--amber);border:1px solid rgba(244,167,50,.25)}
.file-info{flex:1;min-width:0}
.file-name{font-size:13px;font-weight:500;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.file-meta{font-size:11px;color:var(--text3);margin-top:2px;font-family:var(--mono)}
.fstatus{font-size:11px;display:flex;align-items:center;gap:5px;flex-shrink:0}
.fstatus.wait{color:var(--text3)}.fstatus.done{color:var(--green)}.fstatus.active{color:var(--accent2)}
.dot{width:6px;height:6px;border-radius:50%;background:currentColor;flex-shrink:0}
.dot.pulse{animation:p 1.2s infinite}
@keyframes p{0%,100%{opacity:1}50%{opacity:.3}}
.rm-btn{width:26px;height:26px;border-radius:6px;background:transparent;border:none;cursor:pointer;display:flex;align-items:center;justify-content:center;flex-shrink:0;transition:background .15s;color:var(--text3)}
.rm-btn:hover{background:rgba(244,82,106,.15);color:var(--red)}
.rm-btn svg{width:14px;height:14px;fill:none;stroke:currentColor;stroke-width:2;stroke-linecap:round}
.stats-row{display:grid;grid-template-columns:repeat(3,1fr);gap:10px}
.stat-card{background:var(--surface);border:1px solid var(--border);border-radius:var(--radius);padding:14px}
.stat-val{font-size:22px;font-weight:300;font-family:var(--mono);letter-spacing:-1px}
.stat-label{font-size:11px;color:var(--text3);margin-top:4px}
.config-card{background:var(--surface);border:1px solid var(--border);border-radius:var(--radius);padding:16px;display:flex;flex-direction:column;gap:12px}
.cfg-title{font-size:11px;font-weight:500;color:var(--text3);letter-spacing:.6px;text-transform:uppercase}
.cfg-row{display:flex;align-items:center;gap:10px}
.cfg-label{font-size:12px;color:var(--text2);white-space:nowrap;min-width:90px}
.cfg-input{flex:1;background:var(--surface2);border:1px solid var(--border);border-radius:8px;padding:8px 12px;font-size:13px;font-family:var(--mono);color:var(--text);outline:none;transition:border-color .15s}
.cfg-input:focus{border-color:var(--accent2)}
.cfg-divider{height:1px;background:var(--border)}
.sep-btns{display:flex;gap:6px;flex-wrap:wrap}
.sep-btn{font-size:12px;font-weight:500;font-family:var(--font);background:transparent;border:1px solid var(--border2);color:var(--text2);padding:6px 14px;border-radius:8px;cursor:pointer;transition:all .15s}
.sep-btn.active,.sep-btn:hover{border-color:var(--accent2);color:var(--accent2)}
.opts-row{display:flex;gap:16px;flex-wrap:wrap}
.opt-toggle{display:flex;align-items:center;gap:8px;cursor:pointer;user-select:none}
.opt-toggle input[type=checkbox]{display:none}
.trk{width:32px;height:18px;border-radius:9px;background:var(--surface3);border:1px solid var(--border2);position:relative;transition:all .2s;flex-shrink:0}
.opt-toggle input:checked~.trk{background:var(--accent);border-color:var(--accent)}
.thm{width:12px;height:12px;border-radius:50%;background:var(--text3);position:absolute;top:2px;left:2px;transition:all .2s}
.opt-toggle input:checked~.trk .thm{background:#fff;left:16px}
.opt-text{font-size:12px;color:var(--text2)}
.prog-card{background:var(--surface);border:1px solid var(--border);border-radius:var(--radius);padding:20px;display:flex;flex-direction:column;gap:14px}
.prog-header{display:flex;justify-content:space-between;align-items:center}
.prog-title{font-size:13px;font-weight:500}
.prog-pct{font-size:13px;font-family:var(--mono);color:var(--accent2)}
.prog-track{height:4px;background:var(--surface3);border-radius:4px;overflow:hidden}
.prog-fill{height:100%;border-radius:4px;background:linear-gradient(90deg,var(--accent),var(--accent2));transition:width .4s ease}
.prog-fill.complete{background:linear-gradient(90deg,#22c97a,#4de8a0)}
.prog-steps{display:flex;flex-direction:column;gap:8px}
.pstep{display:flex;align-items:center;gap:10px;font-size:12px}
.sicon{width:18px;height:18px;border-radius:50%;display:flex;align-items:center;justify-content:center;flex-shrink:0}
.sicon.done{background:rgba(34,201,122,.15);border:1px solid rgba(34,201,122,.35)}
.sicon.done svg{width:10px;height:10px;fill:none;stroke:var(--green);stroke-width:2.5;stroke-linecap:round;stroke-linejoin:round}
.sicon.active{background:rgba(91,94,244,.15);border:1px solid rgba(91,94,244,.35)}
.sicon.active .spin{width:10px;height:10px;border:1.5px solid transparent;border-top-color:var(--accent2);border-radius:50%;animation:spin .7s linear infinite}
.sicon.pend{background:var(--surface3);border:1px solid var(--border2)}
.sicon.pend .dot-s{width:5px;height:5px;border-radius:50%;background:var(--text3);margin:auto}
@keyframes spin{to{transform:rotate(360deg)}}
.stext{color:var(--text2)}.stext.done-t{color:var(--text)}
.stime{margin-left:auto;font-family:var(--mono);font-size:11px;color:var(--text3)}
.action-row{display:flex;gap:10px}
.btn-primary{flex:1;padding:13px;border-radius:10px;font-size:14px;font-weight:500;font-family:var(--font);cursor:pointer;border:none;background:var(--accent);color:#fff;transition:all .2s;letter-spacing:-.1px}
.btn-primary:hover:not(:disabled){background:var(--accent2);transform:translateY(-1px)}
.btn-primary:active{transform:translateY(0)}
.btn-primary:disabled{background:var(--surface3);color:var(--text3);cursor:not-allowed}
.btn-secondary{padding:13px 18px;border-radius:10px;font-size:14px;font-weight:500;font-family:var(--font);cursor:pointer;border:1px solid var(--border2);background:transparent;color:var(--text2);transition:all .15s}
.btn-secondary:hover{border-color:var(--border);color:var(--text)}
.btn-download{padding:13px 20px;border-radius:10px;font-size:14px;font-weight:500;font-family:var(--font);cursor:pointer;border:1px solid rgba(34,201,122,.4);background:rgba(34,201,122,.1);color:var(--green);transition:all .15s;text-decoration:none;display:flex;align-items:center;gap:8px}
.btn-download:hover{background:rgba(34,201,122,.18)}
.btn-download svg{width:16px;height:16px;fill:none;stroke:currentColor;stroke-width:2;stroke-linecap:round;stroke-linejoin:round}
.toast{position:fixed;bottom:24px;right:24px;background:#1a2e1e;border:1px solid rgba(34,201,122,.4);border-radius:var(--radius);padding:14px 18px;display:flex;align-items:center;gap:12px;font-size:13px;box-shadow:0 8px 32px rgba(0,0,0,.6);transform:translateY(16px);opacity:0;transition:all .3s;pointer-events:none;z-index:100}
.toast.show{transform:translateY(0);opacity:1}
.toast-icon{width:20px;height:20px;border-radius:50%;background:rgba(34,201,122,.2);display:flex;align-items:center;justify-content:center;flex-shrink:0}
.toast-icon svg{width:12px;height:12px;fill:none;stroke:var(--green);stroke-width:2.5;stroke-linecap:round;stroke-linejoin:round}
.hidden{display:none!important}
</style>
</head>
<body>
<div class="app">

  <div class="header">
    <div class="header-left">
      <div class="logo-icon">
        <svg viewBox="0 0 24 24"><path d="M14 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V8z"/><polyline points="14 2 14 8 20 8"/><line x1="8" y1="13" x2="16" y2="13"/><line x1="8" y1="17" x2="16" y2="17"/></svg>
      </div>
      <div>
        <div class="app-title">XLS → CSV</div>
        <div class="app-sub">Conversor de planilhas</div>
      </div>
    </div>
    <span class="badge">v2.0</span>
  </div>

  <div class="dropzone" id="dz" onclick="document.getElementById('fi').click()">
    <div class="dz-icon">
      <svg viewBox="0 0 24 24"><polyline points="16 16 12 12 8 16"/><line x1="12" y1="12" x2="12" y2="21"/><path d="M20.39 18.39A5 5 0 0 0 18 9h-1.26A8 8 0 1 0 3 16.3"/></svg>
    </div>
    <div class="dz-label">Arraste arquivos aqui</div>
    <div class="dz-sub">ou clique para selecionar via Windows Explorer</div>
    <div class="dz-chips">
      <span class="chip">.xlsx</span><span class="chip">.xls</span><span class="chip">múltiplos arquivos</span>
    </div>
    <button class="dz-btn" onclick="event.stopPropagation();document.getElementById('fi').click()">Selecionar arquivos</button>
    <input type="file" id="fi" accept=".xls,.xlsx" multiple style="display:none" onchange="addFiles(this.files)">
  </div>

  <div id="fsec" class="hidden" style="display:flex;flex-direction:column;gap:8px">
    <div class="section-title">Arquivos selecionados</div>
    <div class="file-list" id="flist"></div>
    <div class="stats-row">
      <div class="stat-card"><div class="stat-val" id="sv-files">0</div><div class="stat-label">arquivos</div></div>
      <div class="stat-card"><div class="stat-val" id="sv-size">0 KB</div><div class="stat-label">tamanho total</div></div>
      <div class="stat-card"><div class="stat-val" id="sv-ext">—</div><div class="stat-label">formatos</div></div>
    </div>
  </div>

  <div class="config-card">
    <div class="cfg-title">Configurações de saída</div>
    <div class="cfg-row">
      <span class="cfg-label">Arquivo CSV</span>
      <input class="cfg-input" id="out-name" value="resultado.csv" placeholder="nome-do-arquivo.csv">
    </div>
    <div class="cfg-divider"></div>
    <div class="cfg-row">
      <span class="cfg-label">Separador</span>
      <div class="sep-btns">
        <button class="sep-btn active" data-sep=";">; ponto-vírgula</button>
        <button class="sep-btn" data-sep=",">, vírgula</button>
        <button class="sep-btn" data-sep="&#9;">&#8677; tab</button>
      </div>
    </div>
    <div class="cfg-divider"></div>
    <div class="opts-row">
      <label class="opt-toggle">
        <input type="checkbox" id="opt-bom" checked>
        <div class="trk"><div class="thm"></div></div>
        <span class="opt-text">UTF-8 BOM (acentos PT-BR)</span>
      </label>
      <label class="opt-toggle">
        <input type="checkbox" id="opt-merge" checked>
        <div class="trk"><div class="thm"></div></div>
        <span class="opt-text">Mesclar cabeçalhos</span>
      </label>
    </div>
  </div>

  <div class="prog-card hidden" id="prog-card">
    <div class="prog-header">
      <span class="prog-title" id="prog-lbl">Convertendo…</span>
      <span class="prog-pct" id="prog-pct">0%</span>
    </div>
    <div class="prog-track"><div class="prog-fill" id="prog-fill" style="width:0%"></div></div>
    <div class="prog-steps">
      <div class="pstep" id="step0"><div class="sicon pend"><span class="dot-s"></span></div><span class="stext">Lendo planilhas</span><span class="stime" id="st0"></span></div>
      <div class="pstep" id="step1"><div class="sicon pend"><span class="dot-s"></span></div><span class="stext">Convertendo registros</span><span class="stime" id="st1"></span></div>
      <div class="pstep" id="step2"><div class="sicon pend"><span class="dot-s"></span></div><span class="stext">Salvando CSV</span><span class="stime" id="st2"></span></div>
    </div>
  </div>

  <div class="action-row" id="action-row">
    <button class="btn-secondary" onclick="clearAll()">Limpar</button>
    <button class="btn-primary" id="btn-conv" onclick="startConversion()">Converter para CSV</button>
  </div>

  <div id="dl-row" class="hidden">
    <a class="btn-download" href="/download" id="dl-link">
      <svg viewBox="0 0 24 24"><path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"/><polyline points="7 10 12 15 17 10"/><line x1="12" y1="15" x2="12" y2="3"/></svg>
      Baixar CSV
    </a>
  </div>

</div>

<div class="toast hidden" id="toast">
  <div class="toast-icon"><svg viewBox="0 0 24 24"><polyline points="20 6 9 17 4 12"/></svg></div>
  <div id="toast-msg">Concluído!</div>
</div>

<script>
const files=[];
let selSep=';';
let polling=null;
const startTime={};

function fmtSize(b){return b>1048576?(b/1048576).toFixed(1)+' MB':(b/1024).toFixed(0)+' KB'}

function addFiles(fl){
  Array.from(fl).forEach(f=>{
    if(!files.find(x=>x.name===f.name)){files.push(f);renderFile(f);}
  });
  updateStats();
  show('fsec');
}

function renderFile(f){
  const ext=f.name.split('.').pop().toLowerCase();
  const id='fi_'+Math.random().toString(36).slice(2);
  f._id=id;
  const el=document.createElement('div');
  el.className='file-item';el.id=id;
  el.innerHTML=`<div class="file-ext ext-${ext}">${ext.toUpperCase()}</div>
    <div class="file-info">
      <div class="file-name">${f.name}</div>
      <div class="file-meta">${fmtSize(f.size)} · aguardando</div>
    </div>
    <div class="fstatus wait" id="fs_${id}"><div class="dot"></div>aguardando</div>
    <button class="rm-btn" onclick="removeFile('${f.name}')">
      <svg viewBox="0 0 24 24"><line x1="18" y1="6" x2="6" y2="18"/><line x1="6" y1="6" x2="18" y2="18"/></svg>
    </button>`;
  document.getElementById('flist').appendChild(el);
}

function removeFile(name){
  const i=files.findIndex(f=>f.name===name);
  if(i>-1){const el=document.getElementById(files[i]._id);if(el){el.style.opacity='0';el.style.transform='translateX(8px)';setTimeout(()=>el.remove(),200);}files.splice(i,1);}
  setTimeout(()=>{updateStats();if(!files.length)hide('fsec');},220);
}

function clearAll(){
  files.length=0;
  document.getElementById('flist').innerHTML='';
  hide('fsec');hide('prog-card');hide('dl-row');
}

function updateStats(){
  const total=files.reduce((a,f)=>a+f.size,0);
  document.getElementById('sv-files').textContent=files.length;
  document.getElementById('sv-size').textContent=fmtSize(total);
  const exts=[...new Set(files.map(f=>f.name.split('.').pop().toUpperCase()))];
  document.getElementById('sv-ext').textContent=exts.length?exts.join(' / '):'—';
}

document.querySelectorAll('.sep-btn').forEach(b=>{
  b.onclick=()=>{selSep=b.dataset.sep;document.querySelectorAll('.sep-btn').forEach(x=>x.classList.toggle('active',x===b));};
});

const dz=document.getElementById('dz');
dz.addEventListener('dragover',e=>{e.preventDefault();dz.classList.add('drag');});
dz.addEventListener('dragleave',()=>dz.classList.remove('drag'));
dz.addEventListener('drop',e=>{e.preventDefault();dz.classList.remove('drag');addFiles(e.dataTransfer.files);});

function setStep(n,state,time){
  const s=document.getElementById('step'+n);
  const icon=s.querySelector('.sicon');
  icon.className='sicon '+state;
  if(state==='done')icon.innerHTML='<svg viewBox="0 0 24 24"><polyline points="20 6 9 17 4 12"/></svg>';
  else if(state==='active')icon.innerHTML='<span class="spin"></span>';
  else icon.innerHTML='<span class="dot-s"></span>';
  const txt=s.querySelector('.stext');
  txt.className='stext'+(state==='done'?' done-t':'');
  if(time)document.getElementById('st'+n).textContent=time;
}

async function startConversion(){
  if(!files.length){alert('Adicione ao menos um arquivo.');return;}
  const btn=document.getElementById('btn-conv');
  btn.disabled=true;
  hide('dl-row');
  show('prog-card');
  setStep(0,'active');setStep(1,'pend');setStep(2,'pend');
  document.getElementById('prog-fill').className='prog-fill';
  document.getElementById('prog-fill').style.width='0%';
  document.getElementById('prog-pct').textContent='0%';

  const fd=new FormData();
  files.forEach(f=>fd.append('files',f));
  fd.append('separator',selSep);
  fd.append('bom',document.getElementById('opt-bom').checked);
  fd.append('merge_headers',document.getElementById('opt-merge').checked);
  fd.append('output_name',document.getElementById('out-name').value||'resultado.csv');

  const t0=Date.now();
  await fetch('/convert',{method:'POST',body:fd});

  let lastPct=0;
  polling=setInterval(async()=>{
    const r=await fetch('/progress').then(x=>x.json());
    const pct=r.pct||0;
    const elapsed=((Date.now()-t0)/1000).toFixed(1)+'s';
    document.getElementById('prog-fill').style.width=pct+'%';
    document.getElementById('prog-pct').textContent=pct+'%';
    document.getElementById('prog-lbl').textContent=r.msg||'Convertendo…';

    if(pct>=5&&lastPct<5)setStep(0,'done',elapsed);
    if(pct>=40&&lastPct<40)setStep(1,'active');
    if(pct>=70&&lastPct<70){setStep(1,'done',elapsed);setStep(2,'active');}
    if(pct>=95&&lastPct<95)setStep(2,'done',elapsed);
    lastPct=pct;

    if(r.status==='done'){
      clearInterval(polling);
      document.getElementById('prog-fill').className='prog-fill complete';
      document.getElementById('prog-fill').style.width='100%';
      document.getElementById('prog-pct').textContent='100%';
      document.getElementById('prog-lbl').textContent='Concluído!';
      setStep(0,'done');setStep(1,'done');setStep(2,'done');
      btn.disabled=false;
      show('dl-row');
      toast(`<strong>Concluído!</strong> <span style="font-family:var(--mono);font-size:12px;color:#8888a0">${document.getElementById('out-name').value}</span> pronto para baixar.`);
    }
    if(r.status==='error'){
      clearInterval(polling);
      document.getElementById('prog-lbl').textContent='Erro: '+r.msg;
      btn.disabled=false;
    }
  },400);
}

function show(id){const el=document.getElementById(id);if(el){el.classList.remove('hidden');el.style.display='';}}
function hide(id){const el=document.getElementById(id);if(el)el.classList.add('hidden');}

function toast(html){
  const t=document.getElementById('toast');
  document.getElementById('toast-msg').innerHTML=html;
  t.classList.remove('hidden');
  t.classList.add('show');
  setTimeout(()=>t.classList.remove('show'),4500);
  setTimeout(()=>t.classList.add('hidden'),5000);
}
</script>
</body>
</html>"""


# ── Inicialização ───────────────────────────────────────────────────────────

if __name__ == "__main__":
    port = 5000
    url = f"http://localhost:{port}"
    print(f"\n  XLS → CSV Conversor")
    print(f"  Abrindo em {url}\n")
    threading.Timer(1.2, lambda: webbrowser.open(url)).start()
    app.run(host="127.0.0.1", port=port, debug=False)