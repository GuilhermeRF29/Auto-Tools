"""
Conversor de XLS/XLSX para CSV - Edição Definitiva (Ultra Modern UI)
- Interface gráfica premium com CustomTkinter (UX focada em Cards)
- Botão estilo "Dropzone" (Call to Action principal)
- Remoção da hora nas datas (formato DD/MM/YYYY)
- Saída com separador ; e encoding UTF-8 BOM
- Escrita incremental (streaming)
"""

import sys
import os
import csv
import threading
import datetime
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox

# Nova biblioteca de interface!
try:
    import customtkinter as ctk
except ImportError:
    raise RuntimeError("Por favor, instale o customtkinter rodando: pip install customtkinter")

# Suporte a .xls (legado) e .xlsx
try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import xlrd
    HAS_XLRD = True
except ImportError:
    HAS_XLRD = False


# ─────────────────────────────────────────────
# Utilidades de leitura por tipo de arquivo
# ─────────────────────────────────────────────

def iter_rows_xlsx(filepath):
    """Lê .xlsx em modo read_only (streaming), yielding uma linha por vez."""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    for row in ws.iter_rows(values_only=True):
        new_row = []
        for v in row:
            if v is None:
                new_row.append("")
            # Mágica da Data: formata sem a hora!
            elif isinstance(v, (datetime.datetime, datetime.date)):
                new_row.append(v.strftime("%d/%m/%Y"))
            else:
                new_row.append(str(v))
        yield new_row
    wb.close()

def iter_rows_xls(filepath):
    """Lê .xls via xlrd, yielding uma linha por vez."""
    wb = xlrd.open_workbook(filepath, on_demand=True)
    ws = wb.sheet_by_index(0)
    for i in range(ws.nrows):
        row = []
        for cell in ws.row(i):
            ctype = cell.ctype
            if ctype == xlrd.XL_CELL_EMPTY:
                row.append("")
            elif ctype == xlrd.XL_CELL_NUMBER:
                v = cell.value
                row.append(str(int(v)) if v == int(v) else str(v))
            elif ctype == xlrd.XL_CELL_DATE:
                dt = xlrd.xldate_as_datetime(cell.value, wb.datemode)
                row.append(dt.strftime("%d/%m/%Y"))
            elif ctype == xlrd.XL_CELL_BOOLEAN:
                row.append("VERDADEIRO" if cell.value else "FALSO")
            else:
                row.append(str(cell.value))
        yield row
    wb.release_resources()

def count_rows_xlsx(filepath):
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    n = ws.max_row or 0
    wb.close()
    return n

def count_rows_xls(filepath):
    wb = xlrd.open_workbook(filepath, on_demand=True)
    n = wb.sheet_by_index(0).nrows
    wb.release_resources()
    return n

def get_row_iterator(filepath):
    ext = Path(filepath).suffix.lower()
    if ext == ".xlsx":
        if not HAS_OPENPYXL:
            raise RuntimeError("openpyxl não instalado.")
        return iter_rows_xlsx(filepath)
    elif ext in (".xls", ".xlsb"):
        if not HAS_XLRD:
            raise RuntimeError("xlrd não instalado.")
        return iter_rows_xls(filepath)
    else:
        raise ValueError(f"Formato não suportado: {ext}")

def count_rows(filepath):
    ext = Path(filepath).suffix.lower()
    try:
        if ext == ".xlsx":
            return count_rows_xlsx(filepath)
        elif ext in (".xls", ".xlsb"):
            return count_rows_xls(filepath)
    except Exception:
        return 0
    return 0


# ─────────────────────────────────────────────
# Lógica de conversão (Thread)
# ─────────────────────────────────────────────

def converter(files, output_path, progress_cb, status_cb, done_cb, error_cb):
    try:
        status_cb("Calculando total de linhas...")
        total_rows = sum(count_rows(f) for f in files) or 1

        rows_done = 0
        first_file = True

        with open(output_path, "w", newline="", encoding="utf-8-sig") as fout:
            writer = csv.writer(fout, delimiter=";", quoting=csv.QUOTE_MINIMAL)

            for filepath in files:
                fname = Path(filepath).name
                status_cb(f"Processando: {fname}...")

                skip_header = not first_file 
                first_row = True

                for row in get_row_iterator(filepath):
                    if skip_header and first_row:
                        first_row = False
                        rows_done += 1
                        continue
                    first_row = False
                    writer.writerow(row)
                    rows_done += 1

                    if rows_done % 500 == 0:
                        progress_cb(rows_done / total_rows)

                first_file = False

        progress_cb(1.0)
        done_cb()

    except Exception as e:
        error_cb(str(e))


# ─────────────────────────────────────────────
# Interface Gráfica Premium (CustomTkinter)
# ─────────────────────────────────────────────

# Configuração global: Visual sofisticado e minimalista
ctk.set_appearance_mode("Dark")  # Forçamos o Dark Mode para um visual mais pro
ctk.set_default_color_theme("blue")

class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("DataFlow - Conversor Excel para CSV")
        self.geometry("800x700")
        self.minsize(700, 600)
        self.configure(fg_color="#121212") # Fundo super escuro (estilo Spotify/Discord)
        
        # Grid principal responsivo
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(2, weight=1) # A lista de arquivos vai expandir

        self.files = {}
        self._build_ui()

    def _build_ui(self):
        # ── CABEÇALHO (Título e Subtítulo)
        self.header_frame = ctk.CTkFrame(self, fg_color="transparent")
        self.header_frame.grid(row=0, column=0, padx=30, pady=(30, 15), sticky="ew")
        
        ctk.CTkLabel(
            self.header_frame, text="DataFlow", 
            font=ctk.CTkFont(family="Segoe UI", size=32, weight="bold"),
            text_color="#FFFFFF"
        ).pack(anchor="w")
        
        ctk.CTkLabel(
            self.header_frame, text="Converta múltiplas planilhas Excel em um único CSV perfeitamente.", 
            font=ctk.CTkFont(family="Segoe UI", size=14),
            text_color="#888888"
        ).pack(anchor="w")

        # ── ÁREA DE "DROPZONE" (Hero Button)
        # Ao invés de um botão pequeno, uma área gigante que convida ao clique
        self.btn_hero_add = ctk.CTkButton(
            self, text="➕\n\nClique aqui para selecionar os arquivos\n(XLS ou XLSX)",
            font=ctk.CTkFont(family="Segoe UI", size=16, weight="bold"),
            height=120, fg_color="#1E1E1E", hover_color="#2A2A2A",
            border_width=2, border_color="#333333", text_color="#DDDDDD",
            command=self._add_files
        )
        self.btn_hero_add.grid(row=1, column=0, padx=30, pady=(0, 15), sticky="ew")

        # ── ÁREA DE ARQUIVOS (Card Central)
        self.card_files = ctk.CTkFrame(self, fg_color="#1E1E1E", corner_radius=15)
        self.card_files.grid(row=2, column=0, padx=30, pady=10, sticky="nsew")
        self.card_files.grid_columnconfigure(0, weight=1)
        self.card_files.grid_rowconfigure(1, weight=1)

        # Cabeçalho da lista de arquivos
        self.list_header = ctk.CTkFrame(self.card_files, fg_color="transparent")
        self.list_header.grid(row=0, column=0, padx=20, pady=15, sticky="ew")
        
        self.lbl_file_count = ctk.CTkLabel(
            self.list_header, text="Nenhum arquivo selecionado", 
            font=ctk.CTkFont(family="Segoe UI", size=14, weight="bold"),
            text_color="#AAAAAA"
        )
        self.lbl_file_count.pack(side="left")

        self.btn_clear = ctk.CTkButton(
            self.list_header, text="Limpar Lista", width=100, height=28,
            fg_color="#3B1C1C", hover_color="#5C2525", text_color="#FF6666",
            font=ctk.CTkFont(size=12, weight="bold"), command=self._clear_all
        )
        self.btn_clear.pack(side="right")
        self.btn_clear.configure(state="disabled") # Desabilitado até ter arquivos

        # ScrollFrame limpo (sem fundo para mesclar com o card)
        self.scroll_frame = ctk.CTkScrollableFrame(self.card_files, fg_color="transparent")
        self.scroll_frame.grid(row=1, column=0, padx=10, pady=(0, 10), sticky="nsew")

        # ── ÁREA DE CONFIGURAÇÃO (Saída)
        self.card_config = ctk.CTkFrame(self, fg_color="#1E1E1E", corner_radius=15)
        self.card_config.grid(row=3, column=0, padx=30, pady=15, sticky="ew")
        self.card_config.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(
            self.card_config, text="📁 Salvar em:", 
            font=ctk.CTkFont(weight="bold"), text_color="#AAAAAA"
        ).grid(row=0, column=0, padx=(20, 10), pady=20, sticky="w")

        self.out_var = tk.StringVar()
        self.out_entry = ctk.CTkEntry(
            self.card_config, textvariable=self.out_var, 
            placeholder_text="Escolha onde o CSV será salvo...", 
            state="disabled", border_width=0, fg_color="#2A2A2A"
        )
        self.out_entry.grid(row=0, column=1, padx=(0, 10), pady=20, sticky="ew")

        self.btn_out = ctk.CTkButton(
            self.card_config, text="Alterar Destino", 
            command=self._choose_output, width=120,
            fg_color="#2B5B84", hover_color="#3673A5"
        )
        self.btn_out.grid(row=0, column=2, padx=(0, 20), pady=20)

        # ── ÁREA DE AÇÃO (Converter & Progresso)
        self.action_frame = ctk.CTkFrame(self, fg_color="transparent")
        self.action_frame.grid(row=4, column=0, padx=30, pady=(5, 30), sticky="ew")
        self.action_frame.grid_columnconfigure(0, weight=1)

        # Progresso escondido inicialmente
        self.progress_bar = ctk.CTkProgressBar(self.action_frame, height=8, progress_color="#10B981")
        self.progress_bar.grid(row=0, column=0, sticky="ew", pady=(0, 10))
        self.progress_bar.set(0)
        self.progress_bar.grid_remove() # Ocultar até precisar

        self.status_var = ctk.StringVar(value="")
        self.lbl_status = ctk.CTkLabel(
            self.action_frame, textvariable=self.status_var, 
            font=ctk.CTkFont(size=12), text_color="#888888"
        )
        self.lbl_status.grid(row=1, column=0, sticky="w", pady=(0, 10))
        self.lbl_status.grid_remove() # Ocultar até precisar

        # Botão principal ultra destacado
        self.btn_convert = ctk.CTkButton(
            self.action_frame, text="INICIAR PROCESSAMENTO", 
            command=self._start_conversion,
            height=55, font=ctk.CTkFont(size=16, weight="bold"),
            fg_color="#10B981", hover_color="#059669", text_color="#FFFFFF"
        )
        self.btn_convert.grid(row=2, column=0, sticky="ew")

    # ── Lógica UI
    def _update_file_count(self):
        count = len(self.files)
        if count == 0:
            self.lbl_file_count.configure(text="Nenhum arquivo selecionado", text_color="#AAAAAA")
            self.btn_clear.configure(state="disabled")
        else:
            txt = f"{count} arquivo{'s' if count > 1 else ''} na fila"
            self.lbl_file_count.configure(text=txt, text_color="#10B981")
            self.btn_clear.configure(state="normal")

    def _add_files(self):
        paths = filedialog.askopenfilenames(
            title="Selecionar arquivos Excel",
            filetypes=[("Arquivos Excel", "*.xls *.xlsx"), ("Todos", "*.*")]
        )
        for p in paths:
            if p not in self.files:
                self._create_file_card(p)
        self._update_file_count()

    def _create_file_card(self, filepath):
        """Card individual para cada arquivo - super clean"""
        card = ctk.CTkFrame(self.scroll_frame, fg_color="#2A2A2A", corner_radius=8)
        card.pack(fill="x", pady=4, padx=5)
        
        # Ícone de documento (usando emoji para simplificar sem libs extras)
        lbl_icon = ctk.CTkLabel(card, text="📄", font=ctk.CTkFont(size=18))
        lbl_icon.pack(side="left", padx=(15, 5), pady=10)

        lbl_name = ctk.CTkLabel(
            card, text=Path(filepath).name, anchor="w",
            font=ctk.CTkFont(weight="bold", size=13)
        )
        lbl_name.pack(side="left", padx=5, fill="x", expand=True)

        btn_del = ctk.CTkButton(
            card, text="Remover", width=70, height=28, 
            fg_color="transparent", hover_color="#3B1C1C", text_color="#FF6666",
            command=lambda f=filepath: self._remove_file(f)
        )
        btn_del.pack(side="right", padx=10)
        
        self.files[filepath] = card

    def _remove_file(self, filepath):
        if filepath in self.files:
            self.files[filepath].destroy()
            del self.files[filepath]
            self._update_file_count()

    def _clear_all(self):
        for filepath in list(self.files.keys()):
            self._remove_file(filepath)

    def _choose_output(self):
        path = filedialog.asksaveasfilename(
            title="Salvar CSV como",
            defaultextension=".csv",
            filetypes=[("Arquivo CSV", "*.csv")]
        )
        if path:
            self.out_var.set(path)
            self.out_entry.configure(state="normal")
            self.out_entry.delete(0, 'end')
            self.out_entry.insert(0, path)
            self.out_entry.configure(state="disabled")

    # ── Lógica de Execução
    def _start_conversion(self):
        if not self.files:
            messagebox.showwarning("Atenção", "Adicione arquivos antes de continuar.")
            return
        
        out = self.out_var.get().strip()
        if not out:
            messagebox.showwarning("Atenção", "Selecione o destino de saída.")
            return

        # Prepara a UI para o modo de trabalho
        self.btn_convert.configure(state="disabled", text="PROCESSANDO...", fg_color="#059669")
        self.btn_hero_add.configure(state="disabled")
        self.btn_out.configure(state="disabled")
        self.btn_clear.configure(state="disabled")
        
        # Mostra barra e status
        self.progress_bar.grid()
        self.lbl_status.grid()
        self.progress_bar.set(0)

        def prog(pct):
            self.progress_bar.set(pct)

        def status(msg):
            self.status_var.set(msg)

        def done():
            self.status_var.set("✅ Finalizado com sucesso!")
            self.progress_bar.set(1.0)
            self._reset_ui_state()
            messagebox.showinfo("Sucesso", f"Conversão concluída!\n\nSalvo em: {out}")

        def err(msg):
            self.status_var.set("❌ Falha na conversão.")
            self._reset_ui_state()
            messagebox.showerror("Erro", f"Tivemos um problema:\n\n{msg}")

        # Thread para não travar a UI
        t = threading.Thread(
            target=converter,
            args=(list(self.files.keys()), out, prog, status, done, err),
            daemon=True
        )
        t.start()
        
    def _reset_ui_state(self):
        self.btn_convert.configure(state="normal", text="INICIAR PROCESSAMENTO", fg_color="#10B981")
        self.btn_hero_add.configure(state="normal")
        self.btn_out.configure(state="normal")
        self.btn_clear.configure(state="normal")
        # Opcional: Ocultar o status após 3 segundos
        self.after(3000, self.progress_bar.grid_remove)
        self.after(3000, self.lbl_status.grid_remove)


if __name__ == "__main__":
    app = App()
    app.mainloop()